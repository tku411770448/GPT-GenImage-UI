from __future__ import annotations

import json
import os
import re
import shutil
import subprocess
import sys
import uuid
import urllib.request
import urllib.error
import faulthandler
import traceback
from dataclasses import dataclass, asdict, field
from datetime import datetime
from pathlib import Path
from typing import Optional, Callable, Any

from PIL import Image, ImageDraw

from PySide6.QtCore import Qt, QProcess, QUrl, Signal, QPoint, QRect, QSize, QProcessEnvironment, QEvent, QSignalBlocker, QItemSelectionModel, QTimer
from PySide6.QtGui import (
    QDesktopServices, QPixmap, QCursor, QPainter, QColor, QPen, QTextCursor, QIcon,
    QMouseEvent, QIntValidator, QPolygon, QShortcut, QKeySequence,
    QStandardItem, QStandardItemModel
)
from PySide6.QtWidgets import (
    QApplication, QButtonGroup, QCheckBox, QComboBox, QFileDialog, QFrame, QGridLayout, QGroupBox,
    QHBoxLayout, QLabel, QLineEdit, QListWidget, QListWidgetItem, QMainWindow,
    QMessageBox, QPlainTextEdit, QProgressBar, QProgressDialog, QPushButton, QScrollArea, QSizePolicy,
    QSpinBox, QDoubleSpinBox, QStackedWidget, QTextEdit, QVBoxLayout, QWidget, QInputDialog,
    QListView, QMenu, QToolButton, QAbstractItemView, QStyle, QStyleOptionSpinBox,
    QDialog, QDialogButtonBox, QFormLayout
)

SUPPORTED_EXTS = {".jpg", ".jpeg", ".png", ".webp", ".bmp", ".tif", ".tiff"}
APP_MODE = "defect"
APP_MODE_LABEL = "Defect"
APP_MODE_CARD_COLOR = "#10b981"
MODE_OPTIONS = [("defect", "Defect"), ("food", "Food")]
MAX_INITIAL_THUMBNAILS = 120
_THUMB_CACHE: dict[tuple[str, int, int, int], QIcon] = {}
_THUMB_CACHE_ORDER: list[tuple[str, int, int, int]] = []
_THUMB_CACHE_LIMIT = 512


def cached_thumbnail_icon(path: Path, size: QSize) -> QIcon:
    """Return a small cached thumbnail icon without decoding full images repeatedly."""
    try:
        st = path.stat()
        key = (str(path.resolve()), int(st.st_mtime_ns), int(size.width()), int(size.height()))
        icon = _THUMB_CACHE.get(key)
        if icon is not None:
            return icon
        with Image.open(path) as img:
            img = img.convert("RGB")
            img.thumbnail((max(1, size.width()), max(1, size.height())), Image.Resampling.BILINEAR)
            pix = pil_to_pixmap(img)
        icon = QIcon(pix)
        _THUMB_CACHE[key] = icon
        _THUMB_CACHE_ORDER.append(key)
        while len(_THUMB_CACHE_ORDER) > _THUMB_CACHE_LIMIT:
            old = _THUMB_CACHE_ORDER.pop(0)
            _THUMB_CACHE.pop(old, None)
        return icon
    except Exception:
        return QIcon()



class ReliableSpinBox(QSpinBox):
    """QSpinBox variant with a guaranteed clickable up/down hot zone.

    Some platform/style combinations can render the arrow sub-controls but make
    their mouse hit area unreliable after global QSS padding is applied. This
    class keeps normal QSpinBox behavior for text editing and keyboard input,
    but explicitly handles left-clicks in the arrow area so values always step
    by one. It is used by Step 6 output count.
    """

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setSingleStep(1)
        self.setKeyboardTracking(True)
        self._manual_spin_click = False
        try:
            self.setButtonSymbols(QSpinBox.ButtonSymbols.UpDownArrows)
            self.setAccelerated(False)
        except Exception:
            pass

        # Draw-independent overlay buttons.  Some Windows/Qt style-sheet
        # combinations keep the spin-button hit area but hide the arrow glyphs.
        # Real child buttons stay visible above the editor and also provide a
        # reliable click target for both Step 6 and Step 9.
        self._spin_up_btn = QToolButton(self)
        self._spin_down_btn = QToolButton(self)
        for btn, text, tip in (
            (self._spin_up_btn, "▲", "增加數值"),
            (self._spin_down_btn, "▼", "減少數值"),
        ):
            btn.setObjectName("SpinOverlayButton")
            btn.setText(text)
            btn.setToolTip(tip)
            btn.setFocusPolicy(Qt.FocusPolicy.NoFocus)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setAutoRaise(False)
            btn.setStyleSheet(
                "QToolButton#SpinOverlayButton {"
                "border: 0px; border-left: 1px solid #cbd5e1;"
                "background: #f8fafc; color: #0f172a;"
                "font-size: 10px; font-weight: 900; padding: 0px; margin: 0px;"
                "}"
                "QToolButton#SpinOverlayButton:hover { background: #dbeafe; }"
                "QToolButton#SpinOverlayButton:disabled { color: #94a3b8; background: #f1f5f9; }"
            )
            btn.raise_()
        self._spin_up_btn.clicked.connect(lambda checked=False: self._overlay_step(1))
        self._spin_down_btn.clicked.connect(lambda checked=False: self._overlay_step(-1))
        self._layout_overlay_buttons()

    def _overlay_step(self, delta: int) -> None:
        if not self.isEnabled():
            return
        try:
            self.interpretText()
        except Exception:
            pass
        self.stepBy(delta)
        self.update()

    def _layout_overlay_buttons(self) -> None:
        try:
            btn_w = 28
            h = max(18, self.height())
            top_h = max(10, h // 2)
            bot_h = max(10, h - top_h)
            x = max(0, self.width() - btn_w)
            self._spin_up_btn.setGeometry(x, 1, btn_w, max(10, top_h - 1))
            self._spin_down_btn.setGeometry(x, top_h, btn_w, max(10, bot_h - 1))
            self._spin_up_btn.raise_()
            self._spin_down_btn.raise_()
        except Exception:
            pass

    def resizeEvent(self, event):  # noqa: N802
        super().resizeEvent(event)
        self._layout_overlay_buttons()

    def changeEvent(self, event):  # noqa: N802
        super().changeEvent(event)
        try:
            enabled = self.isEnabled()
            self._spin_up_btn.setEnabled(enabled)
            self._spin_down_btn.setEnabled(enabled)
        except Exception:
            pass

    def _event_xy(self, event: QMouseEvent) -> tuple[float, float]:
        try:
            pos = event.position()
            return float(pos.x()), float(pos.y())
        except Exception:
            pos = event.pos()
            return float(pos.x()), float(pos.y())

    def _manual_step_from_event(self, event: QMouseEvent) -> bool:
        """Return True when this click was handled as a spin-button click."""
        try:
            if (not self.isEnabled()) or event.button() != Qt.LeftButton:
                return False
            x, y = self._event_xy(event)

            # First use Qt's own style geometry. This is the most accurate way
            # to detect the up/down sub-controls across Windows/Linux styles.
            try:
                opt = QStyleOptionSpinBox()
                self.initStyleOption(opt)
                pos = event.pos()
                up_rect = self.style().subControlRect(QStyle.ComplexControl.CC_SpinBox, opt, QStyle.SubControl.SC_SpinBoxUp, self)
                down_rect = self.style().subControlRect(QStyle.ComplexControl.CC_SpinBox, opt, QStyle.SubControl.SC_SpinBoxDown, self)
                if up_rect.contains(pos):
                    self.interpretText()
                    self.stepBy(1)
                    return True
                if down_rect.contains(pos):
                    self.interpretText()
                    self.stepBy(-1)
                    return True
            except Exception:
                pass

            # Fallback hot zone: right side of the field. This catches cases
            # where QSS makes the native sub-control rect empty or unreliable.
            arrow_w = max(34, min(52, int(self.height() * 1.25)))
            if x >= self.width() - arrow_w:
                self.interpretText()
                self.stepBy(1 if y < self.height() / 2 else -1)
                return True
        except Exception:
            return False
        return False

    def paintEvent(self, event):  # noqa: N802
        # Let the native/QSS spinbox paint first, then draw explicit arrow
        # glyphs on top. Some Qt styles hide the arrow image after customizing
        # ::up-button / ::down-button, so this guarantees the visual cue remains
        # visible while preserving the normal editor and border painting.
        super().paintEvent(event)
        try:
            painter = QPainter(self)
            painter.setRenderHint(QPainter.RenderHint.Antialiasing, True)
            arrow_w = max(28, min(44, int(self.height() * 1.15)))
            x0 = max(0, self.width() - arrow_w)
            mid_x = x0 + arrow_w // 2
            top_y = max(7, self.height() // 4)
            bot_y = min(self.height() - 7, (self.height() * 3) // 4)
            color = QColor("#0f172a") if self.isEnabled() else QColor("#94a3b8")
            sep = QColor("#cbd5e1") if self.isEnabled() else QColor("#e5e7eb")
            painter.setPen(QPen(sep, 1))
            painter.drawLine(x0, 2, x0, self.height() - 3)
            painter.setPen(Qt.PenStyle.NoPen)
            painter.setBrush(color)
            up = QPolygon([QPoint(mid_x, top_y - 5), QPoint(mid_x - 6, top_y + 3), QPoint(mid_x + 6, top_y + 3)])
            down = QPolygon([QPoint(mid_x, bot_y + 5), QPoint(mid_x - 6, bot_y - 3), QPoint(mid_x + 6, bot_y - 3)])
            painter.drawPolygon(up)
            painter.drawPolygon(down)
            painter.end()
        except Exception:
            pass

    def mousePressEvent(self, event: QMouseEvent):  # noqa: N802
        if self._manual_step_from_event(event):
            self._manual_spin_click = True
            event.accept()
            self.update()
            return
        self._manual_spin_click = False
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event: QMouseEvent):  # noqa: N802
        if self._manual_spin_click:
            self._manual_spin_click = False
            event.accept()
            return
        super().mouseReleaseEvent(event)

    def keyPressEvent(self, event):  # noqa: N802
        # Keep keyboard stepping reliable even when focus is in the embedded editor.
        if self.isEnabled() and event.key() in (Qt.Key_Up, Qt.Key_Down):
            try:
                self.interpretText()
                self.stepBy(1 if event.key() == Qt.Key_Up else -1)
                event.accept()
                return
            except Exception:
                pass
        super().keyPressEvent(event)


def accept_file_drop(event) -> None:
    """Force external file/folder drags to be treated as copy operations.

    Windows Explorer and some Qt file dialogs propose different drag actions.
    Using CopyAction consistently prevents the cursor from staying in a
    forbidden state even though the dropped payload is valid.
    """
    try:
        event.setDropAction(Qt.CopyAction)
    except Exception:
        pass
    try:
        event.acceptProposedAction()
    except Exception:
        event.accept()

MODELS = ["gpt-image-2", "gpt-image-1.5", "gpt-image-1", "gpt-image-1-mini"]
QUALITIES = ["low", "medium", "high", "auto"]
FIXED_SIZES = ["1024x1024", "1024x1536", "1536x1024", "auto"]
GPT2_SIZE_MODES = ["自訂尺寸", "與原圖尺寸相同"]
SAME_AS_ORIGINAL_SIZE = "same_as_original"
GPT2_SIZE_LIMIT_HINT = "GPT-image-2 尺寸限制：寬高需為 16 的倍數、長邊 ≤ 3840 px、長寬比 ≤ 3:1、總像素 655,360 px～8,294,400 px。不符合時會阻止繼續。"
OPENAI_PRICING_URL = "https://openai.com/api/pricing/"
DEFAULT_GPT_IMAGE2_PRICING_PER_1M = {
    "text_input_tokens": 5.00,
    "text_cached_input_tokens": 1.25,
    "image_input_tokens": 8.00,
    "image_cached_input_tokens": 2.00,
    "image_output_tokens": 30.00,
}
STEP_COUNT = 10
VISIBLE_STEPS = [0, 2, 3, 4, 5, 6, 7, 8, 9]
STEP_TO_STACK_INDEX = {step: idx for idx, step in enumerate(VISIBLE_STEPS)}
STEP_DISPLAY_INDEX = {step: idx + 1 for idx, step in enumerate(VISIBLE_STEPS)}
WINDOWS_RESERVED_NAMES = {
    "CON", "PRN", "AUX", "NUL",
    "COM1", "COM2", "COM3", "COM4", "COM5", "COM6", "COM7", "COM8", "COM9",
    "LPT1", "LPT2", "LPT3", "LPT4", "LPT5", "LPT6", "LPT7", "LPT8", "LPT9",
}


def project_root() -> Path:
    return Path(__file__).resolve().parents[1]


def ensure_dir(path: Path) -> Path:
    path.mkdir(parents=True, exist_ok=True)
    return path


def sanitize_name(text: str) -> str:
    text = str(text or "").strip()
    text = re.sub(r"[\x00-\x1f<>:\"/\\|?*]+", "_", text)
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"_+", "_", text).strip("._-")
    if text.split(".", 1)[0].upper() in WINDOWS_RESERVED_NAMES:
        text = f"{text}_project"
    return text or "custom_class"


def make_project_id(name: str) -> str:
    """Project folder id: exactly the sanitized project name, no timestamp suffix."""
    return sanitize_name(name)


def list_images(folder: Path) -> list[Path]:
    if not folder.exists():
        return []
    return sorted([p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in SUPPORTED_EXTS], key=lambda p: p.name.lower())


def copy_image_to(src: Path, dst_dir: Path, stem_hint: str | None = None) -> Path:
    ensure_dir(dst_dir)
    stem = sanitize_name(stem_hint or src.stem)
    suffix = src.suffix.lower() if src.suffix.lower() in SUPPORTED_EXTS else ".png"
    dst = dst_dir / f"{stem}{suffix}"
    i = 1
    while dst.exists():
        dst = dst_dir / f"{stem}_{i:03d}{suffix}"
        i += 1
    shutil.copy2(src, dst)
    return dst


def open_folder(path: Path) -> None:
    ensure_dir(path)
    QDesktopServices.openUrl(QUrl.fromLocalFile(str(path.resolve())))


def pil_to_pixmap(img: Image.Image) -> QPixmap:
    rgba = img.convert("RGBA")
    data = rgba.tobytes("raw", "RGBA")
    from PySide6.QtGui import QImage
    qimg = QImage(data, rgba.width, rgba.height, rgba.width * 4, QImage.Format_RGBA8888).copy()
    return QPixmap.fromImage(qimg)


def elide_middle(text: str, max_len: int = 34) -> str:
    if len(text) <= max_len:
        return text
    keep = max(6, (max_len - 3) // 2)
    return text[:keep] + "..." + text[-keep:]


def elide_middle_stars(text: str, max_len: int = 34) -> str:
    """Elide long display names with *** in the middle for thumbnail captions."""
    if len(text) <= max_len:
        return text
    keep = max(6, (max_len - 3) // 2)
    return text[:keep] + "***" + text[-keep:]


def parse_region_txt(path: Path) -> dict[str, tuple[int, int, int, int]]:
    """Backward-compatible parser for ROI rectangles and first Target Area bbox."""
    data: dict[str, tuple[int, int, int, int]] = {}
    if not path.exists():
        return data
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        m = re.match(r"^(ROI(?:_\d+)?)\s*:\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)", line)
        if m:
            data[m.group(1)] = tuple(int(m.group(i)) for i in range(2, 6))  # type: ignore[assignment]
            continue
        # Old format: target_area: x1,y1,x2,y2
        m = re.match(r"^(target_area(?:_\d+)?)\s*:\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)", line)
        if m:
            key = "target_area" if "target_area" not in data else m.group(1)
            data[key] = tuple(int(m.group(i)) for i in range(2, 6))  # type: ignore[assignment]
            continue
        # New format: target_area_2: rect:x1,y1,x2,y2
        m = re.match(r"^(target_area(?:_\d+)?)\s*:\s*rect\s*:\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)\s*,\s*(\d+)", line)
        if m:
            key = "target_area" if "target_area" not in data else m.group(1)
            data[key] = tuple(int(m.group(i)) for i in range(2, 6))  # type: ignore[assignment]
    return data


def parse_target_areas(path: Path) -> list[dict[str, Any]]:
    """Parse all target areas. Supports rectangles and closed polygons."""
    shapes: list[dict[str, Any]] = []
    if not path.exists():
        return shapes
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        raw = line.strip()
        if not re.match(r"^target_area(?:_\d+)?\s*:", raw):
            continue
        body = raw.split(":", 1)[1].strip()
        if body.startswith("rect:"):
            vals = [int(v.strip()) for v in body[5:].split(",") if v.strip().lstrip("-").isdigit()]
            if len(vals) >= 4:
                shapes.append({"kind": "rect", "rect": tuple(vals[:4])})
        elif body.startswith("polygon:"):
            pts: list[tuple[int, int]] = []
            for pair in body[8:].split(";"):
                nums = [v.strip() for v in pair.split(",")]
                if len(nums) == 2 and nums[0].lstrip("-").isdigit() and nums[1].lstrip("-").isdigit():
                    pts.append((int(nums[0]), int(nums[1])))
            if len(pts) >= 3:
                shapes.append({"kind": "polygon", "points": pts})
        else:
            vals = [int(v.strip()) for v in body.split(",") if v.strip().lstrip("-").isdigit()]
            if len(vals) >= 4:
                shapes.append({"kind": "rect", "rect": tuple(vals[:4])})
    return shapes


def target_area_bbox(shape: dict[str, Any]) -> tuple[int, int, int, int]:
    if shape.get("kind") == "polygon":
        pts = shape.get("points") or []
        xs = [int(x) for x, _ in pts]; ys = [int(y) for _, y in pts]
        return min(xs), min(ys), max(xs), max(ys)
    rect = shape.get("rect") or (0, 0, 0, 0)
    return tuple(int(v) for v in rect[:4])  # type: ignore[return-value]


class PreviewCanvasLabel(QLabel):
    """Pixmap preview label that never lets the pixmap resize the layout.

    QLabel.setPixmap() changes the label size hint to the image size on some Qt
    versions, which makes large previews look like they progressively enlarge the
    page. This widget stores the pixmap internally and paints a scaled copy inside
    the current label rect, so selecting an image updates immediately without
    triggering layout growth.
    """

    def __init__(self, text: str = "尚無預覽") -> None:
        super().__init__(text)
        self._preview_pixmap: Optional[QPixmap] = None
        self._empty_text = text
        self.setAlignment(Qt.AlignCenter)
        self.setMinimumSize(240, 180)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.setObjectName("PreviewLabel")

    def sizeHint(self):  # noqa: N802
        return QSize(640, 360)

    def minimumSizeHint(self):  # noqa: N802
        return QSize(240, 180)

    def set_preview_pixmap(self, pixmap: QPixmap) -> None:
        self._preview_pixmap = pixmap if pixmap and not pixmap.isNull() else None
        self.setText("")
        self.update()

    def clear_preview(self, text: Optional[str] = None) -> None:
        self._preview_pixmap = None
        self._empty_text = text or self._empty_text
        self.setText(self._empty_text)
        self.update()

    def paintEvent(self, event):  # noqa: N802
        if not self._preview_pixmap or self._preview_pixmap.isNull():
            super().paintEvent(event)
            return
        super().paintEvent(event)
        painter = QPainter(self)
        scaled = self._preview_pixmap.scaled(self.size(), Qt.KeepAspectRatio, Qt.FastTransformation)
        x = (self.width() - scaled.width()) // 2
        y = (self.height() - scaled.height()) // 2
        painter.drawPixmap(x, y, scaled)
        painter.end()

class ImagePreview(QWidget):
    paths_dropped = Signal(list)

    def __init__(self, text: str = "尚無預覽", accept_drops: bool = False) -> None:
        super().__init__()
        self.setAcceptDrops(accept_drops)
        self.setMinimumSize(260, 220)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._pixmap: Optional[QPixmap] = None
        self._image_size: tuple[int, int] = (0, 0)
        self._empty_text = text
        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(6)
        self.image_label = PreviewCanvasLabel(text)
        lay.addWidget(self.image_label, 1)
        self.dimension_label = QLabel("圖像尺寸：0*0")
        self.dimension_label.setAlignment(Qt.AlignCenter)
        self.dimension_label.setObjectName("DimensionLabel")
        lay.addWidget(self.dimension_label, 0)

    def clear(self, empty_text: Optional[str] = None) -> None:
        self._pixmap = None
        self._image_size = (0, 0)
        self.image_label.clear_preview(empty_text or self._empty_text)
        self.dimension_label.setText("圖像尺寸：0*0")

    def set_path(self, path: Optional[Path], empty_text: str = "尚無預覽") -> None:
        self._empty_text = empty_text
        if not path or not path.exists():
            self.clear(empty_text)
            return
        try:
            img = Image.open(path).convert("RGB")
            self._image_size = img.size
            self._pixmap = pil_to_pixmap(img)
            self.image_label.set_preview_pixmap(self._pixmap)
            self.dimension_label.setText(f"圖像尺寸：{img.size[0]}*{img.size[1]}")
        except Exception as exc:
            self._pixmap = None
            self._image_size = (0, 0)
            self.image_label.clear_preview(f"預覽失敗：{exc}")
            self.dimension_label.setText("圖像尺寸：0*0")

    def resizeEvent(self, event):  # noqa: N802
        super().resizeEvent(event)
        self._update_scaled()

    def _update_scaled(self) -> None:
        if self._pixmap and not self._pixmap.isNull():
            self.image_label.update()

    def dragEnterEvent(self, event):  # noqa: N802
        if self._paths_from_event(event):
            accept_file_drop(event)
        else:
            event.ignore()

    def dragMoveEvent(self, event):  # noqa: N802
        self.dragEnterEvent(event)

    def dropEvent(self, event):  # noqa: N802
        paths = self._paths_from_event(event)
        if paths:
            self.paths_dropped.emit([str(p) for p in paths])
            accept_file_drop(event)
        else:
            event.ignore()

    @staticmethod
    def _paths_from_event(event) -> list[Path]:
        if not event.mimeData().hasUrls():
            return []
        out: list[Path] = []
        for url in event.mimeData().urls():
            p = Path(url.toLocalFile())
            if p.is_dir() or p.suffix.lower() in SUPPORTED_EXTS:
                out.append(p)
        return out


class ImageDropList(QListWidget):
    paths_dropped = Signal(list)
    delete_requested = Signal()

    def __init__(self) -> None:
        super().__init__()
        self.setAcceptDrops(True)
        self.setSelectionMode(QListWidget.ExtendedSelection)
        self.setMinimumHeight(180)
        self.setToolTip("可直接拖曳圖片檔或圖片資料夾到此處。")

    def dragEnterEvent(self, event):  # noqa: N802
        if ImagePreview._paths_from_event(event):
            accept_file_drop(event)
        else:
            event.ignore()

    def dragMoveEvent(self, event):  # noqa: N802
        self.dragEnterEvent(event)

    def dropEvent(self, event):  # noqa: N802
        paths = ImagePreview._paths_from_event(event)
        if paths:
            self.paths_dropped.emit([str(p) for p in paths])
            accept_file_drop(event)
        else:
            event.ignore()

    def keyPressEvent(self, event):  # noqa: N802
        if event.key() in (Qt.Key_Delete, Qt.Key_Backspace):
            self.delete_requested.emit()
            return
        super().keyPressEvent(event)


class ThumbGrid(QListWidget):
    selected_path_changed = Signal(str)
    selected_paths_changed = Signal(object)

    def __init__(self) -> None:
        super().__init__()
        self.setViewMode(QListWidget.IconMode)
        self.setResizeMode(QListWidget.Adjust)
        self.setMovement(QListWidget.Static)
        self.setIconSize(QPixmap(120, 90).size())
        self.setGridSize(QPixmap(160, 132).size())
        self.setSpacing(8)
        self.setMinimumHeight(180)
        self.itemSelectionChanged.connect(self._emit_selected)

    def load_paths(self, paths: list[Path]) -> None:
        self.clear()
        for p in paths:
            item = QListWidgetItem(elide_middle(p.name, 26))
            item.setData(Qt.UserRole, str(p))
            item.setTextAlignment(Qt.AlignCenter)
            try:
                item.setIcon(cached_thumbnail_icon(p, QSize(120, 90)))
            except Exception:
                pass
            item.setToolTip(p.name)
            self.addItem(item)

    def _emit_selected(self) -> None:
        try:
            items = self.selectedItems()
            paths = [str(item.data(Qt.UserRole)) for item in items if item.data(Qt.UserRole)]
            if paths:
                current = self.currentItem()
                if current and current.data(Qt.UserRole):
                    self.selected_path_changed.emit(str(current.data(Qt.UserRole)))
                else:
                    self.selected_path_changed.emit(paths[0])
            self.selected_paths_changed.emit(paths)
        except Exception:
            traceback.print_exc()


class CenteredThumbGrid(ThumbGrid):
    """Step 3 crop-result thumbnails.

    Earlier versions centered the icons by changing QListWidget viewport margins
    inside resizeEvent(). That can trigger unstable resize/layout recursion in
    Qt item views when pages are rebuilt after button clicks. Keep item text/icon
    centered, but avoid mutating viewport margins during resize.
    """

    def __init__(self) -> None:
        super().__init__()
        self.setUniformItemSizes(True)
        self.setWrapping(True)
        self.setFlow(QListView.Flow.LeftToRight)
        try:
            self.setItemAlignment(Qt.AlignCenter)
        except Exception:
            pass

    def load_paths(self, paths: list[Path]) -> None:
        super().load_paths(paths)
        for i in range(self.count()):
            self.item(i).setTextAlignment(Qt.AlignCenter)

    def wheelEvent(self, event):  # noqa: N802
        # Hovering this grid and scrolling moves the selection (wheel down = next image,
        # wheel up = previous image) without an explicit click, so the middle crop-frame
        # preview and the crop preview box update to the corresponding image.
        if self.count() <= 0:
            super().wheelEvent(event)
            return
        delta = event.angleDelta().y() or event.angleDelta().x()
        if delta == 0:
            super().wheelEvent(event)
            return
        step = 1 if delta < 0 else -1
        current = self.currentRow()
        if current < 0:
            nxt = 0 if step > 0 else self.count() - 1
        else:
            nxt = max(0, min(self.count() - 1, current + step))
        if nxt != self.currentRow():
            self.setCurrentRow(nxt)
        self.scrollToItem(self.item(nxt), QListWidget.PositionAtCenter)
        event.accept()



class VerticalThumbGrid(ThumbGrid):
    """Thumbnail list for Step 4: vertical, scrollable, and wheel-selectable.

    Important: do not use setViewportMargins() from resizeEvent here. Some
    Qt/PySide builds can enter a resize/layout recursion when a QListWidget is
    refreshed while the page is being shown, which may close the process right
    after Step 3 Submit switches to Step 4. Centering is handled by the parent
    layout instead.
    """

    def __init__(self) -> None:
        super().__init__()
        self.setViewMode(QListWidget.IconMode)
        try:
            self.setFlow(QListView.Flow.TopToBottom)
            self.setItemAlignment(Qt.AlignHCenter | Qt.AlignTop)
        except Exception:
            pass
        self.setWrapping(False)
        self.setResizeMode(QListWidget.Adjust)
        self.setUniformItemSizes(True)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setIconSize(QSize(150, 110))
        self.setGridSize(QSize(190, 150))
        self.setSpacing(10)
        self.setMinimumWidth(230)
        self.setMaximumWidth(260)
        self.setMinimumHeight(500)
        self.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Expanding)

    def load_paths(self, paths: list[Path]) -> None:
        self.clear()
        icon_size = self.iconSize()
        for p in paths:
            item = QListWidgetItem(elide_middle(p.name, 30))
            item.setData(Qt.UserRole, str(p))
            item.setTextAlignment(Qt.AlignCenter)
            try:
                item.setIcon(cached_thumbnail_icon(p, icon_size))
            except Exception:
                pass
            item.setToolTip(p.name)
            self.addItem(item)

    def wheelEvent(self, event):  # noqa: N802
        if self.count() <= 0:
            super().wheelEvent(event)
            return
        delta = event.angleDelta().y() or event.angleDelta().x()
        if delta == 0:
            super().wheelEvent(event)
            return
        current = self.currentRow()
        if current < 0:
            current = 0
        step = 1 if delta < 0 else -1
        nxt = max(0, min(self.count() - 1, current + step))
        self.setCurrentRow(nxt)
        self.scrollToItem(self.item(nxt), QListWidget.PositionAtCenter)
        event.accept()


class OutputThumbList(VerticalThumbGrid):
    """Step 9 output thumbnails: narrow, centered, vertical, wheel-selectable."""

    def __init__(self) -> None:
        super().__init__()
        self.setObjectName("OutputThumbList")
        self.setIconSize(QSize(170, 128))
        self.setGridSize(QSize(220, 174))
        self.setSpacing(12)
        self.setMinimumWidth(230)
        self.setMaximumWidth(290)
        self.setMinimumHeight(420)
        self.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Expanding)

    def load_paths(self, paths: list[Path]) -> None:
        self.clear()
        icon_size = self.iconSize()
        for p in paths:
            # Parent run folder is included because output image names often only
            # contain the seed. Full original path remains available in tooltip.
            run_label = p.parent.parent.name if p.parent.parent.name == sanitize_name(p.parent.parent.name) and p.parent.parent != p.parent else p.parent.name
            raw_name = f"{p.parent.parent.name} / {p.parent.name} / {p.name}" if p.parent.parent.name else f"{p.parent.name} / {p.name}"
            item = QListWidgetItem(elide_middle_stars(raw_name, 34))
            item.setData(Qt.UserRole, str(p))
            item.setTextAlignment(Qt.AlignCenter)
            try:
                item.setIcon(cached_thumbnail_icon(p, icon_size))
            except Exception:
                pass
            item.setToolTip(str(p))
            self.addItem(item)


class PromptGroupList(QListWidget):
    group_changed = Signal(str)
    group_selection_changed = Signal(object)

    def __init__(self) -> None:
        super().__init__()
        self.setObjectName("PromptGroupList")
        self.setViewMode(QListWidget.IconMode)
        try:
            self.setFlow(QListView.Flow.LeftToRight)
        except Exception:
            pass
        self.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.setResizeMode(QListWidget.Adjust)
        self.setMovement(QListWidget.Static)
        self.setWrapping(False)
        self.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.setIconSize(QSize(150, 110))
        self.setGridSize(QSize(190, 160))
        self.setMinimumHeight(190)
        self.itemSelectionChanged.connect(self._emit_selected)

    def _make_all_icon(self) -> QIcon:
        pix = QPixmap(150, 110)
        pix.fill(QColor("#b9f6ca"))
        painter = QPainter(pix)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setPen(QColor("#0f766e"))
        painter.drawText(pix.rect(), Qt.AlignCenter, "ALL")
        painter.end()
        return QIcon(pix)

    def load_groups(self, image_paths: list[Path], regions_dir: Path, selected: str = "", selected_stems: Optional[list[str]] = None) -> None:
        self.clear()
        selected_stems = [str(x) for x in (selected_stems or []) if str(x)]
        # image_paths already contains only the region-ready images (decided by
        # all_region_ready_image_paths). Do NOT re-filter by a legacy
        # regions_dir/<stem>.txt file: those .txt files are no longer written
        # (ROI/Target geometry now lives in state.regions + the rendered
        # reference_image), so the old filter made the 引用組別 list always empty.
        valid_paths: list[Path] = [p for p in image_paths if p.exists()]

        if valid_paths:
            all_item = QListWidgetItem("全部圖像\n位置資訊")
            all_item.setData(Qt.UserRole, "__ALL__")
            all_item.setIcon(self._make_all_icon())
            all_item.setToolTip("快速選取所有已完成 ROI / Target Area 的圖像；若超過 16 組，系統會要求縮減。")
            self.addItem(all_item)

        for p in valid_paths:
            item = QListWidgetItem(elide_middle(p.name, 28))
            item.setData(Qt.UserRole, p.stem)
            try:
                pix = pil_to_pixmap(Image.open(p).convert("RGB"))
                item.setIcon(QIcon(pix.scaled(150, 110, Qt.KeepAspectRatio, Qt.FastTransformation)))
            except Exception:
                pass
            item.setToolTip(p.name)
            self.addItem(item)

        # Restore the final generation selection decided at Step 5.  The ALL
        # card is a command-like helper and is not restored as selected.
        found_any = False
        if selected_stems:
            keep = set(selected_stems)
            for i in range(self.count()):
                item = self.item(i)
                role = str(item.data(Qt.UserRole))
                if role in keep:
                    item.setSelected(True)
                    if not found_any:
                        self.setCurrentRow(i)
                        found_any = True
        elif selected:
            for i in range(self.count()):
                if str(self.item(i).data(Qt.UserRole)) == selected:
                    self.setCurrentRow(i)
                    self.item(i).setSelected(True)
                    found_any = True
                    break
        if not found_any and self.count():
            # Leave nothing selected, but set a current row for keyboard/wheel focus.
            self.setCurrentRow(0, QItemSelectionModel.SelectionFlag.NoUpdate)

    def selected_group_ids(self) -> list[str]:
        return [str(item.data(Qt.UserRole)) for item in self.selectedItems() if item.data(Qt.UserRole)]

    def _emit_selected(self) -> None:
        try:
            ids = self.selected_group_ids()
            current = self.currentItem()
            if current and current.data(Qt.UserRole):
                self.group_changed.emit(str(current.data(Qt.UserRole)))
            elif ids:
                self.group_changed.emit(ids[0])
            self.group_selection_changed.emit(ids)
        except Exception:
            traceback.print_exc()

    def mousePressEvent(self, event):  # noqa: N802
        # The 引用組別 list is selection-by-CLICK only. A press that does not land on an
        # item must not clear the multi-selection, and no press may arm a rubber-band
        # marquee (handled in mouseMoveEvent/mouseReleaseEvent). Otherwise an accidental
        # blank click OR a drag would draw a blue selection box, silently rewrite the
        # user's chosen groups and (via group_selection_changed -> mark_dirty) wrongly
        # un-complete the already-finished later steps.
        if event.button() == Qt.LeftButton:
            pos = event.position().toPoint()
            self._press_pos = pos
            self._dragging = False
            self._press_on_blank = self.itemAt(pos) is None
            if self._press_on_blank:
                event.accept()
                return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event):  # noqa: N802
        # Never start/extend a rubber-band selection rectangle while the left button is
        # held: dragging inside the strip would draw a blue marquee box and rewrite the
        # selection. Swallow every such move (flagging a real drag once it passes the
        # drag-distance threshold).
        if (event.buttons() & Qt.LeftButton) and getattr(self, "_press_pos", None) is not None:
            if (event.position().toPoint() - self._press_pos).manhattanLength() >= QApplication.startDragDistance():
                self._dragging = True
            event.accept()
            return
        super().mouseMoveEvent(event)

    def mouseReleaseEvent(self, event):  # noqa: N802
        # A gesture that began on blank space, or that became a drag, is a complete
        # no-op so the chosen groups (and the later steps) stay intact. A plain
        # press+release on an item still resolves as a normal click selection.
        began_blank = getattr(self, "_press_on_blank", False)
        dragged = getattr(self, "_dragging", False)
        self._press_pos = None
        self._dragging = False
        self._press_on_blank = False
        if began_blank or dragged:
            event.accept()
            return
        super().mouseReleaseEvent(event)

    def wheelEvent(self, event):  # noqa: N802
        # Use vertical mouse wheel to scroll the thumbnail strip horizontally.
        bar = self.horizontalScrollBar()
        delta = event.angleDelta().y() or event.angleDelta().x()
        bar.setValue(bar.value() - delta)
        event.accept()


class CheckableRunCombo(QComboBox):
    """Multi-select dropdown used by Step 9 to choose which runs to export.

    Row 0 is the "all runs" option; the remaining rows are individual run names.
    Each row has a checkbox. Checking "all runs" clears and disables the other rows
    (and a forbidden cursor is shown while hovering them); unchecking re-enables
    them. Clicking a row toggles its checkbox without closing the popup, so several
    runs can be selected in one pass."""

    ALL_LABEL = "全部 runs（包含歷次 runs）"
    selection_changed = Signal()

    def __init__(self, parent=None) -> None:
        super().__init__(parent)
        self._model = QStandardItemModel(self)
        self.setModel(self._model)
        self._list_view = QListView()
        self.setView(self._list_view)
        # Mouse tracking makes the popup deliver MouseMove on plain hover (no button
        # pressed), so the forbidden / normal cursor updates the moment the pointer
        # moves over a row instead of only after a click.
        self._list_view.setMouseTracking(True)
        self._list_view.viewport().setMouseTracking(True)
        self.setEditable(True)
        self.lineEdit().setReadOnly(True)
        self.lineEdit().setPlaceholderText("選擇要匯出的 runs")
        self._guard = False
        self._intended_text = ""
        self._model.itemChanged.connect(self._on_item_changed)
        self._list_view.viewport().installEventFilter(self)
        self.lineEdit().installEventFilter(self)

    def set_runs(self, run_names: list) -> None:
        self._guard = True
        try:
            self._model.clear()
            all_item = QStandardItem(self.ALL_LABEL)
            all_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
            all_item.setData(Qt.Unchecked, Qt.CheckStateRole)
            all_item.setData("__ALL__", Qt.UserRole)
            self._model.appendRow(all_item)
            for name in run_names:
                item = QStandardItem(str(name))
                item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
                item.setData(Qt.Unchecked, Qt.CheckStateRole)
                item.setData(str(name), Qt.UserRole)
                self._model.appendRow(item)
        finally:
            self._guard = False
        self._update_text()

    def set_checked(self, run_names: list) -> None:
        wanted = {str(n) for n in run_names}
        self._guard = True
        try:
            for row in range(1, self._model.rowCount()):
                it = self._model.item(row)
                it.setCheckState(Qt.Checked if it.data(Qt.UserRole) in wanted else Qt.Unchecked)
        finally:
            self._guard = False
        self._update_text()
        self.selection_changed.emit()

    def checked_runs(self) -> list:
        all_item = self._model.item(0)
        run_names = [self._model.item(r).data(Qt.UserRole) for r in range(1, self._model.rowCount())]
        if all_item is not None and all_item.checkState() == Qt.Checked:
            return list(run_names)
        return [self._model.item(r).data(Qt.UserRole) for r in range(1, self._model.rowCount())
                if self._model.item(r).checkState() == Qt.Checked]

    def _on_item_changed(self, item) -> None:
        if self._guard:
            return
        self._guard = True
        try:
            if item.row() == 0:
                all_checked = item.checkState() == Qt.Checked
                for row in range(1, self._model.rowCount()):
                    it = self._model.item(row)
                    if all_checked:
                        it.setCheckState(Qt.Unchecked)
                        it.setFlags(Qt.ItemIsUserCheckable)  # drop ItemIsEnabled -> not selectable
                    else:
                        it.setFlags(Qt.ItemIsEnabled | Qt.ItemIsUserCheckable)
                # The rows' enabled state just changed; refresh the cursor for the row
                # under the pointer right now, so it flips immediately even if the
                # pointer has not moved since the toggle.
                if self._list_view.isVisible():
                    self._update_cursor_for_point(self._list_view.viewport().mapFromGlobal(QCursor.pos()))
            self._update_text()
        finally:
            self._guard = False
        self.selection_changed.emit()

    def _update_text(self) -> None:
        all_item = self._model.item(0)
        if all_item is not None and all_item.checkState() == Qt.Checked:
            text = self.ALL_LABEL
        else:
            checked = [self._model.item(r).text() for r in range(1, self._model.rowCount())
                       if self._model.item(r).checkState() == Qt.Checked]
            text = "、".join(checked)
        self._intended_text = text
        le = self.lineEdit()
        if le is not None:
            le.setText(text)
        # An editable QComboBox re-syncs its line edit to currentText() (the current
        # row's display text) when the current row's data changes — e.g. toggling the
        # checkbox of the highlighted run — which would otherwise replace this
        # multi-select text with a single run name. Re-assert it after Qt's sync.
        QTimer.singleShot(0, self._reassert_text)

    def _reassert_text(self) -> None:
        le = self.lineEdit()
        if le is not None and le.text() != self._intended_text:
            le.setText(self._intended_text)

    def _event_point(self, event):
        if hasattr(event, "position"):
            return event.position().toPoint()
        return event.pos()

    def eventFilter(self, obj, event):
        if obj is self.lineEdit() and event.type() == QEvent.MouseButtonPress:
            self.showPopup()
            return True
        if obj is self._list_view.viewport():
            if event.type() == QEvent.MouseButtonRelease:
                idx = self._list_view.indexAt(self._event_point(event))
                if idx.isValid():
                    item = self._model.itemFromIndex(idx)
                    if item is not None and bool(item.flags() & Qt.ItemIsEnabled):
                        item.setCheckState(Qt.Unchecked if item.checkState() == Qt.Checked else Qt.Checked)
                return True  # keep the popup open for multi-select
            if event.type() == QEvent.MouseMove:
                self._update_cursor_for_point(self._event_point(event))
        return super().eventFilter(obj, event)

    def _update_cursor_for_point(self, pos) -> None:
        """Forbidden cursor over disabled rows, pointing-hand otherwise."""
        forbidden = False
        idx = self._list_view.indexAt(pos)
        if idx.isValid():
            item = self._model.itemFromIndex(idx)
            if item is not None and not bool(item.flags() & Qt.ItemIsEnabled):
                forbidden = True
        self._list_view.viewport().setCursor(Qt.ForbiddenCursor if forbidden else Qt.PointingHandCursor)


class ProjectCard(QFrame):
    selected = Signal(str)
    open_requested = Signal(str)
    copy_requested = Signal(str)
    rename_requested = Signal(str)
    delete_requested = Signal(str)

    def __init__(self, project: dict) -> None:
        super().__init__()
        self.project = project
        self.pid = str(project.get("id", ""))
        self.project_mode = str(project.get("mode") or project.get("project_mode") or APP_MODE).lower()
        if self.project_mode not in {"defect", "food"}:
            self.project_mode = APP_MODE
        self.setObjectName("ProjectCard")
        self.setProperty("selected", "false")
        self.setProperty("mode", self.project_mode)
        self.setCursor(QCursor(Qt.PointingHandCursor))
        self.setMinimumSize(280, 205)
        self.setMaximumHeight(260)
        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 12)
        outer.setSpacing(10)
        header = QFrame(); header.setObjectName("ProjectCardHeader"); header.setProperty("mode", self.project_mode)
        h = QHBoxLayout(header); h.setContentsMargins(14, 12, 14, 12)
        title_badge = QLabel(str(project.get("name", "未命名專案") or "未命名專案")); title_badge.setObjectName("ProjectBadge")
        title_badge.setProperty("mode", self.project_mode)
        title_badge.setWordWrap(False)
        title_badge.setMinimumWidth(150)
        title_badge.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        menu_btn = QToolButton(); menu_btn.setText("…"); menu_btn.setPopupMode(QToolButton.InstantPopup); menu_btn.setAutoRaise(True)
        menu = QMenu(menu_btn)
        open_a = menu.addAction("Open")
        rename_a = menu.addAction("Rename")
        copy_a = menu.addAction("Duplicate")
        del_a = menu.addAction("Remove")
        # Use explicit slots instead of anonymous lambdas. QAction.triggered may pass
        # a checked bool in some Qt/PySide builds; explicit slots keep the menu
        # commands stable and make Open / Rename / Duplicate / Remove unambiguous.
        open_a.triggered.connect(self._emit_open_requested)
        rename_a.triggered.connect(self._emit_rename_requested)
        copy_a.triggered.connect(self._emit_copy_requested)
        del_a.triggered.connect(self._emit_delete_requested)
        menu_btn.setMenu(menu)
        h.addWidget(title_badge, 1); h.addWidget(menu_btn)
        outer.addWidget(header)
        created = str(project.get('created_at') or project.get('updated_at') or '-').replace('T',' ')
        updated = str(project.get('updated_at','-')).replace('T',' ')
        mode_text = 'Food' if self.project_mode == 'food' else 'Defect'
        meta = QLabel(
            f"Mode：{mode_text}\n"
            f"Class：{project.get('class_name','-')}\n"
            f"Model：{project.get('model','-')}｜Quality：{project.get('quality','-')}\n"
            f"Create：{created}\n"
            f"Updated：{updated}"
        )
        meta.setObjectName("CardMeta"); meta.setWordWrap(True)
        outer.addWidget(meta); outer.addStretch(1)


    def _emit_open_requested(self, checked: bool = False) -> None:
        self.open_requested.emit(self.pid)

    def _emit_copy_requested(self, checked: bool = False) -> None:
        self.copy_requested.emit(self.pid)

    def _emit_rename_requested(self, checked: bool = False) -> None:
        self.rename_requested.emit(self.pid)

    def _emit_delete_requested(self, checked: bool = False) -> None:
        self.delete_requested.emit(self.pid)

    def set_selected(self, selected: bool) -> None:
        self.setProperty("selected", "true" if selected else "false")
        self.style().unpolish(self); self.style().polish(self)

    def mousePressEvent(self, event):  # noqa: N802
        if event.button() == Qt.LeftButton:
            self.selected.emit(self.pid)
        super().mousePressEvent(event)

    def mouseDoubleClickEvent(self, event):  # noqa: N802
        if event.button() == Qt.LeftButton:
            self.open_requested.emit(self.pid)
        super().mouseDoubleClickEvent(event)


class CropCanvas(QLabel):
    crop_made = Signal(str)

    def __init__(self) -> None:
        super().__init__("請先選擇縮圖，再設定寬高並按確認。")
        self.setObjectName("PreviewLabel")
        self.setAlignment(Qt.AlignCenter)
        self.setMouseTracking(True)
        self.setMinimumSize(520, 420)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.image_path: Optional[Path] = None
        self.image_size = (0, 0)
        self.pix: Optional[QPixmap] = None
        self.crop_size: Optional[tuple[int, int]] = None
        self.mouse_pos: Optional[QPoint] = None
        self.crop_callback: Optional[Callable[[Path, tuple[int, int, int, int]], Optional[Path]]] = None
        self.active_crop_rect: Optional[tuple[int, int, int, int]] = None
        self.edit_enabled = False
        self.moving_active = False
        self._move_offset = QPoint(0, 0)

    def set_image(self, path: Optional[Path], active_rect: Optional[tuple[int, int, int, int]] = None) -> None:
        self.image_path = path if path and path.exists() else None
        self.mouse_pos = None
        self.active_crop_rect = active_rect
        # Existing crop frames are shown for context first; right-click the frame to enable re-editing.
        self.edit_enabled = False
        self.moving_active = False
        if not self.image_path:
            self.pix = None
            self.image_size = (0, 0)
            self.setText("請先選擇縮圖。")
            self.update()
            return
        try:
            img = Image.open(self.image_path).convert("RGB")
            self.image_size = img.size
            self.pix = pil_to_pixmap(img)
            self.setText("")
            if self.active_crop_rect:
                self._clamp_active_crop()
            self.update()
        except Exception as exc:
            self.pix = None
            self.image_size = (0, 0)
            self.setText(f"載入失敗：{exc}")
            self.update()

    def set_active_crop_rect(self, rect: Optional[tuple[int, int, int, int]]) -> None:
        self.active_crop_rect = rect
        self.edit_enabled = False
        self.moving_active = False
        if rect:
            self.crop_size = (max(1, rect[2] - rect[0]), max(1, rect[3] - rect[1]))
            self._clamp_active_crop()
        self.update()

    def set_crop_size(self, w: int, h: int) -> None:
        self.crop_size = (int(w), int(h))
        self.setCursor(Qt.CrossCursor if not self.edit_enabled else Qt.SizeAllCursor)
        if self.active_crop_rect:
            cx = (self.active_crop_rect[0] + self.active_crop_rect[2]) // 2
            cy = (self.active_crop_rect[1] + self.active_crop_rect[3]) // 2
            self.active_crop_rect = self._rect_from_center(cx, cy, int(w), int(h))
        self.update()

    def paintEvent(self, event):  # noqa: N802
        super().paintEvent(event)
        if not self.pix or self.pix.isNull():
            return
        painter = QPainter(self)
        target = self._target_rect()
        painter.drawPixmap(target, self.pix)
        draw_hover = self.crop_size and self.mouse_pos and target.contains(self.mouse_pos) and not self.moving_active
        if draw_hover and not self.edit_enabled:
            hover_label = f"{self.crop_size[0]}×{self.crop_size[1]}" if self.crop_size else ""
            self._draw_crop_rect(painter, self._display_crop_rect(self.mouse_pos), hover_label)
        if self.active_crop_rect:
            disp = self._image_to_display_rect(self.active_crop_rect)
            label = f"{self.active_crop_rect[2]-self.active_crop_rect[0]}×{self.active_crop_rect[3]-self.active_crop_rect[1]}"
            self._draw_crop_rect(painter, disp, label, active=True)
        painter.end()

    def _draw_crop_rect(self, painter: QPainter, rect: QRect, label: str, active: bool = False) -> None:
        painter.setPen(QPen(QColor("#ff1744"), 4 if active else 3))
        painter.setBrush(QColor(255, 23, 68, 35))
        painter.drawRect(rect)
        text = label.strip()
        if text:
            fm = painter.fontMetrics()
            text_w = max(64, fm.horizontalAdvance(text) + 12)
            text_h = fm.height() + 6
            text_x = rect.center().x() - text_w // 2
            text_y = rect.top() - text_h - 4
            if text_y < 0:
                text_y = rect.bottom() + 4
            text_rect = QRect(text_x, text_y, text_w, text_h)
            painter.setPen(QPen(QColor("#b91c1c"), 1))
            painter.drawText(text_rect, Qt.AlignCenter, text)

    def resizeEvent(self, event):  # noqa: N802
        super().resizeEvent(event)
        self.update()

    def mousePressEvent(self, event: QMouseEvent):  # noqa: N802
        if not self.pix:
            return
        pos = event.position().toPoint()
        if event.button() == Qt.RightButton and self._active_display_rect().contains(pos):
            # Right-click activates edit mode without immediately producing a new crop.
            # After activation, left-drag the red frame to recrop a new area.
            self.edit_enabled = True
            self.moving_active = False
            self.setCursor(Qt.SizeAllCursor)
            self.update()
            event.accept()
            return
        if event.button() == Qt.LeftButton and self.edit_enabled and self._active_display_rect().contains(pos):
            self.moving_active = True
            ar = self._active_display_rect()
            self._move_offset = pos - ar.topLeft()
            self.setCursor(Qt.SizeAllCursor)
            self.update()
            event.accept()
            return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event: QMouseEvent):  # noqa: N802
        pos = event.position().toPoint()
        if self.moving_active and self.active_crop_rect and self.crop_size:
            target = self._target_rect()
            disp_w = self._active_display_rect().width()
            disp_h = self._active_display_rect().height()
            x = max(target.left(), min(pos.x() - self._move_offset.x(), target.right() - disp_w + 1))
            y = max(target.top(), min(pos.y() - self._move_offset.y(), target.bottom() - disp_h + 1))
            center = QPoint(x + disp_w // 2, y + disp_h // 2)
            img_rect = self._image_crop_rect(center)
            if img_rect:
                self.active_crop_rect = img_rect
            self.update()
            event.accept()
            return
        self.mouse_pos = pos
        if self.active_crop_rect and self._active_display_rect().contains(pos):
            self.setCursor(Qt.SizeAllCursor if self.edit_enabled else Qt.CrossCursor)
        else:
            self.setCursor(Qt.CrossCursor)
        self.update()

    def leaveEvent(self, event):  # noqa: N802
        if not self.moving_active:
            self.mouse_pos = None
        self.update()
        super().leaveEvent(event)

    def mouseReleaseEvent(self, event: QMouseEvent):  # noqa: N802
        if not self.image_path or not self.crop_size:
            return
        if self.moving_active and self.active_crop_rect:
            self.moving_active = False
            self.setCursor(Qt.SizeAllCursor if self.edit_enabled else Qt.CrossCursor)
            if self.crop_callback:
                out = self.crop_callback(self.image_path, self.active_crop_rect)
                if out:
                    self.crop_made.emit(str(out))
            self.update()
            event.accept()
            return
        if event.button() != Qt.LeftButton:
            return
        pos = event.position().toPoint()
        target = self._target_rect()
        if not target.contains(pos):
            return
        rect = self._image_crop_rect(pos)
        if rect:
            self.active_crop_rect = rect
            self.edit_enabled = True
            if self.crop_callback:
                out = self.crop_callback(self.image_path, rect)
                if out:
                    self.crop_made.emit(str(out))
        self.mouse_pos = None
        self.update()

    def _target_rect(self) -> QRect:
        if not self.pix:
            return QRect(0, 0, 0, 0)
        scaled = self.pix.size().scaled(self.size(), Qt.KeepAspectRatio)
        x = (self.width() - scaled.width()) // 2
        y = (self.height() - scaled.height()) // 2
        return QRect(x, y, scaled.width(), scaled.height())

    def _display_crop_rect(self, pos: QPoint) -> QRect:
        assert self.crop_size is not None
        target = self._target_rect()
        sx = target.width() / max(1, self.image_size[0])
        sy = target.height() / max(1, self.image_size[1])
        dw = max(1, int(self.crop_size[0] * sx))
        dh = max(1, int(self.crop_size[1] * sy))
        x = max(target.left(), min(pos.x() - dw // 2, target.right() - dw + 1))
        y = max(target.top(), min(pos.y() - dh // 2, target.bottom() - dh + 1))
        return QRect(x, y, dw, dh)

    def _image_crop_rect(self, pos: QPoint) -> Optional[tuple[int, int, int, int]]:
        if not self.crop_size:
            return None
        target = self._target_rect()
        rx = (pos.x() - target.left()) / max(1, target.width())
        ry = (pos.y() - target.top()) / max(1, target.height())
        cx = int(rx * self.image_size[0])
        cy = int(ry * self.image_size[1])
        w, h = self.crop_size
        return self._rect_from_center(cx, cy, w, h)

    def _rect_from_center(self, cx: int, cy: int, w: int, h: int) -> Optional[tuple[int, int, int, int]]:
        if w <= 0 or h <= 0 or w > self.image_size[0] or h > self.image_size[1]:
            return None
        x1 = max(0, min(cx - w // 2, self.image_size[0] - w))
        y1 = max(0, min(cy - h // 2, self.image_size[1] - h))
        x2 = x1 + w
        y2 = y1 + h
        if x2 > self.image_size[0] or y2 > self.image_size[1]:
            return None
        return x1, y1, x2, y2

    def _image_to_display_rect(self, rect: tuple[int, int, int, int]) -> QRect:
        target = self._target_rect()
        x1, y1, x2, y2 = rect
        sx = target.width() / max(1, self.image_size[0])
        sy = target.height() / max(1, self.image_size[1])
        return QRect(
            target.left() + int(x1 * sx),
            target.top() + int(y1 * sy),
            max(1, int((x2 - x1) * sx)),
            max(1, int((y2 - y1) * sy)),
        )

    def _active_display_rect(self) -> QRect:
        if not self.active_crop_rect:
            return QRect(0, 0, 0, 0)
        return self._image_to_display_rect(self.active_crop_rect)

    def _clamp_active_crop(self) -> None:
        if not self.active_crop_rect or not self.image_size:
            return
        x1, y1, x2, y2 = self.active_crop_rect
        w = max(1, x2 - x1)
        h = max(1, y2 - y1)
        rect = self._rect_from_center((x1 + x2) // 2, (y1 + y2) // 2, min(w, self.image_size[0]), min(h, self.image_size[1]))
        self.active_crop_rect = rect


class RoiTargetCanvas(QLabel):
    region_changed = Signal()
    canvas_error = Signal(str)

    def __init__(self) -> None:
        super().__init__("請選擇裁切完成的圖像。")
        self.setObjectName("PreviewLabel")
        self.setAlignment(Qt.AlignCenter)
        self.setMouseTracking(True)
        self.setMinimumSize(920, 640)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self.setCursor(Qt.CrossCursor)
        self.image_path: Optional[Path] = None
        self.image_size = (0, 0)
        self.pix: Optional[QPixmap] = None
        self.mode = "roi"
        self.drag_start: Optional[QPoint] = None
        self.drag_current: Optional[QPoint] = None
        self.rois: list[tuple[int, int, int, int]] = []
        self.selected_roi_indices: set[int] = set()
        self.target_areas: list[dict[str, Any]] = []
        self.selected_target_indices: set[int] = set()
        self.poly_points: list[QPoint] = []
        self.zoom_factor = 1.0
        self.pan_offset = QPoint(0, 0)
        self.panning = False
        self.pan_start: Optional[QPoint] = None
        self.pan_origin = QPoint(0, 0)
        self.hover_pos: Optional[QPoint] = None
        # Resize support for selected ROI / rectangular Target Area boxes.
        # resize_active is a small state dict: {kind, index, handle, original}.
        self.resize_active: Optional[dict[str, Any]] = None
        self.handle_size = 10
        self.min_region_size = 4
        # Fixed ROI/Target box colours: ROI red, Target Area blue.
        self.roi_color = QColor("#ff1744")
        self.target_color = QColor("#00c2ff")

    @property
    def roi(self) -> Optional[tuple[int, int, int, int]]:
        return self.rois[0] if self.rois else None

    @property
    def target_area(self) -> Optional[tuple[int, int, int, int]]:
        return target_area_bbox(self.target_areas[0]) if self.target_areas else None

    def set_image(self, path: Optional[Path]) -> None:
        self.image_path = path if path and path.exists() else None
        self.rois = []
        self.selected_roi_indices.clear()
        self.target_areas = []
        self.selected_target_indices.clear()
        self.poly_points = []
        self.drag_start = None
        self.drag_current = None
        self.hover_pos = None
        self.resize_active = None
        self.zoom_factor = 1.0
        self.pan_offset = QPoint(0, 0)
        self.panning = False
        if not self.image_path:
            self.pix = None
            self.image_size = (0, 0)
            self.setText("請選擇裁切完成的圖像。")
            self.update()
            return
        try:
            img = Image.open(self.image_path).convert("RGB")
            self.image_size = img.size
            self.pix = pil_to_pixmap(img)
            self.setText("")
            reg = parse_region_txt(self.region_txt_path())
            self.rois = []
            def roi_sort_key(x: str):
                return (0 if x == "ROI" else 1, int(x.split("_")[-1]) if "_" in x and x.split("_")[-1].isdigit() else 1)
            for key in sorted([k for k in reg.keys() if k.startswith("ROI")], key=roi_sort_key):
                self.rois.append(reg[key])
            self.target_areas = parse_target_areas(self.region_txt_path())
            self.update()
        except Exception as exc:
            self.pix = None
            self.image_size = (0, 0)
            self.setText(f"載入失敗：{exc}")
            self.update()

    def region_txt_path(self) -> Path:
        if not self.image_path:
            return Path()
        return self.image_path.parent.parent / "regions" / f"{self.image_path.stem}.txt"

    def set_mode(self, mode: str) -> None:
        if self.mode == "target_poly" and self.poly_points and mode != "target_poly":
            self.canvas_error.emit("Target Area 直線繪製尚未封閉，已取消本次未完成線段。請重新繪製並點回第一個點封閉。")
            self.poly_points = []
        self.mode = mode
        self.drag_start = None
        self.drag_current = None
        self.resize_active = None
        if mode in {"select_roi", "select_target"}:
            self.setCursor(Qt.ArrowCursor)
        else:
            self.setCursor(Qt.CrossCursor)
        self.update()

    def clear_rois(self) -> None:
        self.rois = []
        self.selected_roi_indices.clear()
        self.region_changed.emit()
        self.update()

    def clear_target_areas(self) -> bool:
        if not self.target_areas:
            return False
        self.target_areas = []
        self.selected_target_indices.clear()
        self.poly_points = []
        if self.mode.startswith("target"):
            self.drag_start = None
            self.drag_current = None
        self.region_changed.emit()
        self.update()
        return True

    def clear_target_area(self) -> bool:
        return self.clear_target_areas()

    def delete_selected_roi(self) -> bool:
        valid_indices = sorted([idx for idx in self.selected_roi_indices if 0 <= idx < len(self.rois)], reverse=True)
        if not valid_indices:
            self.selected_roi_indices.clear()
            return False
        for idx in valid_indices:
            del self.rois[idx]
        self.selected_roi_indices.clear()
        self.region_changed.emit()
        self.update()
        return True

    def delete_selected_target_area(self) -> bool:
        valid_indices = sorted([idx for idx in self.selected_target_indices if 0 <= idx < len(self.target_areas)], reverse=True)
        if not valid_indices:
            self.selected_target_indices.clear()
            return False
        for idx in valid_indices:
            del self.target_areas[idx]
        self.selected_target_indices.clear()
        self.region_changed.emit()
        self.update()
        return True

    def paintEvent(self, event):  # noqa: N802
        super().paintEvent(event)
        if not self.pix or self.pix.isNull():
            return
        painter = QPainter(self)
        target = self._target_rect()
        painter.drawPixmap(target, self.pix)
        for idx, roi in enumerate(self.rois, start=1):
            self._draw_rect_region(painter, roi, self.roi_color, f"ROI {idx}", selected=((idx - 1) in self.selected_roi_indices))
        for idx, shape in enumerate(self.target_areas, start=1):
            label = "Target Area" if len(self.target_areas) == 1 else f"Target Area {idx}"
            self._draw_target_shape(painter, shape, label, selected=((idx - 1) in self.selected_target_indices))
        self._draw_resize_handles(painter)
        if self.drag_start and self.drag_current:
            rect = QRect(self.drag_start, self.drag_current).normalized()
            if self.mode == "roi":
                color = self.roi_color
            elif self.mode == "target_rect":
                color = self.target_color
            else:
                color = QColor("#fbbf24")
            painter.setPen(QPen(color, 3, Qt.PenStyle.DashLine if self.mode in {"select_roi", "select_target"} else Qt.PenStyle.SolidLine))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            painter.drawRect(rect)
        if self.mode == "target_poly" and self.poly_points:
            painter.setPen(QPen(self.target_color, 3))
            painter.setBrush(Qt.BrushStyle.NoBrush)
            pts = list(self.poly_points)
            if self.drag_current:
                pts.append(self.drag_current)
            if len(pts) >= 2:
                painter.drawPolyline(QPolygon(pts))
            for i, pt in enumerate(self.poly_points):
                painter.setBrush(self.target_color)
                painter.drawEllipse(pt, 4, 4)
                if i == 0:
                    painter.setPen(QPen(QColor("#ffffff"), 2))
                    painter.drawEllipse(pt, 8, 8)
                    painter.setPen(QPen(self.target_color, 3))
        self._draw_cursor_mode_label(painter)
        painter.end()

    def _draw_rect_region(self, painter: QPainter, region: tuple[int, int, int, int], color: QColor, label: str, selected: bool = False) -> None:
        disp = self._image_to_display_rect(region)
        painter.setPen(QPen(QColor("#fbbf24") if selected else color, 5 if selected else 3))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        painter.drawRect(disp)
        self._draw_label_above(painter, disp, label + ("  已選取" if selected else ""))

    def _draw_target_shape(self, painter: QPainter, shape: dict[str, Any], label: str, selected: bool = False) -> None:
        color = self.target_color
        pen_color = QColor("#fbbf24") if selected else color
        painter.setPen(QPen(pen_color, 5 if selected else 3))
        painter.setBrush(Qt.BrushStyle.NoBrush)
        if shape.get("kind") == "polygon":
            pts = [self._image_to_display_point(x, y) for x, y in shape.get("points", [])]
            if len(pts) >= 3:
                poly = QPolygon(pts)
                painter.drawPolygon(poly)
                bounds = poly.boundingRect()
                self._draw_label_above(painter, bounds, label + ("  已選取" if selected else ""))
        else:
            rect = target_area_bbox(shape)
            self._draw_rect_region(painter, rect, color, label, selected)

    def _draw_label_above(self, painter: QPainter, disp: QRect, text: str) -> None:
        fm = painter.fontMetrics()
        text_w = fm.horizontalAdvance(text) + 10
        text_h = fm.height() + 6
        text_x = disp.center().x() - text_w // 2
        text_y = disp.top() - text_h - 4
        target_rect = self._target_rect()
        if text_y < target_rect.top():
            text_y = target_rect.top() + 2
        text_rect = QRect(text_x, text_y, text_w, text_h)
        painter.fillRect(text_rect, QColor(255, 255, 255, 220))
        painter.setPen(QPen(QColor("#111827"), 1))
        painter.drawText(text_rect, Qt.AlignCenter, text)

    def _handle_rects(self, disp: QRect) -> dict[str, QRect]:
        """Return visible resize handles for a displayed rectangle.

        Handles support four-corner resizing, left/right side-midpoint
        horizontal resizing, and top/bottom side-midpoint vertical resizing.
        """
        hs = max(8, int(self.handle_size))
        half = hs // 2
        points = {
            "nw": disp.topLeft(),
            "n": QPoint(disp.center().x(), disp.top()),
            "ne": disp.topRight(),
            "w": QPoint(disp.left(), disp.center().y()),
            "e": QPoint(disp.right(), disp.center().y()),
            "sw": disp.bottomLeft(),
            "s": QPoint(disp.center().x(), disp.bottom()),
            "se": disp.bottomRight(),
        }
        return {name: QRect(pt.x() - half, pt.y() - half, hs, hs) for name, pt in points.items()}

    def _polygon_handle_rects(self, shape: dict[str, Any]) -> dict[str, QRect]:
        pts = [self._image_to_display_point(x, y) for x, y in shape.get("points", [])]
        if len(pts) < 3:
            return {}
        hs = max(8, int(self.handle_size))
        half = hs // 2
        handles: dict[str, QRect] = {}
        for idx, pt in enumerate(pts):
            handles[f"poly_vertex_{idx}"] = QRect(pt.x() - half, pt.y() - half, hs, hs)
        for idx, pt in enumerate(pts):
            nxt = pts[(idx + 1) % len(pts)]
            mid = QPoint((pt.x() + nxt.x()) // 2, (pt.y() + nxt.y()) // 2)
            handles[f"poly_edge_{idx}"] = QRect(mid.x() - half, mid.y() - half, hs, hs)
        return handles

    def _draw_resize_handles(self, painter: QPainter) -> None:
        """Draw resize handles on selected regions."""
        if self.mode not in {"select_roi", "select_target"}:
            return
        painter.setBrush(QColor(255, 255, 255, 245))
        if self.mode == "select_roi":
            painter.setPen(QPen(QColor("#fbbf24"), 2))
            for idx in sorted(self.selected_roi_indices):
                if 0 <= idx < len(self.rois):
                    for rect in self._handle_rects(self._image_to_display_rect(self.rois[idx])).values():
                        painter.drawRect(rect)
        elif self.mode == "select_target":
            painter.setPen(QPen(QColor("#fbbf24"), 2))
            for idx in sorted(self.selected_target_indices):
                if 0 <= idx < len(self.target_areas):
                    shape = self.target_areas[idx]
                    if shape.get("kind") == "polygon":
                        handles = self._polygon_handle_rects(shape)
                    else:
                        handles = self._handle_rects(self._image_to_display_rect(target_area_bbox(shape)))
                    for rect in handles.values():
                        painter.drawRect(rect)

    def _cursor_for_handle(self, handle: str) -> QCursor:
        if handle.startswith("poly_"):
            return QCursor(Qt.SizeAllCursor)
        if handle in {"nw", "se"}:
            return QCursor(Qt.SizeFDiagCursor)
        if handle in {"ne", "sw"}:
            return QCursor(Qt.SizeBDiagCursor)
        if handle in {"w", "e"}:
            return QCursor(Qt.SizeHorCursor)
        if handle in {"n", "s"}:
            return QCursor(Qt.SizeVerCursor)
        return QCursor(Qt.ArrowCursor)

    def _hit_rect_resize_handle(self, disp: QRect, point: QPoint) -> Optional[str]:
        for handle, hrect in self._handle_rects(disp).items():
            if hrect.contains(point):
                return handle

        tol = max(6, int(self.handle_size // 2) + 2)
        left = disp.left()
        right = disp.right()
        top = disp.top()
        bottom = disp.bottom()
        within_x = left - tol <= point.x() <= right + tol
        within_y = top - tol <= point.y() <= bottom + tol
        candidates: list[tuple[int, str]] = []
        if within_x and abs(point.y() - top) <= tol:
            candidates.append((abs(point.y() - top), "n"))
        if within_x and abs(point.y() - bottom) <= tol:
            candidates.append((abs(point.y() - bottom), "s"))
        if within_y and abs(point.x() - left) <= tol:
            candidates.append((abs(point.x() - left), "w"))
        if within_y and abs(point.x() - right) <= tol:
            candidates.append((abs(point.x() - right), "e"))
        if not candidates:
            return None
        return min(candidates, key=lambda item: item[0])[1]

    def _distance_to_segment_sq(self, point: QPoint, start: QPoint, end: QPoint) -> float:
        dx = end.x() - start.x()
        dy = end.y() - start.y()
        if dx == 0 and dy == 0:
            px = point.x() - start.x()
            py = point.y() - start.y()
            return float(px * px + py * py)
        t = ((point.x() - start.x()) * dx + (point.y() - start.y()) * dy) / float(dx * dx + dy * dy)
        t = max(0.0, min(1.0, t))
        proj_x = start.x() + t * dx
        proj_y = start.y() + t * dy
        px = point.x() - proj_x
        py = point.y() - proj_y
        return px * px + py * py

    def _hit_polygon_resize_handle(self, shape: dict[str, Any], point: QPoint) -> Optional[str]:
        for handle, hrect in self._polygon_handle_rects(shape).items():
            if hrect.contains(point):
                return handle

        pts = [self._image_to_display_point(x, y) for x, y in shape.get("points", [])]
        if len(pts) < 3:
            return None
        tol = max(6, int(self.handle_size // 2) + 2)
        candidates: list[tuple[float, str]] = []
        for idx, start in enumerate(pts):
            end = pts[(idx + 1) % len(pts)]
            dist_sq = self._distance_to_segment_sq(point, start, end)
            if dist_sq <= float(tol * tol):
                candidates.append((dist_sq, f"poly_edge_{idx}"))
        if not candidates:
            return None
        return min(candidates, key=lambda item: item[0])[1]

    def _hit_resize_handle(self, point: QPoint) -> Optional[tuple[str, int, str]]:
        """Return (kind, index, handle) when pointer is on a resize handle."""
        if self.mode == "select_roi":
            for idx in sorted(self.selected_roi_indices, reverse=True):
                if 0 <= idx < len(self.rois):
                    disp = self._image_to_display_rect(self.rois[idx])
                    handle = self._hit_rect_resize_handle(disp, point)
                    if handle is not None:
                        return "roi", idx, handle
        if self.mode == "select_target":
            for idx in sorted(self.selected_target_indices, reverse=True):
                if 0 <= idx < len(self.target_areas):
                    shape = self.target_areas[idx]
                    if shape.get("kind") == "polygon":
                        handle = self._hit_polygon_resize_handle(shape, point)
                        if handle is not None:
                            return "target_polygon", idx, handle
                    else:
                        disp = self._image_to_display_rect(target_area_bbox(shape))
                        handle = self._hit_rect_resize_handle(disp, point)
                        if handle is not None:
                            return "target", idx, handle
        return None

    def _resize_rect_from_handle(self, original: tuple[int, int, int, int], handle: str, point: QPoint) -> tuple[int, int, int, int]:
        px, py = self._display_to_image_point(point)
        x1, y1, x2, y2 = [int(v) for v in original]
        min_size = max(1, int(self.min_region_size))
        img_w, img_h = self.image_size
        if "w" in handle:
            x1 = max(0, min(px, x2 - min_size))
        if "e" in handle:
            x2 = min(img_w, max(px, x1 + min_size))
        if "n" in handle:
            y1 = max(0, min(py, y2 - min_size))
        if "s" in handle:
            y2 = min(img_h, max(py, y1 + min_size))
        return int(x1), int(y1), int(x2), int(y2)

    def _apply_resize_rect(self, kind: str, index: int, rect: tuple[int, int, int, int]) -> None:
        if kind == "roi" and 0 <= index < len(self.rois):
            self.rois[index] = rect
        elif kind == "target" and 0 <= index < len(self.target_areas):
            if self.target_areas[index].get("kind") != "polygon":
                self.target_areas[index] = {"kind": "rect", "rect": rect}

    def _apply_resize_polygon(self, index: int, handle: str, original_points: list[tuple[int, int]], original_point: tuple[int, int], point: QPoint) -> None:
        if not (0 <= index < len(self.target_areas)) or len(original_points) < 3:
            return
        current = self._display_to_image_point(point)
        img_w, img_h = self.image_size
        points = [(int(x), int(y)) for x, y in original_points]
        if handle.startswith("poly_vertex_"):
            vertex_idx = int(handle.rsplit("_", 1)[1])
            if 0 <= vertex_idx < len(points):
                points[vertex_idx] = (max(0, min(img_w, current[0])), max(0, min(img_h, current[1])))
        elif handle.startswith("poly_edge_"):
            edge_idx = int(handle.rsplit("_", 1)[1])
            next_idx = (edge_idx + 1) % len(points)
            if 0 <= edge_idx < len(points):
                endpoints = [points[edge_idx], points[next_idx]]
                dx = current[0] - original_point[0]
                dy = current[1] - original_point[1]
                dx = max(max(-x for x, _ in endpoints), min(dx, min(img_w - x for x, _ in endpoints)))
                dy = max(max(-y for _, y in endpoints), min(dy, min(img_h - y for _, y in endpoints)))
                points[edge_idx] = (points[edge_idx][0] + dx, points[edge_idx][1] + dy)
                points[next_idx] = (points[next_idx][0] + dx, points[next_idx][1] + dy)
        if len(set(points)) >= 3:
            self.target_areas[index] = {"kind": "polygon", "points": points}

    def _cursor_mode_label(self) -> str:
        if self.mode == "roi":
            return "ROI"
        if self.mode in {"target_rect", "target_poly"}:
            return "Target Area"
        return ""

    def _draw_cursor_mode_label(self, painter: QPainter) -> None:
        label = self._cursor_mode_label()
        if not label or self.hover_pos is None or not self._target_rect().contains(self.hover_pos):
            return
        fm = painter.fontMetrics()
        text_w = fm.horizontalAdvance(label) + 12
        text_h = fm.height() + 6
        x = self.hover_pos.x() + 14
        y = self.hover_pos.y() + 14
        if x + text_w > self.width() - 4:
            x = self.hover_pos.x() - text_w - 14
        if y + text_h > self.height() - 4:
            y = self.hover_pos.y() - text_h - 14
        rect = QRect(x, y, text_w, text_h)
        painter.fillRect(rect, QColor(255, 255, 255, 230))
        painter.setPen(QPen(QColor("#111827"), 1))
        painter.drawText(rect, Qt.AlignCenter, label)

    def mousePressEvent(self, event: QMouseEvent):  # noqa: N802
        if not self.pix:
            return
        p = event.position().toPoint()
        if event.button() == Qt.MiddleButton:
            self.panning = True; self.pan_start = p; self.pan_origin = QPoint(self.pan_offset); self.setCursor(Qt.ClosedHandCursor); event.accept(); return
        if event.button() == Qt.LeftButton and event.modifiers() & Qt.AltModifier:
            self.panning = True; self.pan_start = p; self.pan_origin = QPoint(self.pan_offset); self.setCursor(Qt.ClosedHandCursor); event.accept(); return
        if event.button() == Qt.RightButton and self.mode == "target_poly" and self.poly_points:
            self.canvas_error.emit("Target Area 直線繪製尚未封閉。請點回第一個點形成封閉區域，或重新繪製。")
            event.accept(); return
        if event.button() == Qt.LeftButton and self.mode in {"select_roi", "select_target"}:
            resize_hit = self._hit_resize_handle(p)
            if resize_hit is not None:
                kind, idx, handle = resize_hit
                if kind == "target_polygon":
                    original = [(int(x), int(y)) for x, y in self.target_areas[idx].get("points", [])]
                    original_point = self._display_to_image_point(p)
                    self.resize_active = {"kind": kind, "index": idx, "handle": handle, "original": original, "original_point": original_point}
                else:
                    original = self.rois[idx] if kind == "roi" else target_area_bbox(self.target_areas[idx])
                    self.resize_active = {"kind": kind, "index": idx, "handle": handle, "original": original}
                self.setCursor(self._cursor_for_handle(handle))
                event.accept(); return
        if event.button() == Qt.LeftButton and self.mode == "select_roi":
            hit_idx = self._hit_roi_index(p)
            if hit_idx is not None:
                if hit_idx in self.selected_roi_indices: self.selected_roi_indices.remove(hit_idx)
                else: self.selected_roi_indices.add(hit_idx)
            elif self._target_rect().contains(p):
                if not (event.modifiers() & Qt.ControlModifier): self.selected_roi_indices.clear()
                self.drag_start = p; self.drag_current = p
            else:
                if not (event.modifiers() & Qt.ControlModifier): self.selected_roi_indices.clear()
            self.update(); event.accept(); return
        if event.button() == Qt.LeftButton and self.mode == "select_target":
            hit_idx = self._hit_target_index(p)
            if hit_idx is not None:
                if hit_idx in self.selected_target_indices: self.selected_target_indices.remove(hit_idx)
                else: self.selected_target_indices.add(hit_idx)
            elif self._target_rect().contains(p):
                if not (event.modifiers() & Qt.ControlModifier): self.selected_target_indices.clear()
                self.drag_start = p; self.drag_current = p
            else:
                if not (event.modifiers() & Qt.ControlModifier): self.selected_target_indices.clear()
            self.update(); event.accept(); return
        if event.button() == Qt.LeftButton and self.mode == "target_poly" and self._target_rect().contains(p):
            if len(self.poly_points) >= 3 and self._point_distance(p, self.poly_points[0]) <= 14:
                self._finish_polygon_target()
            else:
                self.poly_points.append(p)
                self.drag_current = p
                self.update()
            event.accept(); return
        if event.button() == Qt.LeftButton and self._target_rect().contains(p) and self.mode in {"roi", "target_rect"}:
            self.drag_start = p; self.drag_current = p; self.update(); event.accept(); return
        super().mousePressEvent(event)

    def mouseMoveEvent(self, event: QMouseEvent):  # noqa: N802
        p = event.position().toPoint()
        self.hover_pos = p
        if self.panning and self.pan_start is not None:
            self.pan_offset = self.pan_origin + (p - self.pan_start); self.update(); event.accept(); return
        if self.resize_active is not None:
            info = self.resize_active
            if info["kind"] == "target_polygon":
                self._apply_resize_polygon(info["index"], info["handle"], info["original"], info["original_point"], p)
            else:
                rect = self._resize_rect_from_handle(info["original"], info["handle"], p)
                self._apply_resize_rect(info["kind"], info["index"], rect)
            self.setCursor(self._cursor_for_handle(info["handle"]))
            self.update(); event.accept(); return
        if self.mode == "target_poly" and self.poly_points:
            self.drag_current = p; self.update(); event.accept(); return
        if self.drag_start:
            self.drag_current = p; self.update(); event.accept(); return
        resize_hit = self._hit_resize_handle(p) if self.mode in {"select_roi", "select_target"} else None
        if resize_hit is not None:
            self.setCursor(self._cursor_for_handle(resize_hit[2]))
        elif self.mode == "select_roi":
            self.setCursor(Qt.PointingHandCursor if self._hit_roi_index(p) is not None else Qt.ArrowCursor)
        elif self.mode == "select_target":
            self.setCursor(Qt.PointingHandCursor if self._hit_target_index(p) is not None else Qt.ArrowCursor)
        elif self.mode in {"roi", "target_rect", "target_poly"}:
            self.setCursor(Qt.CrossCursor)
        self.update()

    def leaveEvent(self, event):  # noqa: N802
        self.hover_pos = None
        self.update()
        super().leaveEvent(event)

    def mouseReleaseEvent(self, event: QMouseEvent):  # noqa: N802
        if event.button() in {Qt.MiddleButton, Qt.LeftButton} and self.panning:
            self.panning = False; self.pan_start = None; self.setCursor(Qt.ArrowCursor if self.mode in {"select_roi", "select_target"} else Qt.CrossCursor); event.accept(); return
        if event.button() == Qt.LeftButton and self.resize_active is not None:
            self.resize_active = None
            self.region_changed.emit()
            self.update(); event.accept(); return
        if event.button() == Qt.LeftButton and self.drag_start and self.drag_current:
            r = QRect(self.drag_start, self.drag_current).normalized()
            if self.mode == "select_roi":
                hits = {idx for idx, roi in enumerate(self.rois) if self._image_to_display_rect(roi).intersects(r)}
                if hits: self.selected_roi_indices.update(hits)
            elif self.mode == "select_target":
                hits = {idx for idx, shape in enumerate(self.target_areas) if self._target_shape_display_bbox(shape).intersects(r)}
                if hits: self.selected_target_indices.update(hits)
            else:
                img_rect = self._display_to_image_rect(r)
                if img_rect:
                    if self.mode == "roi":
                        self.rois.append(img_rect); self.selected_roi_indices = {len(self.rois) - 1}
                    elif self.mode == "target_rect":
                        self.target_areas.append({"kind": "rect", "rect": img_rect}); self.selected_target_indices = {len(self.target_areas) - 1}
                    self.region_changed.emit()
            self.drag_start = None; self.drag_current = None; self.update(); event.accept(); return
        super().mouseReleaseEvent(event)

    def wheelEvent(self, event):  # noqa: N802
        if event.modifiers() & Qt.ControlModifier:
            old_rect = self._target_rect(); cursor = event.position().toPoint(); delta = event.angleDelta().y(); old_zoom = self.zoom_factor
            if delta > 0: self.zoom_factor = min(8.0, self.zoom_factor * 1.18)
            elif delta < 0: self.zoom_factor = max(0.35, self.zoom_factor / 1.18)
            if self.zoom_factor != old_zoom and old_rect.width() > 0 and old_rect.height() > 0:
                rx = (cursor.x() - old_rect.left()) / max(1, old_rect.width()); ry = (cursor.y() - old_rect.top()) / max(1, old_rect.height())
                new_rect = self._target_rect()
                self.pan_offset += QPoint(int(cursor.x() - (new_rect.left() + rx * new_rect.width())), int(cursor.y() - (new_rect.top() + ry * new_rect.height())))
            self.update(); event.accept(); return
        super().wheelEvent(event)

    def resizeEvent(self, event):  # noqa: N802
        super().resizeEvent(event); self.update()

    def _target_rect(self) -> QRect:
        if not self.pix: return QRect(0, 0, 0, 0)
        base = self.pix.size().scaled(self.size(), Qt.KeepAspectRatio)
        w = int(base.width() * self.zoom_factor); h = int(base.height() * self.zoom_factor)
        x = (self.width() - w) // 2 + self.pan_offset.x(); y = (self.height() - h) // 2 + self.pan_offset.y()
        return QRect(x, y, w, h)

    def _image_to_display_point(self, x: int, y: int) -> QPoint:
        target = self._target_rect()
        sx = target.width() / max(1, self.image_size[0]); sy = target.height() / max(1, self.image_size[1])
        return QPoint(target.left() + int(x * sx), target.top() + int(y * sy))

    def _display_to_image_point(self, p: QPoint) -> tuple[int, int]:
        target = self._target_rect()
        sx = self.image_size[0] / max(1, target.width()); sy = self.image_size[1] / max(1, target.height())
        x = max(0, min(self.image_size[0], int((p.x() - target.left()) * sx)))
        y = max(0, min(self.image_size[1], int((p.y() - target.top()) * sy)))
        return x, y

    def _image_to_display_rect(self, region: tuple[int, int, int, int]) -> QRect:
        target = self._target_rect(); x1, y1, x2, y2 = region
        sx = target.width() / max(1, self.image_size[0]); sy = target.height() / max(1, self.image_size[1])
        return QRect(target.left() + int(x1 * sx), target.top() + int(y1 * sy), max(1, int((x2 - x1) * sx)), max(1, int((y2 - y1) * sy)))

    def _display_to_image_rect(self, rect: QRect) -> Optional[tuple[int, int, int, int]]:
        target = self._target_rect(); inter = rect.intersected(target)
        if inter.width() < 4 or inter.height() < 4: return None
        sx = self.image_size[0] / max(1, target.width()); sy = self.image_size[1] / max(1, target.height())
        x1 = max(0, int((inter.left() - target.left()) * sx)); y1 = max(0, int((inter.top() - target.top()) * sy))
        x2 = min(self.image_size[0], int((inter.right() - target.left()) * sx)); y2 = min(self.image_size[1], int((inter.bottom() - target.top()) * sy))
        if x2 <= x1 or y2 <= y1: return None
        return x1, y1, x2, y2

    def _hit_roi_index(self, point: QPoint) -> Optional[int]:
        for idx in range(len(self.rois) - 1, -1, -1):
            if self._image_to_display_rect(self.rois[idx]).contains(point): return idx
        return None

    def _target_shape_display_bbox(self, shape: dict[str, Any]) -> QRect:
        if shape.get("kind") == "polygon":
            pts = [self._image_to_display_point(x, y) for x, y in shape.get("points", [])]
            return QPolygon(pts).boundingRect() if pts else QRect(0, 0, 0, 0)
        return self._image_to_display_rect(target_area_bbox(shape))

    def _hit_target_index(self, point: QPoint) -> Optional[int]:
        for idx in range(len(self.target_areas) - 1, -1, -1):
            shape = self.target_areas[idx]
            if shape.get("kind") == "polygon":
                poly = QPolygon([self._image_to_display_point(x, y) for x, y in shape.get("points", [])])
                if poly.containsPoint(point, Qt.FillRule.OddEvenFill): return idx
            elif self._image_to_display_rect(target_area_bbox(shape)).contains(point):
                return idx
        return None

    def _finish_polygon_target(self) -> None:
        if len(self.poly_points) < 3:
            self.canvas_error.emit("Target Area 至少需要 3 個點才能形成封閉區域。")
            return
        img_pts = [self._display_to_image_point(pt) for pt in self.poly_points]
        # Remove repeated adjacent points and ensure at least three unique vertices.
        cleaned: list[tuple[int, int]] = []
        for pt in img_pts:
            if not cleaned or cleaned[-1] != pt:
                cleaned.append(pt)
        if len(set(cleaned)) < 3:
            self.canvas_error.emit("Target Area 直線繪製未形成有效封閉區域，請重新繪製。")
        else:
            self.target_areas.append({"kind": "polygon", "points": cleaned})
            self.selected_target_indices = {len(self.target_areas) - 1}
            self.region_changed.emit()
        self.poly_points = []
        self.drag_current = None
        self.update()

    @staticmethod
    def _point_distance(a: QPoint, b: QPoint) -> float:
        dx = a.x() - b.x(); dy = a.y() - b.y()
        return (dx * dx + dy * dy) ** 0.5


@dataclass
class UIState:
    project_id: str = ""
    project_name: str = ""
    created_at: str = ""
    updated_at: str = ""
    class_name: str = "stain"
    project_mode: str = APP_MODE
    api_key_set: bool = False
    crop_width: int = 640
    crop_height: int = 640
    crop_mode: str = "crop"
    crop_mode_selected: bool = False
    model: str = "gpt-image-2"
    quality: str = "low"
    size: str = "1280x1280"
    num_outputs: int = 1
    run_name: str = "run"
    dry_run: bool = False
    prompt_mode: str = "custom"
    prompt_template: str = "瑕疵"
    prompt_group: str = ""
    selected_region_stems: list[str] = field(default_factory=list)
    prompt_input: str = ""
    completed_steps: list[bool] = field(default_factory=lambda: [False] * STEP_COUNT)
    # Export scope controls which runs are included in final zip/artifacts.
    # current = current UI run name/latest batch; all = every run under this class/project.
    export_scope: str = "current"
    estimated_run_cost_usd: float = 0.0
    actual_run_cost_usd: float = 0.0
    saved_project: bool = False
    last_active_step: int = 0
    generation_status: str = "idle"
    regions: dict = field(default_factory=dict)
    last_generation_return_code: int = 0
    last_generation_error: str = ""


def ui_state_data(data: dict[str, Any] | None = None) -> dict[str, Any]:
    base = asdict(UIState())
    if isinstance(data, dict):
        for key, value in data.items():
            if key in base:
                base[key] = value
    return base


class StepButton(QPushButton):
    _locked_cursor: Optional[QCursor] = None

    def __init__(self, index: int, title: str) -> None:
        self.raw_title = title
        label = f"Step {index}\n{title}" if index > 0 else f"Step 0\n{title}"
        super().__init__(label)
        self.index = index
        self.is_locked_step = False
        self._override_cursor_active = False
        self.setCheckable(True)
        self.setMouseTracking(True)
        self.setMinimumHeight(64)
        self.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)
        self.setObjectName("StepButton")
        self.setProperty("status", "locked")
        self.setProperty("density", "normal")
        self.setCursor(QCursor(Qt.PointingHandCursor))
        self.setFocusPolicy(Qt.NoFocus)

    def set_responsive_density(self, density: str) -> None:
        """Adapt the Step button height for smaller screens.

        The left workflow rail is frequently used on laptop displays.  Fixed
        64px buttons make the rail feel cramped and may clip the second text
        line when Windows display scaling is high, so the main window adjusts
        the button density on resize while keeping every step readable.
        """
        density = density if density in {"normal", "compact", "tiny"} else "normal"
        heights = {"normal": 64, "compact": 54, "tiny": 48}
        h = heights[density]
        self.setMinimumHeight(h)
        self.setMaximumHeight(h)
        self.setProperty("density", density)
        self.style().unpolish(self)
        self.style().polish(self)

    @classmethod
    def forbidden_cursor(cls) -> QCursor:
        if cls._locked_cursor is not None:
            return cls._locked_cursor
        pix = QPixmap(36, 36)
        pix.fill(Qt.transparent)
        painter = QPainter(pix)
        painter.setRenderHint(QPainter.Antialiasing)
        painter.setPen(QPen(QColor("#ff1744"), 5))
        painter.drawEllipse(5, 5, 26, 26)
        painter.drawLine(11, 25, 25, 11)
        painter.end()
        cls._locked_cursor = QCursor(pix, 18, 18)
        return cls._locked_cursor

    def set_locked(self, locked: bool) -> None:
        self.is_locked_step = locked
        self.setEnabled(True)
        self.setCursor(self.forbidden_cursor() if locked else QCursor(Qt.PointingHandCursor))

    def enterEvent(self, event):  # noqa: N802
        # Do not use QApplication.setOverrideCursor() here. Global override cursor
        # stacks can remain unbalanced if a widget is rebuilt while hovered. The
        # button-level cursor set in set_locked() is sufficient and safer.
        super().enterEvent(event)

    def leaveEvent(self, event):  # noqa: N802
        super().leaveEvent(event)

    def mousePressEvent(self, event):  # noqa: N802
        if self.is_locked_step:
            event.ignore()
            return
        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):  # noqa: N802
        if self.is_locked_step:
            event.ignore()
            return
        super().mouseReleaseEvent(event)


class MainWindow(QMainWindow):
    TR = {
        "app_title": "GPT GenImage UI",
        "workflow_hint": "Step 1 管理專案；專案卡綠色＝Defect、藍色＝Food；紅框＝尚未儲存；綠框＝已完成；灰色＝不可跳步。",
        "step0": "Homepage",
        "step1": "專案與 API",
        "step2": "資料上傳",
        "step3": "裁切尺寸與圖像裁切",
        "step4": "ROI / Target Area",
        "step5": "Prompt 編輯",
        "step6": "模型參數",
        "step7": "Aggregate",
        "step8": "執行生成",
        "step9": "Export / 輸出",
        "help": "HELP？",
        "back": "Back",
        "submit": "Submit / Save Step",
    }

    HELP = {
        0: [("可在此新增、開啟、複製或刪除專案。新增專案時請選擇 Mode：Defect 走瑕疵資料流程，Food 走食物變化流程。專案卡綠色代表 Defect，藍色代表 Food。", False)],
        1: [("第一次使用請輸入 OpenAI API Key，按『儲存 / 替換 API Key』後會立即顯示是否已保存成功。", False), ("重要：若輸入不同 API Key，系統會先要求確認，避免誤改；套件版本檢查已省略。", True)],
        2: [("可拖曳圖片或資料夾上傳；刪除只會移除本專案複製檔，不會刪除原始圖片。", False), ("上方清單用來管理檔案；下方大預覽框用來檢查目前選取圖片。", False)],
        3: [("先點選左側已上傳圖片縮圖，中間才會載入該原圖預覽。", False), ("若要保留全部原圖不裁切，請按底部 Back 與 Submit 之間的『使用原始圖片』；系統會把 Step 2 的全部上傳圖直接作為 Step 4 輸入並跳到 Step 4。", True), ("若需要裁切，請設定裁切框寬高後在中間圖像點選/拖曳裁切框，裁切結果會加入右側 Step 4 輸入圖像。裁切寬高僅作為輸入圖尺寸；建議值為 320(px)～1280(px)，真正的生成輸出尺寸限制在 Step 6。", False)],
        4: [("請針對每張輸入圖框選 ROI 與 Target Area。ROI 是原瑕疵位置；Target Area 是允許隨機生成的位置。", False), ("本頁已啟用自動儲存：新增、刪除、切換 ROI / Target Area 後會自動寫入座標與 mask，不需要另外按儲存。", True), ("左側圖像清單可多選，但此處只用於 Step 4 檢視與快速切換；最終要送進生成的圖像組合改在 Step 5『引用組別』決定。", True), ("ROI 與 Target Area 都支援多個區域、逐一選取刪除與全部刪除；Target Area 重疊時會以聯集方式輸出到同一張 mask。", False), ("直線繪製 Target Area 時，請依序點選多個頂點，最後點回第一個點形成封閉區域；未封閉會提示重新繪製。", True), ("快捷鍵：R=框選 ROI、S=選取 ROI、A=刪除全部 ROI、D=刪除選取 ROI、T=矩形 Target、L=直線 Target、Y=選取 Target、G=刪除全部 Target、H=刪除選取 Target、↑/↓=上一張/下一張圖並同步左側縮圖選取。選取 ROI 或矩形 Target Area 後，可拖曳四角控制點調整寬高，也可拖曳左右邊中點水平調整寬度、上下邊中點垂直調整高度。", False), ("Ctrl + 滑鼠滾輪可縮放，滑鼠中鍵可平移；繪製時游標旁會顯示 ROI 或 Target Area，方便確認目前模式。Target Area 是隨機生成位置的唯一範圍；可依需求畫在任何區域，系統不會再自動排除黑色區域或材質外區域。", False)],
        5: [("請在上方『引用組別』用 Ctrl / Shift 多選最終要送入生成的圖像組合；一次最多 16 組，超過會提示並擋下。", True), ("自訂 prompt 預設空白；使用模板時請按『套用模板到輸入指令』。實際傳送指令不再附加 ROI / Target Area 座標，而是由 Step 8 傳送乾淨原圖 + 標註參考圖讓模型理解位置。", False), ("新版生成模式：Image 1 是乾淨原圖；Image 2 是 ROI / Target Area 標註參考圖。程式不會把座標量化文字送進 prompt，也不會把 ROI / Target Area 當 API mask 修補。", False)],
        6: [("gpt-image-2 輸出尺寸限制：寬高皆需為 16 的倍數、長邊不可超過 3840 px、長寬比不可超過 3:1、總像素需介於 655,360 px～8,294,400 px。", True), ("已恢復正式檢查：按『確認參數並估算本次成本』時，無論是『自訂尺寸』或『與原圖尺寸相同』，只要低於下界、高於上界或比例不合規，都會跳出提示並阻止進入下一步。", True), ("例如 640×640 = 409,600 px，低於最低總像素限制；5472×3648 長邊超過 3840 px 且總像素過高。成本預估會嘗試讀取 OpenAI Pricing；實際金額仍以 API usage / 帳單為準。", False)],
        7: [("Aggregate 會即時顯示目前重點設定。確認內容無誤後按 Submit 進入正式生成。", False)],
        8: [("生成期間會鎖定其他 Step，避免生成中被修改設定。", False), ("重新開始生成不會清除舊資料；系統會依 Step 6 的 Run name 自動建立新資料夾，例如 run1、run2、run3，避免覆蓋已完成圖片。從上次進度開始生成會沿用目前 Run name 並加入 --resume-existing。", True), ("若設定輸出 500 張，批次程式會分配到選定圖像後逐張呼叫後端；目前每張輸出由 run_gpt_image2.py 以 n=1 呼叫一次 OpenAI Image Edit API，因此通常是 500 次 API edit 呼叫。", True), ("Step 8 每次 API 輸入包含原始/裁切輸入圖、Image 2 標註參考圖、prompt、model、size、quality 等參數；reference-guided-edit 不把 ROI/Target Area 當 API mask。", False), ("正式生成會消耗 API 額度。", True)],
        9: [("在『匯出範圍』勾選要匯出的 runs（可複選）；勾選『全部 runs（包含歷次 runs）』會包含所有 runs 並停用其他選項。按『確認』可將所選 runs 的生成圖載入左側預覽。", False), ("按『Export』時需選擇匯出路徑；系統會在該路徑建立一個名為 <專案名稱>-<時間戳> 的資料夾放入所有生成圖。若勾選『打包成 .zip』則改輸出對應的 .zip 檔。", False), ("每個 run 的生成圖都存放在該 run 的 Gen_Images 資料夾中；同層另有記錄每張圖 token/成本的 Excel 與該次 prompt.txt。", True)],
    }

    def __init__(self) -> None:
        super().__init__()
        self.root = project_root()
        self.projects_root = ensure_dir(self.root / "project")
        self.index_path = self.projects_root / "project_index.json"
        self.cleanup_legacy_unsaved_project()
        self.state = self.load_initial_state()
        # Keep the project card selection aligned with the actually loaded project.
        # This prevents Homepage from showing one project's summary while the side
        # navigation/state still belongs to another project after relaunch.
        self.selected_project_card_id = self.state.project_id if self.state.saved_project else ""
        self.current_step = 0
        # Guard programmatic widget refresh/load from marking saved project steps dirty.
        # Without this, reopening a completed project can briefly reload Step 5/6 widgets,
        # emit textChanged/currentTextChanged, and turn later steps red even though artifacts exist.
        self._suspend_dirty_tracking = True
        self.current_process: Optional[QProcess] = None
        self.current_process_label = ""
        # UI-level lock is separate from QProcess.state(). QProcess starts asynchronously,
        # so relying only on state() can leave other Steps clickable for a short window.
        self.generation_ui_locked = False
        self.generation_total = 0
        self.generation_completed = 0
        self._generation_progress_buffer = ""
        self._stop_requested = False
        self._receiving_dialog = None
        self.step_buttons: list[StepButton] = []
        self.footer_buttons: dict[int, dict[str, QPushButton]] = {}
        self.header_widgets: list[tuple[QLabel, QLabel, str, str]] = []
        self.dirty_steps = [not done for done in self.state.completed_steps]
        self.setWindowTitle("GPT GenImage UI - Gen Defect")
        self.resize(1480, 920)
        self.setMinimumSize(1120, 740)
        # Enable file/folder drag-and-drop at the window level.
        # This is intentionally global so drops work even when the cursor is
        # over Step 2 whitespace, the scroll area, the list, or the preview box.
        self.setAcceptDrops(True)
        app = QApplication.instance()
        if app is not None:
            app.installEventFilter(self)
        self.build_ui()
        self.apply_style()
        self.update_sidebar_responsive()
        # Startup always lands on Homepage for recovery safety. The loaded project
        # remains active, but the visible page is Step 0 so the user can choose
        # where to continue. QStackedWidget exposes the active page through
        # currentIndex()/setCurrentIndex(), so set it explicitly on every launch.
        self.current_step = 0
        self.state.last_active_step = 0
        if hasattr(self, "stack"):
            self.stack.setCurrentIndex(0)
        self.save_state()
        self.update_step_buttons()
        try:
            self.refresh_all()
        finally:
            self._suspend_dirty_tracking = False
            self.update_step_buttons()

    def resizeEvent(self, event):  # noqa: N802
        super().resizeEvent(event)
        self.update_sidebar_responsive()

    def update_sidebar_responsive(self) -> None:
        """Keep the left Step rail readable on laptop/smaller screens.

        The main page is intentionally scrollable, but the side workflow rail is
        always visible.  On shorter displays, fixed 64px step cards plus large
        margins can clip the two-line labels and make cards look stacked on top
        of each other.  This method changes only density/spacing/width of the
        sidebar; it does not change workflow state.
        """
        if not hasattr(self, "sidebar"):
            return
        h = max(1, self.height())
        w = max(1, self.width())
        if h < 760:
            density = "tiny"
            margins = (10, 10, 10, 10)
            side_spacing = 6
            step_spacing = 4
            min_w, max_w = 220, 300
            title_pt = 18
            hint_pt = 11
            status_pt = 11
            hint_text = "Step 1 管理專案；紅框＝尚未儲存；綠框＝已完成；灰色＝不可跳步。"
        elif h < 860 or w < 1320:
            density = "compact"
            margins = (12, 12, 12, 12)
            side_spacing = 8
            step_spacing = 6
            min_w, max_w = 235, 320
            title_pt = 20
            hint_pt = 12
            status_pt = 12
            hint_text = self.TR["workflow_hint"]
        else:
            density = "normal"
            margins = (16, 16, 16, 16)
            side_spacing = 10
            step_spacing = 9
            min_w, max_w = 260, 340
            title_pt = 24
            hint_pt = 14
            status_pt = 14
            hint_text = self.TR["workflow_hint"]
        self.sidebar.setMinimumWidth(min_w)
        self.sidebar.setMaximumWidth(max_w)
        if hasattr(self, "sidebar_layout"):
            self.sidebar_layout.setContentsMargins(*margins)
            self.sidebar_layout.setSpacing(side_spacing)
        if hasattr(self, "step_layout"):
            self.step_layout.setSpacing(step_spacing)
        if hasattr(self, "sidebar_title"):
            self.sidebar_title.setMinimumWidth(max(160, min_w - margins[0] - margins[2] - 10))
            self.sidebar_title.setStyleSheet(f"font-size: {title_pt}px;")
        if hasattr(self, "sidebar_hint"):
            self.sidebar_hint.setText(hint_text)
            self.sidebar_hint.setStyleSheet(f"font-size: {hint_pt}px;")
            self.sidebar_hint.setMaximumHeight(54 if density == "tiny" else 88)
        if hasattr(self, "status_label"):
            self.status_label.setStyleSheet(f"font-size: {status_pt}px;")
            self.status_label.setMinimumHeight(48 if density == "tiny" else 58)
            self.status_label.setMaximumHeight(76 if density == "tiny" else 120)
        for btn in getattr(self, "step_buttons", []):
            btn.set_responsive_density(density)
        if hasattr(self, "step_scroll"):
            # Keep vertical scrolling as the safety net.  The density changes make
            # each card readable; the scroll area guarantees no card is squeezed
            # or over-painted when the window is very short.
            self.step_scroll.setVerticalScrollBarPolicy(Qt.ScrollBarPolicy.ScrollBarAsNeeded)

    # ---------- project persistence ----------
    def load_index(self) -> dict:
        if self.index_path.exists():
            try:
                return json.loads(self.index_path.read_text(encoding="utf-8"))
            except Exception:
                return {"active_project_id": "", "projects": []}
        return {"active_project_id": "", "projects": []}

    def save_index(self, data: dict) -> None:
        ensure_dir(self.index_path.parent)
        self.index_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")

    def mode_projects_root(self, mode: str) -> Path:
        """Return the per-mode project root used by the integrated launcher."""
        mode = "food" if str(mode).lower() == "food" else "defect"
        mode_dir = self.root.parent / mode
        folder_name = "project"
        return ensure_dir(mode_dir / folder_name)

    def load_mode_index(self, mode: str) -> dict:
        """Load one mode's project index without changing the active workflow."""
        idx_path = self.mode_projects_root(mode) / "project_index.json"
        if idx_path.exists():
            try:
                data = json.loads(idx_path.read_text(encoding="utf-8"))
                if isinstance(data, dict):
                    return data
            except Exception:
                pass
        return {"active_project_id": "", "projects": []}

    def load_combined_projects(self) -> list[dict]:
        """Return Defect and Food projects together for the shared Homepage."""
        combined: list[dict] = []
        seen: set[tuple[str, str]] = set()
        for mode in ("defect", "food"):
            proot = self.mode_projects_root(mode)
            idx = self.load_mode_index(mode)
            for item in idx.get("projects", []) or []:
                if not isinstance(item, dict):
                    continue
                pid = str(item.get("id") or "").strip()
                if not pid:
                    continue
                key = (mode, pid)
                if key in seen:
                    continue
                rec = dict(item)
                rec["id"] = pid
                rec["mode"] = mode
                rec.setdefault("project_mode", mode)
                rec["_project_root"] = str(proot)
                combined.append(rec)
                seen.add(key)
            # Recovery fallback: include projects whose folder exists even if the
            # index is stale. project_state.json no longer exists, so a project is
            # detected purely by its folder under the mode's project root.
            if proot.exists():
                for child in sorted(proot.iterdir()):
                    if not child.is_dir():
                        continue
                    pid = child.name
                    key = (mode, pid)
                    if not pid or key in seen:
                        continue
                    combined.append({
                        "id": pid,
                        "name": pid,
                        "class_name": "-",
                        "mode": mode,
                        "project_mode": mode,
                        "model": "-",
                        "quality": "-",
                        "num_outputs": 1,
                        "created_at": "-",
                        "updated_at": "-",
                        "_project_root": str(proot),
                    })
                    seen.add(key)
        # Fixed display order: sort by creation time (then id) so selecting or
        # opening a project never reorders the cards on the Homepage.
        return sorted(combined, key=lambda p: (str(p.get("created_at") or ""), str(p.get("id") or "")))

    def find_project_record(self, pid: str) -> dict | None:
        for proj in self.load_combined_projects():
            if str(proj.get("id") or "") == str(pid):
                return proj
        return None

    def record_state_dict(self, pid: str) -> dict | None:
        """Build a UIState dict for a project from the lightweight
        project_index.json record. project_state.json is no longer persisted, so
        detailed progress is re-derived from disk by
        reconcile_completed_steps_for_loaded_project()."""
        rec = self.find_project_record(pid)
        if rec is None:
            return None
        data: dict[str, Any] = {
            "project_id": str(rec.get("id") or pid),
            "project_name": str(rec.get("name") or rec.get("id") or pid),
            "project_mode": str(rec.get("project_mode") or rec.get("mode") or APP_MODE),
            "saved_project": True,
            "created_at": str(rec.get("created_at") or ""),
            "updated_at": str(rec.get("updated_at") or ""),
        }
        for key in ("class_name", "model", "quality"):
            value = rec.get(key)
            if value and str(value) != "-":
                data[key] = str(value)
        try:
            if rec.get("num_outputs"):
                data["num_outputs"] = int(rec["num_outputs"])
        except Exception:
            pass
        # Restore the recorded per-step completion flags and the ROI/Target geometry
        # so reopening a project keeps its progress. project_state.json is no longer
        # used; these now live inside the project_index.json record.
        if isinstance(rec.get("completed_steps"), list):
            data["completed_steps"] = [bool(v) for v in rec["completed_steps"]][:STEP_COUNT]
        if isinstance(rec.get("regions"), dict):
            data["regions"] = rec["regions"]
        return ui_state_data(data)

    def project_mode_for_id(self, pid: str) -> str:
        rec = self.find_project_record(pid)
        mode = str((rec or {}).get("mode") or APP_MODE).lower()
        return mode if mode in {"defect", "food"} else APP_MODE

    def cleanup_legacy_unsaved_project(self) -> None:
        """Remove the old transient project/unsaved_project folder if it exists."""
        legacy = self.projects_root / "unsaved_project"
        if not legacy.exists():
            return
        try:
            shutil.rmtree(legacy, ignore_errors=True)
        except Exception:
            pass

    def load_initial_state(self) -> UIState:
        idx = self.load_index()
        pid = idx.get("active_project_id") or ""
        if pid:
            base = self.record_state_dict(pid)
            if base is not None:
                try:
                    st = UIState(**base)
                    self.reconcile_completed_steps_for_loaded_project(st)
                    return st
                except Exception:
                    pass
        return UIState()

    def candidate_project_dirs(self, state: UIState | None = None) -> list[Path]:
        """Return possible project folders for old/new saved states.

        Older builds used either project_id or project_name/class-name based folders.
        Reopen/recovery must inspect all plausible folders so a completed project is
        not shown as stuck at Step 5 merely because artifacts live under another
        compatible path.
        """
        st = state or self.state
        names = [st.project_name, st.project_id, sanitize_name(st.project_name or ""), sanitize_name(st.project_id or "")]
        out: list[Path] = []
        seen: set[str] = set()
        for name in names:
            key = str(name or "").strip()
            if not key:
                continue
            p = self.projects_root / sanitize_name(key)
            sk = str(p.resolve()) if p.exists() else str(p)
            if sk not in seen:
                seen.add(sk); out.append(p)
        return out

    def project_has_generation_outputs(self, state: UIState | None = None) -> bool:
        """Detect real generated outputs across both old and new run layouts.

        The current layout stores outputs under runs/<run>/Gen_Images/<name>.<ext>.
        Older project versions used runs/<class>/<run>/<seed>/edited_seed*.png plus a
        metadata.json, and class names can differ after project duplication. Searching
        the whole project runs tree prevents false unfinished Step 5/6/8 status when
        generated files already exist.
        """
        for pdir in self.candidate_project_dirs(state):
            runs_root = pdir / "runs"
            if not runs_root.exists():
                continue
            # Current layout: runs/<run_name>/Gen_Images/<run_name>-<Y>-<M>-<D>-<H>-<m>-<counter>.<ext>
            for gen_dir in runs_root.glob("*/Gen_Images"):
                if not gen_dir.is_dir():
                    continue
                for img in gen_dir.iterdir():
                    if img.is_file() and img.suffix.lower() in SUPPORTED_EXTS and img.stat().st_size > 0:
                        return True
            # Legacy layouts: per-image metadata.json or seed-named final images.
            if any(runs_root.rglob("metadata.json")):
                return True
            patterns = ["edited_seed*.png", "generated_seed*.png", "*_seed*.png"]
            for pat in patterns:
                for img in runs_root.rglob(pat):
                    if img.is_file() and img.stat().st_size > 0:
                        return True
        return False

    def project_has_export_outputs(self, state: UIState | None = None) -> bool:
        for pdir in self.candidate_project_dirs(state):
            exp_root = pdir / "exports"
            if exp_root.exists() and any(p.is_file() and p.stat().st_size > 0 for p in exp_root.rglob("*")):
                return True
        return False

    def _class_input_dirs_for_state(self, state: UIState | None = None) -> list[tuple[Path, Path]]:
        st = state or self.state
        dirs: list[tuple[Path, Path]] = []
        seen: set[str] = set()
        preferred_classes = [sanitize_name(st.class_name or ""), sanitize_name(st.project_name or ""), sanitize_name(st.project_id or "")]
        for pdir in self.candidate_project_dirs(st):
            data_root = pdir / "data" / "01_inputs"
            for cname in preferred_classes:
                if cname:
                    img_dir = data_root / cname / "images"
                    reg_dir = data_root / cname / "regions"
                    key = f"{img_dir}|{reg_dir}"
                    if key not in seen:
                        seen.add(key); dirs.append((img_dir, reg_dir))
            if data_root.exists():
                for img_dir in data_root.glob("*/images"):
                    reg_dir = img_dir.parent / "regions"
                    key = f"{img_dir}|{reg_dir}"
                    if key not in seen:
                        seen.add(key); dirs.append((img_dir, reg_dir))
        return dirs

    def project_has_required_regions(self, state: UIState | None = None) -> bool:
        st = state or self.state
        regions = st.regions if isinstance(getattr(st, "regions", None), dict) else {}
        if regions and any((g.get("rois") and g.get("targets")) for g in regions.values() if isinstance(g, dict)):
            return True
        for inputs, regions in self._class_input_dirs_for_state(state):
            crops = list_images(inputs)
            if not crops:
                continue
            ok = True
            for img in crops:
                reg_path = regions / f"{img.stem}.txt"
                reg = parse_region_txt(reg_path)
                if not any(k.startswith("ROI") for k in reg) or not parse_target_areas(reg_path):
                    ok = False
                    break
            if ok:
                return True
        return False

    def project_prompt_exists(self, state: UIState | None = None) -> bool:
        st = state or self.state
        preferred = [sanitize_name(st.class_name or ""), sanitize_name(st.project_name or ""), sanitize_name(st.project_id or "")]
        for pdir in self.candidate_project_dirs(st):
            cfg_root = pdir / "data" / "01_inputs"
            for cname in preferred:
                if cname:
                    prompt = cfg_root / cname / "prompt.txt"
                    if prompt.exists() and bool(prompt.read_text(encoding="utf-8", errors="ignore").strip()):
                        return True
            if cfg_root.exists():
                for prompt in cfg_root.glob("*/prompt.txt"):
                    if prompt.exists() and bool(prompt.read_text(encoding="utf-8", errors="ignore").strip()):
                        return True
        return False

    def reconcile_completed_steps_for_loaded_project(self, st: UIState | None = None) -> None:
        """Restore sidebar green status from durable artifacts when reopening projects.

        Project state from older builds can be stale because UI refresh signals may
        have marked Step 5/6 dirty after generation.  Durable artifacts are the
        source of truth: if generation outputs exist, the workflow reached Step 8;
        if exports exist, Step 9 is also complete.
        """
        target = st or self.state
        if not target.project_id:
            return
        if len(target.completed_steps) != STEP_COUNT:
            fixed = [False] * STEP_COUNT
            for i, v in enumerate(target.completed_steps[:STEP_COUNT]):
                fixed[i] = bool(v)
            target.completed_steps = fixed
        target.completed_steps[0] = True
        target.completed_steps[1] = True
        if target.completed_steps[3] or self.project_has_required_regions(target):
            target.crop_mode_selected = True
        has_runs = self.project_has_generation_outputs(target)
        has_exports = self.project_has_export_outputs(target)
        has_regions = self.project_has_required_regions(target)
        has_prompt = bool(target.prompt_input.strip()) or self.project_prompt_exists(target)
        if has_runs:
            for i in range(0, 9):
                target.completed_steps[i] = True
            if has_exports or bool(target.completed_steps[9]):
                target.completed_steps[9] = True
        elif has_regions and has_prompt:
            for i in range(0, 7):
                target.completed_steps[i] = True
        elif has_regions:
            for i in range(0, 5):
                target.completed_steps[i] = True

    def save_state(self) -> None:
        if self.state.project_id:
            ensure_dir(self.project_dir())
            now_iso = datetime.now().isoformat(timespec="seconds")
            if not self.state.created_at:
                self.state.created_at = now_iso
            self.state.updated_at = now_iso
            # project_state.json is no longer written; only the lightweight
            # project_index.json registry is kept so the Homepage can list projects.
            idx = self.load_index()
            idx["active_project_id"] = self.state.project_id
            if self.state.saved_project:
                projects = [p for p in idx.get("projects", []) if p.get("id") != self.state.project_id]
                projects.append({
                    "id": self.state.project_id,
                    "name": self.state.project_name or self.state.project_id,
                    "class_name": self.state.class_name,
                    "mode": getattr(self.state, "project_mode", APP_MODE),
                    "model": self.state.model,
                    "quality": self.state.quality,
                    "num_outputs": self.state.num_outputs,
                    # Persisted so reopening restores exact progress + annotations.
                    "completed_steps": list(self.state.completed_steps),
                    "regions": self.state.regions if isinstance(self.state.regions, dict) else {},
                    "created_at": self.state.created_at,
                    "updated_at": now_iso,
                })
                idx["projects"] = sorted(projects, key=lambda p: p.get("updated_at", ""), reverse=True)
            self.save_index(idx)

    # ---------- paths ----------
    def class_name(self) -> str:
        return sanitize_name(self.class_edit.text() if hasattr(self, "class_edit") else self.state.class_name)

    def project_dir(self) -> Path:
        # All user task products are isolated under project/<project_name>/, without timestamp suffix.
        # Before a real project exists, keep transient UI state under _ui_state instead of creating
        # an unnecessary project/unsaved_project folder.
        name = self.state.project_name or self.state.project_id or "_ui_state"
        return ensure_dir(self.projects_root / sanitize_name(name))

    def raw_dir(self) -> Path:
        # Image 1 / input images live flat under data/raw_image/.
        return ensure_dir(self.project_dir() / "data" / "raw_image")

    def inputs_dir(self) -> Path:
        # Generation Image 1 source == raw_image folder.
        return self.raw_dir()

    def reference_dir(self) -> Path:
        # Image 2: pre-rendered ROI/Target annotation reference, paired by stem.
        return ensure_dir(self.project_dir() / "data" / "reference_image")

    def regions_dir(self) -> Path:
        # Legacy: ROI/Target geometry is now stored in project_state.json. Keep a
        # hidden, non-created path so old callers do not crash.
        return self.project_dir() / "data" / "_unused"

    def masks_dir(self) -> Path:
        return self.project_dir() / "data" / "_unused"

    def target_masks_dir(self) -> Path:
        return self.project_dir() / "data" / "_unused"

    def class_config_dir(self) -> Path:
        return ensure_dir(self.project_dir() / "data")

    def prompt_path(self) -> Path:
        return self.class_config_dir() / "prompt.txt"

    def runs_dir(self) -> Path:
        # No <class_name> folder layer: each run folder lives directly under runs/.
        return self.project_dir() / "runs"

    def exports_dir(self) -> Path:
        return self.project_dir() / "exports"

    def current_run_name(self, create_if_empty: bool = False) -> str:
        name = sanitize_name(self.state.run_name or "")
        if not name and create_if_empty:
            name = "run"
            self.state.run_name = name
            if hasattr(self, "run_name_edit"):
                self._set_widget_value_blocked(self.run_name_edit, "setText", name)
        return name

    def export_scope_is_all(self) -> bool:
        return getattr(self.state, "export_scope", "current") == "all"

    def current_export_name(self) -> str:
        if self.export_scope_is_all():
            return "all_runs"
        return self.current_run_name(create_if_empty=True)

    def current_export_dir(self) -> Path:
        return self.exports_dir() / self.current_export_name()

    def app_log_path(self) -> Path:
        return self.root.parent.parent / "log.txt"

    def log_event(self, message: str) -> None:
        try:
            with self.app_log_path().open("a", encoding="utf-8") as f:
                f.write(f"[{datetime.now().isoformat(timespec='seconds')}] [{APP_MODE}] {message}\n")
        except Exception:
            pass

    # ---------- UI construction ----------
    def build_ui(self) -> None:
        central = QWidget()
        main = QHBoxLayout(central)
        main.setContentsMargins(0, 0, 0, 0)
        main.setSpacing(0)
        sidebar = QFrame(); sidebar.setObjectName("Sidebar")
        self.sidebar = sidebar
        sidebar.setMinimumWidth(240); sidebar.setMaximumWidth(340)
        side_lay = QVBoxLayout(sidebar); side_lay.setContentsMargins(16, 16, 16, 16); side_lay.setSpacing(10)
        self.sidebar_layout = side_lay
        title = QLabel(self.TR["app_title"]); title.setObjectName("AppTitle"); title.setWordWrap(False); title.setMinimumWidth(190); side_lay.addWidget(title)
        self.sidebar_title = title
        hint = QLabel(self.TR["workflow_hint"]); hint.setObjectName("SideHint"); hint.setWordWrap(True); side_lay.addWidget(hint)
        self.sidebar_hint = hint
        step_container = QWidget(); step_container.setObjectName("StepContainer"); step_container.setAutoFillBackground(False); step_container.setAttribute(Qt.WA_StyledBackground, True); step_lay = QVBoxLayout(step_container); step_lay.setContentsMargins(0,0,0,0); step_lay.setSpacing(9)
        self.step_layout = step_lay
        for display_idx, internal_idx in enumerate(VISIBLE_STEPS):
            btn = StepButton(display_idx + 1, self.TR[f"step{internal_idx}"])
            btn.index = internal_idx
            btn.display_index = display_idx + 1
            btn.clicked.connect(self.safe_action(f"goto_step_{internal_idx}", lambda idx=internal_idx: self.goto_step(idx)))
            self.step_buttons.append(btn); step_lay.addWidget(btn)
        step_lay.addStretch(1)
        step_scroll = QScrollArea(); step_scroll.setObjectName("StepScroll"); step_scroll.setWidgetResizable(True); step_scroll.setFrameShape(QFrame.NoFrame); step_scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff); step_scroll.viewport().setObjectName("StepScrollViewport"); step_scroll.viewport().setAutoFillBackground(False); step_scroll.viewport().setAttribute(Qt.WA_StyledBackground, True); step_scroll.setWidget(step_container)
        self.step_scroll = step_scroll
        side_lay.addWidget(step_scroll, 1)
        self.status_label = QLabel("Status: 尚未執行"); self.status_label.setWordWrap(True); self.status_label.setObjectName("StatusBox")
        side_lay.addWidget(self.status_label)
        main.addWidget(sidebar, 0)
        self.stack = QStackedWidget(); self.stack.setObjectName("MainStack"); main.addWidget(self.stack, 1)
        self.stack.addWidget(self.page_home())
        self.stack.addWidget(self.page_upload())
        self.stack.addWidget(self.page_crop())
        self.stack.addWidget(self.page_regions())
        self.stack.addWidget(self.page_prompt())
        self.stack.addWidget(self.page_model())
        self.stack.addWidget(self.page_aggregate())
        self.stack.addWidget(self.page_run())
        self.stack.addWidget(self.page_export())
        self.setCentralWidget(central)
        # Make the major containers eligible drop targets too. Some child
        # widgets inside a QScrollArea may otherwise prevent MainWindow from
        # receiving the first dragEnterEvent.
        for w in (central, self.stack):
            w.setAcceptDrops(True)

    def _event_paths(self, event) -> list[Path]:
        """Return dropped image files / image folders from a Qt drag/drop event."""
        if not hasattr(event, "mimeData"):
            return []
        mime = event.mimeData()
        if not mime or not mime.hasUrls():
            return []
        paths: list[Path] = []
        for url in mime.urls():
            p = Path(url.toLocalFile())
            if p.is_dir() or (p.is_file() and p.suffix.lower() in SUPPORTED_EXTS):
                paths.append(p)
        return paths

    def eventFilter(self, obj, event):  # noqa: N802
        # Soft-locked footer buttons stay enabled so the forbidden cursor can be
        # shown, but all mouse/keyboard activation events are swallowed here.
        # This is used while Step 8 generation is running.
        try:
            if bool(obj.property("softLocked")):
                blocked_types = {
                    QEvent.MouseButtonPress, QEvent.MouseButtonRelease, QEvent.MouseButtonDblClick,
                    QEvent.KeyPress, QEvent.KeyRelease, QEvent.Shortcut, QEvent.ShortcutOverride,
                }
                if event.type() in blocked_types:
                    try:
                        event.accept()
                    except Exception:
                        pass
                    return True
        except Exception:
            pass

        # Step 2 should accept drag-and-drop from anywhere inside the window,
        # not only from the QListWidget / preview widget. This also fixes cases
        # where QScrollArea's viewport captures the drag event before the child
        # drop widgets receive it.
        if self.current_step == 2 and event.type() in (QEvent.DragEnter, QEvent.DragMove, QEvent.Drop):
            paths = self._event_paths(event)
            if paths:
                accept_file_drop(event)
                if event.type() == QEvent.Drop:
                    self.import_source_paths(paths)
                return True
        return super().eventFilter(obj, event)

    def dragEnterEvent(self, event):  # noqa: N802
        if self.current_step == 2 and self._event_paths(event):
            accept_file_drop(event)
        else:
            super().dragEnterEvent(event)

    def dragMoveEvent(self, event):  # noqa: N802
        self.dragEnterEvent(event)

    def dropEvent(self, event):  # noqa: N802
        if self.current_step == 2:
            paths = self._event_paths(event)
            if paths:
                self.import_source_paths(paths)
                accept_file_drop(event)
                return
        super().dropEvent(event)

    def wrap_page(self, step_idx: int, content_layout: QVBoxLayout) -> QWidget:
        page = QWidget(); page.setAcceptDrops(True); page_lay = QVBoxLayout(page); page_lay.setContentsMargins(16,16,16,16); page_lay.setSpacing(12)
        body = QWidget(); body.setAcceptDrops(True); body_lay = QVBoxLayout(body); body_lay.setContentsMargins(0,0,0,0); body_lay.setSpacing(12)
        while content_layout.count():
            item = content_layout.takeAt(0)
            if item.widget(): body_lay.addWidget(item.widget())
            elif item.layout(): body_lay.addLayout(item.layout())
        if step_idx not in {0, 2, 3, 4, 7, 8, 9}:
            body_lay.addStretch(1)
        scroll = QScrollArea(); scroll.setObjectName("PageScroll"); scroll.setAcceptDrops(True); scroll.viewport().setAcceptDrops(True); scroll.setWidgetResizable(True); scroll.setFrameShape(QFrame.NoFrame); scroll.setWidget(body)
        page_lay.addWidget(scroll, 1)
        footer = QHBoxLayout(); footer.setSpacing(10)
        help_btn = QPushButton(self.TR["help"]); help_btn.setObjectName("HelpButton"); help_btn.clicked.connect(self.safe_action(f"show_help_{step_idx}", lambda: self.show_help(step_idx)))
        back_btn = QPushButton(self.TR["back"]); back_btn.clicked.connect(self.safe_action("go_back", self.go_back))
        self.footer_buttons[step_idx] = {"help": help_btn, "back": back_btn}
        footer.addWidget(help_btn); footer.addStretch(1); footer.addWidget(back_btn)
        if step_idx == 3:
            use_all_original_btn = QPushButton("使用原始圖片")
            use_all_original_btn.setObjectName("PrimaryButton")
            use_all_original_btn.clicked.connect(self.safe_action("use_all_original_images_and_go_step4", self.use_all_original_images_and_go_step4))
            self.footer_buttons[step_idx]["use_original_all"] = use_all_original_btn
            footer.addWidget(use_all_original_btn)
        if step_idx == 9:
            # Step 9 has no footer Submit; the Export button is inside the Export 規劃 section.
            home_btn = QPushButton("回首頁")
            home_btn.setObjectName("PrimaryButton")
            home_btn.setMinimumWidth(118)
            home_btn.clicked.connect(self.safe_action("go_home_from_export", self.go_home_from_export))
            self.footer_buttons[step_idx]["home"] = home_btn
            footer.addWidget(home_btn)
        else:
            submit_btn = QPushButton(self.TR["submit"]); submit_btn.setObjectName("SubmitButton"); submit_btn.clicked.connect(self.safe_action(f"submit_step_{step_idx}", lambda: self.submit_step(step_idx)))
            self.footer_buttons[step_idx]["submit"] = submit_btn
            footer.addWidget(submit_btn)
        page_lay.addLayout(footer)
        return page

    def header(self, title: str, subtitle: str) -> QWidget:
        box = QFrame(); box.setObjectName("Header")
        lay = QVBoxLayout(box)
        m = re.match(r"^(Step\s+)(\d+)(.*)$", title)
        if m:
            internal_step = int(m.group(2))
            if internal_step in STEP_DISPLAY_INDEX:
                title = f"{m.group(1)}{STEP_DISPLAY_INDEX[internal_step]}{m.group(3)}"
        t = QLabel(title); t.setObjectName("PageTitle")
        if subtitle:
            box.setToolTip(subtitle)
        # 操作說明統一放到各 Step 左下角 HELP？；標題區只保留頁名，避免外部版面被提示文字占用。
        lay.addWidget(t)
        return box

    def page_home(self) -> QWidget:
        lay = QVBoxLayout()
        lay.addWidget(self.header("Step 0｜Homepage / 專案管理", "建立與管理專案；專案名稱不可重複，所有產物會儲存在該專案底下。"))
        row = QHBoxLayout()
        new_btn = QPushButton("新增專案"); new_btn.setObjectName("PrimaryButton"); new_btn.clicked.connect(self.safe_action("create_project", self.create_project))
        row.addWidget(new_btn); row.addStretch(1)
        lay.addLayout(row)

        api_box = QGroupBox("Shared OpenAI API")
        api_lay = QGridLayout(api_box)
        api_lay.setColumnStretch(1, 1)
        self.home_api_edit = QLineEdit()
        self.home_api_edit.setEchoMode(QLineEdit.Password)
        save_api = QPushButton("Save / Replace API Key")
        save_api.clicked.connect(self.safe_action("save_home_api_key", lambda: self.save_api_key_from_editor(self.home_api_edit)))
        api_lay.addWidget(QLabel("OpenAI API Key"), 0, 0)
        api_lay.addWidget(self.home_api_edit, 0, 1)
        api_lay.addWidget(save_api, 0, 2)
        lay.addWidget(api_box)
        self.refresh_api_placeholder()

        self.home_current_project_label = QLabel("")
        self.home_current_project_label.setObjectName("ActiveProjectLabel")
        self.home_current_project_label.setWordWrap(True)
        lay.addWidget(self.home_current_project_label)

        self.project_error = QLabel(""); self.project_error.setObjectName("DangerText"); self.project_error.setVisible(False); lay.addWidget(self.project_error)

        cards_box = QGroupBox("Projects")
        # Tall enough to show two rows of project cards without scrolling.
        cards_box.setMinimumHeight(560)
        cards_box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        cards_lay = QVBoxLayout(cards_box)
        cards_lay.setContentsMargins(14, 18, 14, 14)
        cards_lay.setSpacing(12)
        rule = QLabel("Rule: project names must be unique; each card keeps its own settings and outputs.")
        rule.setObjectName("SideHint"); rule.setWordWrap(True); cards_lay.addWidget(rule)
        self.project_cards_container = QWidget(); self.project_cards_layout = QGridLayout(self.project_cards_container)
        self.project_cards_layout.setContentsMargins(0, 0, 0, 0); self.project_cards_layout.setSpacing(18)
        self.project_cards_layout.setColumnStretch(0, 1); self.project_cards_layout.setColumnStretch(1, 1); self.project_cards_layout.setColumnStretch(2, 1)
        self.project_cards_scroll = QScrollArea(); self.project_cards_scroll.setWidgetResizable(True); self.project_cards_scroll.setFrameShape(QFrame.NoFrame); self.project_cards_scroll.setMinimumHeight(470); self.project_cards_scroll.setWidget(self.project_cards_container)
        cards_lay.addWidget(self.project_cards_scroll, 1)
        lay.addWidget(cards_box, 3)

        self.project_summary = QTextEdit(); self.project_summary.setReadOnly(True)
        self.project_summary.setMinimumHeight(150)
        self.project_summary.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        lay.addWidget(self.project_summary, 1)
        return self.wrap_page(0, lay)

    def page_upload(self) -> QWidget:
        lay = QVBoxLayout()
        lay.addWidget(self.header("Step 2｜資料上傳", "選擇或拖曳原始圖片；下一步會依照指定尺寸進行裁切。"))
        row = QHBoxLayout()
        choose = QPushButton("選擇圖片資料夾"); choose.clicked.connect(self.safe_action("choose_source_folder", self.choose_source_folder))
        delete_btn = QPushButton("刪除選取圖片"); delete_btn.clicked.connect(self.safe_action("delete_selected_uploads", self.delete_selected_uploads))
        row.addWidget(choose); row.addWidget(delete_btn); row.addStretch(1)
        lay.addLayout(row)
        self.upload_list = ImageDropList(); self.upload_list.paths_dropped.connect(self.safe_slot("upload_list_drop", lambda paths: self.import_source_paths([Path(p) for p in paths]))); self.upload_list.delete_requested.connect(self.safe_action("delete_selected_uploads_key", self.delete_selected_uploads)); self.upload_list.itemSelectionChanged.connect(self.safe_action("preview_upload_selection", self.preview_upload_selection))
        self.upload_list.setMaximumHeight(165)
        lay.addWidget(self.upload_list, 0)
        self.upload_preview = ImagePreview("尚無預覽，請選取圖片或直接拖曳圖片到此處。", accept_drops=True); self.upload_preview.paths_dropped.connect(self.safe_slot("upload_preview_drop", lambda paths: self.import_source_paths([Path(p) for p in paths])))
        self.upload_preview.setMinimumHeight(520)
        lay.addWidget(self.upload_preview, 10)
        return self.wrap_page(2, lay)

    def page_crop(self) -> QWidget:
        lay = QVBoxLayout()
        lay.addWidget(self.header("Step 3｜裁切尺寸與圖像裁切", "請先點選左側縮圖；可直接使用原始圖片，也可用裁切框裁切後送入 Step 4。"))

        lay.addWidget(self.build_step3_workflow(), 1)
        self.validate_crop_inputs()
        return self.wrap_page(3, lay)

    def build_step3_mode_menu(self) -> QWidget:
        page = QWidget()
        outer = QVBoxLayout(page)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.setSpacing(16)
        row = QHBoxLayout()
        row.setSpacing(18)

        def make_mode_card(title: str, body: str, button_text: str, mode: str) -> QFrame:
            card = QFrame()
            card.setObjectName("ModeCard")
            card.setMinimumHeight(520)
            card.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            card_lay = QVBoxLayout(card)
            card_lay.setContentsMargins(28, 28, 28, 28)
            card_lay.setSpacing(18)
            t = QLabel(title)
            t.setObjectName("CardTitle")
            t.setAlignment(Qt.AlignCenter)
            t.setWordWrap(True)
            b = QLabel(body)
            b.setObjectName("CardMeta")
            b.setAlignment(Qt.AlignCenter)
            b.setWordWrap(True)
            btn = QPushButton(button_text)
            btn.setObjectName("PrimaryButton")
            btn.setMinimumHeight(54)
            btn.clicked.connect(self.safe_action(f"select_step3_{mode}", lambda m=mode: self.select_step3_mode(m)))
            card_lay.addStretch(1)
            card_lay.addWidget(t)
            card_lay.addWidget(b)
            card_lay.addStretch(1)
            card_lay.addWidget(btn)
            return card

        row.addWidget(make_mode_card("使用原圖尺寸", "不做裁切，直接將 Step 2 原始圖片作為 Step 4 的輸入。", "使用此模式", "no_crop"), 1)
        row.addWidget(make_mode_card("裁切尺寸", "依固定寬高在原圖上選擇裁切區域，產生 Step 4 的輸入圖片。", "使用此模式", "crop"), 1)
        outer.addLayout(row, 1)
        return page

    def build_step3_workflow(self) -> QWidget:
        page = QWidget()
        lay = QVBoxLayout(page)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(12)

        ctrl = QGroupBox("裁切框設定")
        self.crop_controls_box = ctrl
        g = QGridLayout(ctrl); g.setColumnStretch(4,1)
        self.crop_w_edit = QLineEdit(str(self.state.crop_width)); self.crop_w_edit.setValidator(QIntValidator(0, 999999, self))
        self.crop_h_edit = QLineEdit(str(self.state.crop_height)); self.crop_h_edit.setValidator(QIntValidator(0, 999999, self))
        self.crop_confirm_btn = QPushButton("確認裁切框尺寸"); self.crop_confirm_btn.clicked.connect(self.safe_action("confirm_crop_size", self.confirm_crop_size))
        self.crop_error = QLabel("建議輸入值範圍：320(px)~1280(px)")
        self.crop_error.setObjectName("DangerText"); self.crop_error.setVisible(True)
        self.crop_w_edit.textChanged.connect(self.safe_slot("validate_crop_w", lambda *_: self.validate_crop_inputs())); self.crop_h_edit.textChanged.connect(self.safe_slot("validate_crop_h", lambda *_: self.validate_crop_inputs()))
        g.addWidget(QLabel("寬"), 0, 0); g.addWidget(self.crop_w_edit, 0, 1)
        g.addWidget(QLabel("高"), 0, 2); g.addWidget(self.crop_h_edit, 0, 3); g.addWidget(self.crop_confirm_btn, 0, 4)
        g.addWidget(self.crop_error, 1, 0, 1, 5)
        lay.addWidget(ctrl)

        body = QHBoxLayout()
        body.setSpacing(14)
        left = QVBoxLayout()
        left.addWidget(QLabel("已上傳圖像縮圖"))
        self.raw_thumb_grid = VerticalThumbGrid()
        self.raw_thumb_grid.setMinimumWidth(230)
        self.raw_thumb_grid.setMaximumWidth(280)
        self.raw_thumb_grid.selected_path_changed.connect(self.safe_slot("crop_select_raw_image", self.crop_select_raw_image))
        raw_wrap = QWidget()
        raw_wrap_lay = QHBoxLayout(raw_wrap)
        raw_wrap_lay.setContentsMargins(0, 0, 0, 0)
        raw_wrap_lay.addStretch(1)
        raw_wrap_lay.addWidget(self.raw_thumb_grid)
        raw_wrap_lay.addStretch(1)
        left.addWidget(raw_wrap, 1)

        mid = QVBoxLayout()
        mid.addWidget(QLabel("原尺寸圖像 / 固定裁切框"))
        self.crop_canvas = CropCanvas()
        self.crop_canvas.crop_callback = self.make_crop_from_rect
        self.crop_canvas.crop_made.connect(self.safe_slot("crop_made_refresh", lambda p="": self.refresh_crops(auto_select=True, select_path=(Path(p) if p else None))))
        mid.addWidget(self.crop_canvas, 1)

        right = QVBoxLayout()
        right.addWidget(QLabel("Step 4 輸入圖像"))
        self.crop_done_grid = CenteredThumbGrid()
        self.crop_done_grid.setMinimumWidth(240)
        self.crop_done_grid.selected_path_changed.connect(self.safe_slot("crop_done_selected", self.on_crop_done_selected))
        right.addWidget(self.crop_done_grid, 1)
        del_crop_btn = QPushButton("刪除選取裁切圖")
        del_crop_btn.setObjectName("DangerButton")
        del_crop_btn.clicked.connect(self.safe_action("delete_selected_crop", self.delete_selected_crop))
        del_all_crop_btn = QPushButton("刪除所有裁切圖")
        del_all_crop_btn.setObjectName("DangerButton")
        del_all_crop_btn.clicked.connect(self.safe_action("delete_all_crops", self.delete_all_crops))
        del_btn_row = QHBoxLayout(); del_btn_row.setContentsMargins(0, 0, 0, 0); del_btn_row.setSpacing(8)
        del_btn_row.addWidget(del_crop_btn); del_btn_row.addWidget(del_all_crop_btn)
        right.addLayout(del_btn_row)
        self.crop_done_preview = ImagePreview("尚無 Step 4 輸入圖像")
        right.addWidget(self.crop_done_preview, 1)

        w1=QWidget(); w1.setLayout(left); w2=QWidget(); w2.setLayout(mid); w3=QWidget(); w3.setLayout(right)
        body.addWidget(w1, 1); body.addWidget(w2, 3); body.addWidget(w3, 1)
        lay.addLayout(body, 1)
        return page

    def page_regions(self) -> QWidget:
        lay = QVBoxLayout()
        lay.addWidget(self.header("Step 4｜ROI 與 Target Area 框選", ""))
        top = QHBoxLayout()
        self.region_roi_btn = QPushButton("框選 ROI"); self.region_roi_btn.setCheckable(True); self.region_roi_btn.setChecked(True)
        self.region_select_roi_btn = QPushButton("選取 ROI"); self.region_select_roi_btn.setCheckable(True)
        self.region_target_btn = QPushButton("框選 Target Area 矩形"); self.region_target_btn.setCheckable(True)
        self.region_target_poly_btn = QPushButton("直線繪製 Target Area"); self.region_target_poly_btn.setCheckable(True)
        self.region_select_target_btn = QPushButton("選取 Target Area"); self.region_select_target_btn.setCheckable(True)
        self.region_mode_group = QButtonGroup(self)
        self.region_mode_group.setExclusive(True)
        for b in [self.region_roi_btn, self.region_select_roi_btn, self.region_target_btn, self.region_target_poly_btn, self.region_select_target_btn]:
            self.region_mode_group.addButton(b)
        self.region_roi_btn.clicked.connect(self.safe_action("region_mode_roi", lambda: self.region_canvas.set_mode("roi")))
        self.region_select_roi_btn.clicked.connect(self.safe_action("region_mode_select_roi", lambda: self.region_canvas.set_mode("select_roi")))
        self.region_target_btn.clicked.connect(self.safe_action("region_mode_target_rect", lambda: self.region_canvas.set_mode("target_rect")))
        self.region_target_poly_btn.clicked.connect(self.safe_action("region_mode_target_poly", lambda: self.region_canvas.set_mode("target_poly")))
        self.region_select_target_btn.clicked.connect(self.safe_action("region_mode_select_target", lambda: self.region_canvas.set_mode("select_target")))
        self.region_clear_roi_btn = QPushButton("刪除全部 ROI"); self.region_clear_roi_btn.setObjectName("DangerButton"); self.region_clear_roi_btn.clicked.connect(self.safe_action("clear_all_rois", lambda: self.region_canvas.clear_rois()))
        self.region_delete_selected_roi_btn = QPushButton("刪除選取 ROI"); self.region_delete_selected_roi_btn.setObjectName("DangerButton"); self.region_delete_selected_roi_btn.clicked.connect(self.safe_action("delete_selected_roi", self.delete_selected_roi))
        self.region_clear_target_btn = QPushButton("刪除全部 Target Area"); self.region_clear_target_btn.setObjectName("DangerButton"); self.region_clear_target_btn.clicked.connect(self.safe_action("delete_target_area", self.delete_target_area))
        self.region_delete_selected_target_btn = QPushButton("刪除選取 Target Area"); self.region_delete_selected_target_btn.setObjectName("DangerButton"); self.region_delete_selected_target_btn.clicked.connect(self.safe_action("delete_selected_target_area", self.delete_selected_target_area))
        for b in [self.region_roi_btn, self.region_select_roi_btn, self.region_clear_roi_btn, self.region_delete_selected_roi_btn, self.region_target_btn, self.region_target_poly_btn, self.region_select_target_btn, self.region_clear_target_btn, self.region_delete_selected_target_btn]:
            top.addWidget(b)
        top.addStretch(1)
        lay.addLayout(top)
        body = QHBoxLayout()
        left = QVBoxLayout(); left.addWidget(QLabel("圖像檢視 / 切換")); self.region_thumb_grid = VerticalThumbGrid(); self.region_thumb_grid.setSelectionMode(QAbstractItemView.ExtendedSelection); self.region_thumb_grid.selected_path_changed.connect(self.safe_slot("region_select_crop_image", self.region_select_crop_image)); self.region_thumb_grid.selected_paths_changed.connect(self.safe_slot("region_selected_groups_changed", self.region_selected_groups_changed))
        self.region_selection_status = QLabel("目前選取：0 組（Step 5 決定輸入組合）"); self.region_selection_status.setObjectName("ProgressStatus"); left.addWidget(self.region_selection_status)
        region_thumb_wrap = QWidget(); region_thumb_wrap_lay = QHBoxLayout(region_thumb_wrap); region_thumb_wrap_lay.setContentsMargins(0, 0, 0, 0); region_thumb_wrap_lay.addStretch(1); region_thumb_wrap_lay.addWidget(self.region_thumb_grid); region_thumb_wrap_lay.addStretch(1); left.addWidget(region_thumb_wrap, 1)
        right = QVBoxLayout(); self.region_status = QLabel("尚未選擇圖像"); self.region_status.setObjectName("ProgressStatus"); right.addWidget(self.region_status)
        self.region_canvas = RoiTargetCanvas(); self.region_canvas.region_changed.connect(self.safe_action("auto_save_current_regions", self.auto_save_current_regions)); self.region_canvas.canvas_error.connect(self.safe_slot("region_canvas_error", lambda msg: QMessageBox.warning(self, "Target Area", msg))); right.addWidget(self.region_canvas, 1)
        wl=QWidget(); wl.setLayout(left); wr=QWidget(); wr.setLayout(right); body.addWidget(wl,1); body.addWidget(wr,4)
        lay.addLayout(body, 1)
        page = self.wrap_page(4, lay)
        self.setup_step4_shortcuts(page)
        return page

    def setup_step4_shortcuts(self, page: QWidget) -> None:
        """Attach Step 4 keyboard shortcuts to the Step 4 page only."""
        self.step4_shortcuts = []
        def add(seq: str, func: Callable[[], object], tip: str = "") -> None:
            sc = QShortcut(QKeySequence(seq), page)
            sc.setContext(Qt.WidgetWithChildrenShortcut)
            sc.activated.connect(self.safe_action(f"step4_shortcut_{seq}", func))
            self.step4_shortcuts.append(sc)
            if tip:
                try:
                    func_name = getattr(func, "__name__", "")
                except Exception:
                    func_name = ""

        add("R", lambda: self.region_roi_btn.click())
        add("S", lambda: self.region_select_roi_btn.click())
        add("A", lambda: self.region_canvas.clear_rois())
        add("D", self.delete_selected_roi)
        add("T", lambda: self.region_target_btn.click())
        add("L", lambda: self.region_target_poly_btn.click())
        add("Y", lambda: self.region_select_target_btn.click())
        add("G", self.delete_target_area)
        add("H", self.delete_selected_target_area)
        add("Down", lambda: self.step4_navigate_image(+1))
        add("Up", lambda: self.step4_navigate_image(-1))
        tooltips = {
            self.region_roi_btn: "快捷鍵 R",
            self.region_select_roi_btn: "快捷鍵 S",
            self.region_clear_roi_btn: "快捷鍵 A",
            self.region_delete_selected_roi_btn: "快捷鍵 D",
            self.region_target_btn: "快捷鍵 T",
            self.region_target_poly_btn: "快捷鍵 L",
            self.region_select_target_btn: "快捷鍵 Y",
            self.region_clear_target_btn: "快捷鍵 G",
            self.region_delete_selected_target_btn: "快捷鍵 H",
        }
        for btn, tip in tooltips.items():
            try:
                btn.setToolTip(tip)
            except Exception:
                pass

    def step4_navigate_image(self, delta: int) -> None:
        """Switch visible editing image with ↑/↓ without destroying multi-selection."""
        if getattr(self, "current_step", None) != 4 or not hasattr(self, "region_thumb_grid"):
            return
        grid = self.region_thumb_grid
        if grid.count() <= 0:
            return
        current_path = getattr(self.region_canvas, "image_path", None) if hasattr(self, "region_canvas") else None
        row = -1
        if current_path:
            cur = str(Path(current_path))
            for i in range(grid.count()):
                if str(grid.item(i).data(Qt.UserRole)) == cur:
                    row = i
                    break
        if row < 0:
            row = grid.currentRow() if grid.currentRow() >= 0 else 0
        nxt = max(0, min(grid.count() - 1, row + int(delta)))
        item = grid.item(nxt)
        if not item:
            return
        path = str(item.data(Qt.UserRole))
        # Keep the visible thumbnail selection synchronized with ↑/↓ navigation.
        # Step 4 selection is only for viewing/editing; final generation groups
        # are still selected in Step 5.
        blocker = QSignalBlocker(grid)
        try:
            grid.clearSelection()
            grid.setCurrentRow(nxt)
            item.setSelected(True)
            grid.scrollToItem(item, QListWidget.PositionAtCenter)
        finally:
            del blocker
        self.step4_selected_view_stems = [Path(path).stem]
        self.update_region_selection_status()
        self.region_select_crop_image(path)

    def page_prompt(self) -> QWidget:
        lay = QVBoxLayout()
        lay.addWidget(self.header("Step 5｜Prompt 編輯", ""))

        group_box = QGroupBox("引用組別")
        group_lay = QVBoxLayout(group_box)
        self.prompt_group_grid = PromptGroupList(); self.prompt_group_grid.group_changed.connect(self.safe_slot("prompt_group_selected", self.on_prompt_group_selected)); self.prompt_group_grid.group_selection_changed.connect(self.safe_slot("prompt_group_selection_changed", self.on_prompt_group_selection_changed))
        self.prompt_group_grid.setMaximumHeight(180)
        group_lay.addWidget(self.prompt_group_grid)
        self.prompt_selection_status = QLabel("已選定：0/16 組")
        self.prompt_selection_status.setObjectName("ProgressStatus")
        group_lay.addWidget(self.prompt_selection_status)
        lay.addWidget(group_box, 0)

        mode_box = QGroupBox("Prompt 來源設定"); mg = QGridLayout(mode_box); mg.setColumnStretch(1,1); mg.setColumnStretch(2,1)
        self.prompt_mode_combo = QComboBox(); self.prompt_mode_combo.addItems(["自訂 prompt", "使用模板"]); self.prompt_mode_combo.setCurrentText("使用模板" if self.state.prompt_mode == "template" else "自訂 prompt")
        self.prompt_template_label = QLabel("模板樣式"); self.prompt_template_combo = QComboBox(); self.prompt_template_combo.addItems(["瑕疵"]); self.prompt_template_combo.setCurrentText(self.state.prompt_template if self.state.prompt_template in ["瑕疵"] else "瑕疵")
        self.apply_template_btn = QPushButton("套用模板到輸入指令"); self.apply_template_btn.clicked.connect(self.safe_action("apply_prompt_template", self.apply_prompt_template))
        self.prompt_mode_combo.currentTextChanged.connect(self.safe_slot("prompt_mode_changed", lambda *_: self.on_prompt_mode_changed()))
        mg.addWidget(QLabel("Prompt 模式"),0,0); mg.addWidget(self.prompt_mode_combo,0,1,1,2)
        mg.addWidget(self.prompt_template_label,1,0); mg.addWidget(self.prompt_template_combo,1,1); mg.addWidget(self.apply_template_btn,1,2)
        lay.addWidget(mode_box, 0)

        # Only the 輸入指令 section remains below; it fills the rest of the page and
        # uses an enlarged font. The 實際傳送指令 preview section was removed.
        left = QGroupBox("輸入指令"); left_lay = QVBoxLayout(left)
        self.prompt_edit = QPlainTextEdit(); self.prompt_edit.setMinimumHeight(360); self.prompt_edit.setStyleSheet("font-size: 20px;"); self.prompt_edit.textChanged.connect(self.safe_action("prompt_text_changed", self.on_prompt_text_changed)); left_lay.addWidget(self.prompt_edit)
        lay.addWidget(left, 2)
        self.on_prompt_mode_changed()
        return self.wrap_page(5, lay)

    def page_model(self) -> QWidget:
        lay=QVBoxLayout(); lay.addWidget(self.header("Step 6｜模型與生成參數", "設定模型、品質、尺寸、輸出張數與輸出資料夾名稱。"))
        box=QGroupBox("OpenAI Image Parameters"); g=QGridLayout(box); g.setColumnStretch(1,1)
        self.model_combo=QComboBox(); self.model_combo.addItems(MODELS); self.model_combo.setCurrentText(self.state.model)
        self.quality_combo=QComboBox(); self.quality_combo.addItems(QUALITIES); self.quality_combo.setCurrentText(self.state.quality if self.state.quality in QUALITIES else "low")
        self.gpt2_size_mode_label=QLabel("尺寸模式"); self.gpt2_size_mode_combo=QComboBox(); self.gpt2_size_mode_combo.addItems(GPT2_SIZE_MODES); self.gpt2_size_mode_combo.setCurrentText("與原圖尺寸相同" if self.state.size == SAME_AS_ORIGINAL_SIZE else "自訂尺寸"); self.gpt2_size_mode_combo.setToolTip(GPT2_SIZE_LIMIT_HINT)
        self.size_label=QLabel("固定尺寸"); self.size_combo=QComboBox(); self.size_combo.addItems(FIXED_SIZES); self.size_combo.setCurrentText(self.state.size if self.state.size in FIXED_SIZES else "1024x1024")
        w,h=self.parse_size_or_default(self.state.size)
        self.width_label=QLabel("寬度"); self.width_edit=QLineEdit(str(w)); self.width_edit.setValidator(QIntValidator(0,9999,self))
        self.height_label=QLabel("高度"); self.height_edit=QLineEdit(str(h)); self.height_edit.setValidator(QIntValidator(0,9999,self))
        self.size_warning=QLabel(GPT2_SIZE_LIMIT_HINT)
        self.size_warning.setObjectName("DangerText")
        self.num_spin=ReliableSpinBox(); self.num_spin.setRange(1,100000); self.num_spin.setValue(self.state.num_outputs)
        self.run_name_edit=QLineEdit(self.state.run_name or "run"); self.run_name_edit.setPlaceholderText("預設：run，可自行設定")
        self.estimate_cost_btn=QPushButton("確認參數並估算本次成本")
        self.estimate_cost_btn.setObjectName("PrimaryButton")
        self.estimate_cost_btn.clicked.connect(self.safe_action("estimate_model_cost", self.estimate_model_cost))
        self.estimated_cost_label=QLabel("預估成本：尚未確認")
        self.estimated_cost_label.setObjectName("ProgressStatus")
        self.model_combo.currentTextChanged.connect(self.safe_slot("model_changed", lambda *_: (self.update_size_controls(), self.mark_dirty(6))))
        self.gpt2_size_mode_combo.currentTextChanged.connect(self.safe_slot("gpt2_size_mode_changed", lambda *_: (self.update_size_controls(), self.mark_dirty(6))))
        self.quality_combo.currentTextChanged.connect(self.safe_slot("quality_changed", lambda *_: self.mark_dirty(6)))
        self.size_combo.currentTextChanged.connect(self.safe_slot("fixed_size_changed", lambda *_: self.mark_dirty(6)))
        self.num_spin.valueChanged.connect(self.safe_slot("num_outputs_changed", lambda *_: self.mark_dirty(6)))
        for wdg in [self.width_edit,self.height_edit,self.model_combo,self.quality_combo,self.size_combo,self.gpt2_size_mode_combo,self.num_spin,self.run_name_edit]:
            try: wdg.textChanged.connect(self.safe_slot("model_param_changed", lambda *_: self.mark_dirty(6)))
            except Exception: pass
        g.addWidget(QLabel("模型"),0,0); g.addWidget(self.model_combo,0,1)
        g.addWidget(QLabel("品質"),1,0); g.addWidget(self.quality_combo,1,1)
        g.addWidget(self.gpt2_size_mode_label,2,0); g.addWidget(self.gpt2_size_mode_combo,2,1)
        g.addWidget(self.size_label,3,0); g.addWidget(self.size_combo,3,1)
        g.addWidget(self.width_label,4,0); g.addWidget(self.width_edit,4,1)
        g.addWidget(self.height_label,5,0); g.addWidget(self.height_edit,5,1)
        g.addWidget(self.size_warning,6,0,1,2)
        g.addWidget(QLabel("輸出張數"),7,0); g.addWidget(self.num_spin,7,1)
        g.addWidget(QLabel("輸出資料夾名稱"),8,0); g.addWidget(self.run_name_edit,8,1)
        g.addWidget(self.estimate_cost_btn,9,1)
        g.addWidget(self.estimated_cost_label,10,0,1,2)
        lay.addWidget(box); self.update_size_controls()
        return self.wrap_page(6, lay)

    def page_aggregate(self) -> QWidget:
        lay=QVBoxLayout(); lay.addWidget(self.header("Step 7｜Aggregate / 設定彙整確認", "確認 Class Name、圖像數量、ROI / target area、prompt 與模型參數，無誤後才能進入生成。"))
        self.aggregate_box=QPlainTextEdit(); self.aggregate_box.setReadOnly(True); self.aggregate_box.setMinimumHeight(760); self.aggregate_box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding); lay.addWidget(self.aggregate_box,10)
        return self.wrap_page(7, lay)

    def page_run(self) -> QWidget:
        lay=QVBoxLayout(); lay.addWidget(self.header("Step 8｜執行生成", "依照目前設定呼叫後端批次生成，並顯示即時 log 與進度。"))
        row=QHBoxLayout(); restart=QPushButton("重新開始生成"); restart.setObjectName("PrimaryButton"); restart.clicked.connect(self.safe_action("restart_generation", lambda: self.start_generation(resume=False))); resume=QPushButton("從上次進度開始生成"); resume.clicked.connect(self.safe_action("resume_generation", lambda: self.start_generation(resume=True))); stop=QPushButton("停止目前程序"); stop.clicked.connect(self.safe_action("stop_process", self.stop_process)); row.addWidget(restart); row.addWidget(resume); row.addWidget(stop); row.addStretch(1); lay.addLayout(row)
        self.generation_progress_bar=QProgressBar(); self.generation_progress_bar.setRange(0,1); self.generation_progress_bar.setValue(0); self.generation_progress_bar.setFormat("0/0 張")
        self.generation_progress_label=QLabel("尚未開始生成"); self.generation_progress_label.setObjectName("ProgressStatus")
        self.actual_cost_label=QLabel("本次實際成本：尚未完成生成")
        self.actual_cost_label.setObjectName("ProgressStatus")
        lay.addWidget(self.generation_progress_bar); lay.addWidget(self.generation_progress_label); lay.addWidget(self.actual_cost_label)
        self.command_preview=QPlainTextEdit(); self.command_preview.setReadOnly(True); self.command_preview.setMaximumHeight(110); lay.addWidget(self.command_preview)
        self.log_box=QPlainTextEdit(); self.log_box.setReadOnly(True); self.log_box.setMinimumHeight(560); self.log_box.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        try: self.log_box.setMaximumBlockCount(2500)
        except Exception: pass
        lay.addWidget(self.log_box,10)
        return self.wrap_page(8, lay)

    def page_export(self) -> QWidget:
        lay=QVBoxLayout(); lay.addWidget(self.header("Step 9｜Export / 輸出資料整理", "勾選要匯出的 runs，按『確認』預覽，再按『Export』選擇路徑匯出。"))

        box=QGroupBox("Export 規劃"); g=QGridLayout(box)
        g.setColumnStretch(0,0); g.setColumnStretch(1,1); g.setColumnStretch(2,0)
        self.export_runs_combo=CheckableRunCombo(); self.export_runs_combo.setMinimumWidth(360)
        confirm_btn=QPushButton("確認"); confirm_btn.setObjectName("PrimaryButton"); confirm_btn.setMinimumWidth(120)
        confirm_btn.clicked.connect(self.safe_action("preview_selected_runs", self.load_selected_runs_preview))
        export_btn=QPushButton("Export"); export_btn.setObjectName("SubmitButton"); export_btn.setMinimumWidth(120)
        export_btn.clicked.connect(self.safe_action("submit_step_9", lambda: self.submit_step(9)))
        self.export_run_btn=export_btn
        self.zip_check=QCheckBox("打包成 .zip"); self.zip_check.setChecked(True)
        # 確認 (top) and Export (bottom) are stacked vertically in the right column.
        g.addWidget(QLabel("匯出範圍"),0,0)
        g.addWidget(self.export_runs_combo,0,1)
        g.addWidget(confirm_btn,0,2)
        g.addWidget(self.zip_check,1,0,1,2)
        g.addWidget(export_btn,1,2)
        lay.addWidget(box,0)

        mid=QHBoxLayout(); mid.setSpacing(12)
        out_box=QGroupBox("欲輸出圖片")
        out_box.setMaximumWidth(360)
        out_box.setMinimumWidth(280)
        out_lay=QVBoxLayout(out_box); out_lay.setContentsMargins(8, 10, 8, 8)
        self.output_list=OutputThumbList(); self.output_list.selected_path_changed.connect(self.safe_action("preview_output_selection", self.preview_output_selection))
        out_lay.addWidget(self.output_list,1)
        mid.addWidget(out_box,0)
        self.output_preview=ImagePreview("尚無預覽")
        self.output_preview.setMinimumSize(900, 560)
        self.output_preview.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        mid.addWidget(self.output_preview,1)
        lay.addLayout(mid,1)
        page=self.wrap_page(9, lay)
        # The Export button lives in the section; register it so the generation
        # soft-lock (forbidden cursor) still applies to it like a footer submit.
        self.footer_buttons.setdefault(9, {})["submit"]=export_btn
        return page

    # ---------- style ----------
    def apply_style(self) -> None:
        self.setStyleSheet("""
            QWidget { font-family: 'Microsoft JhengHei', 'Noto Sans TC', Arial; font-size: 14px; }
            QMainWindow { background: #06111f; }
            QWidget { background-clip: border; }
            #Sidebar { background: #0b1220; border-right: 1px solid #1e293b; }
            #StepScroll, #StepScrollViewport, #StepContainer { background: #0b1220; border: 0px; }
            #StepScroll QScrollBar:vertical { background: #0b1220; width: 8px; margin: 0px; }
            #StepScroll QScrollBar::handle:vertical { background: #334155; border-radius: 4px; min-height: 24px; }
            #StepScroll QScrollBar::add-line:vertical, #StepScroll QScrollBar::sub-line:vertical { height: 0px; background: #0b1220; }
            #PageScroll { background: transparent; border: 0px; }
            #AppTitle { color: #e5e7eb; font-size: 24px; font-weight: 800; padding: 8px 0 6px 0 }
            #SideHint { color: #b8d4f5; background: #0f1a2e; padding: 8px; border-radius: 10px; }
            #StepButton { text-align: left; padding: 9px 12px; border-radius: 16px; color: #f8fafc; background: #111827; border: 2px solid #334155; margin: 2px 0; outline: none; }
            #StepButton[density="compact"] { padding: 7px 10px; border-radius: 14px; font-size: 13px; margin: 1px 0; }
            #StepButton[density="tiny"] { padding: 5px 9px; border-radius: 13px; font-size: 12px; margin: 1px 0; }
            #StepButton:hover { background: #18283d; border: 2px solid #38bdf8; margin-top: 2px; margin-bottom: 2px; }
            #StepButton[density="compact"]:hover, #StepButton[density="tiny"]:hover { margin-top: 1px; margin-bottom: 1px; }
            #StepButton[status="locked"] { color: #64748b; background: #111827; border: 2px solid #334155; }
            #StepButton[status="locked"]:hover { color: #64748b; background: #0f172a; border: 2px dashed #ef4444; }
            #StepButton[status="dirty"] { color: #ffffff; background: #3b111d; border: 2px solid #ff477e; }
            #StepButton[status="done"] { color: #ffffff; background: #07351f; border: 2px solid #00f5a0; }
            #StepButton[status="current_dirty"] { color: #ffffff; background: #421323; border: 3px solid #ff3f7f; }
            #StepButton[status="current_done"] { color: #ffffff; background: #0a3d2a; border: 3px solid #00f5a0; }
            #StatusBox { color: #d1d5db; background: #08101d; border: 1px solid #334155; border-radius: 12px; padding: 10px; }
            QStackedWidget#MainStack, QStackedWidget#MainStack > QWidget { background: #f8fafc; border-radius: 0px; }
            #Header { background: #eef4fb; border-radius: 18px; padding: 8px; }
            #ModeCard { background: #ffffff; border: 2px solid #cbd5e1; border-radius: 22px; }
            #ModeCard:hover { border: 3px solid #2563eb; background: #f8fbff; }
            #ModeBar { background:#eef4fb; border:1px solid #cbd5e1; border-radius:14px; }
            #PageTitle { color: #0f172a; font-size: 25px; font-weight: 800; }
            #PageSubtitle { color: #475569; }
            QGroupBox { color:#0f172a; font-weight: 700; border: 1px solid #cbd5e1; border-radius: 14px; margin-top: 12px; padding: 14px; background: #ffffff; }
            QGroupBox::title { subcontrol-origin: margin; left: 14px; padding: 0 6px; }
            QPushButton { background: #e2e8f0; border: 1px solid #cbd5e1; border-radius: 11px; padding: 9px 14px; min-height: 28px; outline: none; }
            QPushButton:hover { background: #dbeafe; border: 1px solid #60a5fa; margin-top: -1px; margin-bottom: 1px; }
            QPushButton:pressed { background: #bfdbfe; padding-top: 11px; padding-left: 16px; }
            QPushButton[softLocked="true"] { background: #e5e7eb; color: #94a3b8; border: 1px solid #cbd5e1; }
            QPushButton[softLocked="true"]:hover { background: #e5e7eb; border: 1px solid #cbd5e1; margin-top: 0px; margin-bottom: 0px; }
            QPushButton[softLocked="true"]:pressed { background: #e5e7eb; padding: 9px 14px; }
            #PrimaryButton { background: #2563eb; color: white; border: 1px solid #1d4ed8; font-weight: 700; }
            #DangerButton { background: #ef4444; color: white; border: 1px solid #b91c1c; font-weight: 800; }
            #SubmitButton { background: #059669; color: white; border: 1px solid #047857; font-weight: 800; }
            #HelpButton { background: #e11d48; color: white; border: 1px solid #be123c; font-weight: 800; }
            QLineEdit, QPlainTextEdit, QTextEdit, QComboBox, QSpinBox, QListWidget { border: 1px solid #cbd5e1; border-radius: 10px; padding: 6px; background: #ffffff; color: #0f172a; }
            QLineEdit, QComboBox, QSpinBox { min-height: 30px; }
            QSpinBox { padding-right: 34px; }
            QSpinBox::up-button { subcontrol-origin: border; subcontrol-position: top right; width: 28px; border-left: 1px solid #cbd5e1; border-top-right-radius: 10px; background: transparent; }
            QSpinBox::down-button { subcontrol-origin: border; subcontrol-position: bottom right; width: 28px; border-left: 1px solid #cbd5e1; border-bottom-right-radius: 10px; background: transparent; }
            QSpinBox::up-arrow, QSpinBox::down-arrow { width: 0px; height: 0px; }
            QSpinBox::up-button:hover, QSpinBox::down-button:hover { background: transparent; }
            #PreviewLabel { background: #e2e8f0; border: 1px dashed #94a3b8; border-radius: 14px; color: #64748b; }
            #DimensionLabel { color: #334155; background: #f1f5f9; border: 1px solid #cbd5e1; border-radius: 10px; padding: 6px; font-weight: 700; }
            #DangerText { color: #dc2626; font-weight: 800; }
            #ProgressStatus { color: #334155; font-weight: 700; padding-top: 4px; }
            QProgressBar { border: 1px solid #94a3b8; border-radius: 10px; background: #e2e8f0; text-align: center; min-height: 24px; color: #0f172a; font-weight: 800; }
            QProgressBar::chunk { background: #2563eb; border-radius: 9px; }
            #ProjectCard { background: #ffffff; border: 2px solid #cbd5e1; border-radius: 18px; padding: 0px; }
            #ProjectCard[mode="defect"] { border: 3px solid #10b981; }
            #ProjectCard[mode="food"] { border: 3px solid #2563eb; }
            #ProjectCard[mode="defect"]:hover { border: 3px solid #10b981; background: #f0fdf9; }
            #ProjectCard[mode="food"]:hover { border: 3px solid #2563eb; background: #f8fbff; }
            #ProjectCard[mode="defect"][selected="true"] { border: 4px solid #047857; }
            #ProjectCard[mode="food"][selected="true"] { border: 4px solid #1e40af; }
            #ActiveProjectLabel { background: #e0f2fe; color: #0f172a; border: 1px solid #93c5fd; border-radius: 12px; padding: 10px 14px; font-weight: 700; }
            #ProjectCardHeader { background: #a7f3d0; border-top-left-radius: 16px; border-top-right-radius: 16px; padding: 10px; }
            #ProjectCardHeader[mode="defect"] { background: #a7f3d0; }
            #ProjectCardHeader[mode="food"] { background: #bfdbfe; }
            #ProjectBadge { color: #0f766e; background: #ecfeff; border-radius: 10px; padding: 6px 12px; font-weight: 800; font-size: 15px; }
            #ProjectBadge[mode="defect"] { color: #047857; background: #ecfdf5; }
            #ProjectBadge[mode="food"] { color: #1d4ed8; background: #eff6ff; }
            #CardTitle { color:#0f172a; font-size:16px; font-weight:800; }
            #CardMeta { color:#334155; font-size:12px; padding: 4px 14px; line-height: 1.35; }
            #PromptGroupList { background:#ffffff; border:1px solid #cbd5e1; border-radius:14px; padding:8px; }
        """)

    # ---------- navigation ----------
    def max_accessible_step(self) -> int:
        if not self.state.project_id:
            return 0
        for i, done in enumerate(self.state.completed_steps):
            if not done:
                return i
        return STEP_COUNT - 1

    def is_generation_running(self) -> bool:
        return bool(
            getattr(self, "generation_ui_locked", False)
            or (
                self.current_process
                and self.current_process_label == "generation"
                and self.current_process.state() != QProcess.NotRunning
            )
        )

    def set_generation_ui_locked(self, locked: bool) -> None:
        """Lock/unlock all non-Step-8 navigation while image generation is pending/running."""
        self.generation_ui_locked = bool(locked)
        self.update_step_buttons()

    def update_step_buttons(self) -> None:
        max_access = self.max_accessible_step()
        generation_locked = self.is_generation_running()
        for btn in self.step_buttons:
            i = int(getattr(btn, "index", 0))
            display_i = int(getattr(btn, "display_index", STEP_DISPLAY_INDEX.get(i, i)))
            btn.setText((f"Step {display_i}\n{self.TR[f'step{i}']}"))
            locked = (i > max_access) or (generation_locked and i != 8)
            btn.set_locked(locked)
            if locked:
                status = "locked"
            elif i == self.current_step:
                status = "current_done" if self.state.completed_steps[i] and not self.dirty_steps[i] else "current_dirty"
            else:
                status = "done" if self.state.completed_steps[i] and not self.dirty_steps[i] else "dirty"
            if i == 0 and self.state.project_id and not locked:
                status = "done" if i != self.current_step else "current_done"
            btn.setProperty("status", status); btn.setChecked(i == self.current_step); btn.style().unpolish(btn); btn.style().polish(btn)
        # During generation, prevent accidental page changes or Submit/Back actions
        # outside Step 8. Keep Help available so the user can still read guidance.
        for step_idx, btns in getattr(self, "footer_buttons", {}).items():
            # During generation, Back and Submit must be blocked on every page,
            # including Step 8 itself.  They are kept enabled but soft-locked so
            # the forbidden cursor is visible while clicks/keyboard activation are
            # swallowed by eventFilter(). Help remains available.
            for name, fbtn in btns.items():
                soft_locked = bool(generation_locked and name in {"back", "submit", "use_original_all"})
                fbtn.setProperty("softLocked", soft_locked)
                fbtn.setEnabled(True)
                fbtn.setCursor(StepButton.forbidden_cursor() if soft_locked else QCursor(Qt.PointingHandCursor))
                fbtn.style().unpolish(fbtn); fbtn.style().polish(fbtn)

    def goto_step(self, idx: int) -> None:
        if idx == 1:
            idx = 2 if self.state.project_id else 0
        if idx not in STEP_TO_STACK_INDEX:
            return
        if self.is_generation_running() and idx != 8:
            QMessageBox.warning(self, "Generating", "生成圖片期間已鎖定其他 Step，請等待完成或停止目前程序。")
            return
        if idx > self.max_accessible_step():
            QMessageBox.warning(self, "Locked", "請先完成目前 Step 並按下 Submit。")
            return
        prev = getattr(self, "current_step", None)
        if prev is not None and prev != idx:
            self.clear_step_preview(prev)
        self.current_step = idx
        self.state.last_active_step = int(idx)
        self.log_event(f"enter Step {STEP_DISPLAY_INDEX.get(idx, idx)} (internal {idx})")
        self.save_state()
        self.stack.setCurrentIndex(STEP_TO_STACK_INDEX[idx])
        self.update_step_buttons()
        self.refresh_all(light=True)
        if idx == 3:
            self.prepare_step3_entry()

    def go_back(self) -> None:
        if self.is_generation_running():
            QMessageBox.warning(self, "Generating", "生成圖片期間 Back 已鎖定，請等待完成或停止目前程序。")
            return
        if self.current_step in VISIBLE_STEPS:
            pos = VISIBLE_STEPS.index(self.current_step)
            if pos > 0:
                self.goto_step(VISIBLE_STEPS[pos - 1])

    def mark_dirty(self, idx: int) -> None:
        if getattr(self, "_suspend_dirty_tracking", False):
            return
        if idx < 0 or idx >= STEP_COUNT: return
        self.dirty_steps[idx] = True
        if self.state.completed_steps[idx]:
            for j in range(idx, STEP_COUNT):
                self.state.completed_steps[j] = False; self.dirty_steps[j] = True
        self.update_step_buttons()

    def complete_step(self, idx: int) -> None:
        self.state.completed_steps[idx] = True; self.dirty_steps[idx] = False; self.save_state(); self.update_step_buttons()
        if idx in VISIBLE_STEPS:
            pos = VISIBLE_STEPS.index(idx)
            if pos < len(VISIBLE_STEPS) - 1:
                self.goto_step(VISIBLE_STEPS[pos + 1])
        elif idx < STEP_COUNT - 1:
            self.goto_step(idx + 1)

    def _log_ui_exception(self, context: str, exc: BaseException) -> None:
        try:
            log_path = self.app_log_path()
            with log_path.open("a", encoding="utf-8") as f:
                f.write(f"\n===== {datetime.now().isoformat(timespec='seconds')} | {context} =====\n")
                traceback.print_exception(type(exc), exc, exc.__traceback__, file=f)
        except Exception:
            pass

    def safe_action(self, context: str, func: Callable[[], object]) -> Callable[..., object | None]:
        """Wrap no-argument UI callbacks so signal errors are logged instead of escaping Qt."""
        def _wrapped(*_args, **_kwargs):
            try:
                return func()
            except Exception as exc:
                self._log_ui_exception(context, exc)
                try:
                    QMessageBox.critical(self, "UI action failed", f"執行按鍵動作時發生錯誤，程式已阻止視窗直接關閉。\n\n{type(exc).__name__}: {exc}")
                except Exception:
                    pass
                return None
        return _wrapped

    def safe_slot(self, context: str, func: Callable[..., object]) -> Callable[..., object | None]:
        """Wrap Qt slots that receive signal arguments such as text, path, or checked."""
        def _wrapped(*args, **kwargs):
            try:
                return func(*args, **kwargs)
            except Exception as exc:
                self._log_ui_exception(context, exc)
                try:
                    QMessageBox.critical(self, "UI signal failed", f"執行 UI 事件時發生錯誤，程式已阻止視窗直接關閉。\n\n{type(exc).__name__}: {exc}")
                except Exception:
                    pass
                return None
        return _wrapped

    def submit_step(self, idx: int) -> None:
        if self.is_generation_running():
            QMessageBox.warning(self, "Generating", "生成圖片期間 Submit / Save Step 已鎖定，請等待完成或停止目前程序。")
            return
        funcs = [self.submit_home, self.submit_project, self.submit_upload, self.submit_crop, self.submit_regions, self.submit_prompt, self.submit_model, self.submit_aggregate, self.submit_run, self.submit_export]
        disp = STEP_DISPLAY_INDEX.get(idx, idx)
        self.log_event(f"Step {disp} (internal {idx}) submit pressed")
        try:
            ok = bool(funcs[idx]())
            if ok:
                self.complete_step(idx)
            self.log_event(f"Step {disp} submit result={ok}")
        except Exception as exc:
            self._log_ui_exception(f"submit_step_{idx}", exc)
            QMessageBox.critical(self, "Submit failed", f"完成 Step {idx} 時發生錯誤，程式已阻止視窗直接關閉。\n\n{type(exc).__name__}: {exc}")

    def show_help(self, step_idx: int) -> None:
        rows = self.HELP.get(step_idx, [])
        html = ["<div style='font-size:14px;'>"]
        for text, important in rows:
            if important:
                html.append(f"<p style='background:#fee2e2;color:#b91c1c;padding:8px;border-radius:6px;'><b>{text}</b></p>")
            else:
                html.append(f"<p>{text}</p>")
        html.append("</div>")
        QMessageBox.information(self, "HELP", "".join(html))

    def clear_step_preview(self, step_idx: int) -> None:
        if step_idx == 2 and hasattr(self, "upload_preview"):
            self.upload_preview.clear("尚無預覽，請選取圖片或直接拖曳圖片到此處。")
            self.upload_list.clearSelection()
        elif step_idx == 3 and hasattr(self, "crop_canvas"):
            self.crop_canvas.set_image(None); self.crop_done_preview.clear("尚無裁切完成圖")
        elif step_idx == 9 and hasattr(self, "output_preview"):
            self.output_preview.clear("尚無預覽"); self.output_list.clearSelection()

    # ---------- project actions ----------
    def all_project_names(self, exclude_id: str = "") -> set[str]:
        names: set[str] = set()
        for item in self.load_combined_projects():
            if str(item.get("id", "")) != exclude_id:
                names.add(str(item.get("name", "")).strip().casefold())
        return {n for n in names if n}

    def all_project_ids(self, exclude_id: str = "") -> set[str]:
        ids: set[str] = set()
        for item in self.load_combined_projects():
            item_id = str(item.get("id", "")).strip()
            if item_id and item_id != exclude_id:
                ids.add(item_id)
        return ids

    def show_project_error(self, text: str) -> None:
        if hasattr(self, "project_error"):
            self.project_error.setText(text)
            self.project_error.setVisible(bool(text))
        if text:
            self.status_label.setText("Status: project name error")

    def update_home_current_project_label(self) -> None:
        if not hasattr(self, "home_current_project_label"):
            return
        active_name = (self.state.project_name or self.state.project_id or "").strip()
        selected_pid = getattr(self, "selected_project_card_id", "")
        selected_name = ""
        if selected_pid:
            for proj in self.load_combined_projects():
                if str(proj.get("id", "")) == selected_pid:
                    selected_name = str(proj.get("name", selected_pid))
                    break
        if active_name and selected_name and selected_name != active_name:
            text = f"目前開啟專案：{active_name}　｜　目前選取卡片：{selected_name}"
        elif active_name:
            text = f"目前開啟專案：{active_name}"
        elif selected_name:
            text = f"目前尚未開啟專案　｜　目前選取卡片：{selected_name}"
        else:
            text = "目前尚未開啟專案。請新增或選擇既有專案。"
        self.home_current_project_label.setText(text)

    def refresh_project_list(self) -> None:
        if not hasattr(self, "project_cards_layout"):
            return
        sb = None
        prev = 0
        if hasattr(self, "project_cards_scroll"):
            sb = self.project_cards_scroll.verticalScrollBar()
            prev = sb.value()
        self._project_cards = []
        while self.project_cards_layout.count():
            item = self.project_cards_layout.takeAt(0)
            w = item.widget()
            if w:
                w.hide()
                w.setParent(None)
                w.deleteLater()
        projects = self.load_combined_projects()
        if not projects:
            empty = QLabel("尚無專案。請按「新增專案」開始。")
            empty.setObjectName("ProgressStatus")
            empty.setAlignment(Qt.AlignCenter)
            self.project_cards_layout.addWidget(empty, 0, 0, 1, 3)
        cols = 3
        for n, proj in enumerate(projects):
            card = ProjectCard(proj)
            card.selected.connect(self.safe_slot("select_project_card", self.select_project_card))
            card.open_requested.connect(self.safe_slot("open_project_by_id", self.open_project_by_id))
            card.copy_requested.connect(self.safe_slot("copy_project_by_id", self.copy_project_by_id))
            card.rename_requested.connect(self.safe_slot("rename_project_by_id", self.rename_project_by_id))
            card.delete_requested.connect(self.safe_slot("delete_project_by_id", self.delete_project_by_id))
            card.set_selected(proj.get("id") == getattr(self, "selected_project_card_id", ""))
            self.project_cards_layout.addWidget(card, n // cols, n % cols)
            self._project_cards.append(card)
        self.project_cards_layout.setRowStretch((len(projects) + cols - 1) // cols, 1)
        self.project_cards_container.updateGeometry()
        self.project_cards_container.update()
        app = QApplication.instance()
        if app is not None:
            app.processEvents()
        self.update_home_current_project_label()
        self.preview_project_summary()
        if sb is not None:
            sb.setValue(prev)
            QTimer.singleShot(0, lambda: sb.setValue(prev))

    def select_project_card(self, pid: str) -> None:
        # A single click on a project card now makes that project the active one.
        # Earlier versions only previewed the card, which made Homepage show a
        # different project than the Step sidebar/state that would actually run.
        if not pid:
            return
        if pid != getattr(self.state, "project_id", ""):
            self.open_project_by_id(pid)
            return
        self.selected_project_card_id = pid
        for card in getattr(self, "_project_cards", []):
            card.set_selected(card.pid == pid)
        self.update_home_current_project_label()
        self.preview_project_summary()

    def preview_project_summary(self) -> None:
        if not hasattr(self, "project_summary"):
            return
        pid = getattr(self, "selected_project_card_id", "") or (self.state.project_id if self.state.saved_project else "")
        if not pid:
            self.project_summary.setText("尚未選擇專案。")
            return
        rec = self.find_project_record(pid)
        if rec is None:
            self.project_summary.setText("此專案尚無設定檔。")
            return
        self.project_summary.setText(
            f"專案名稱：{rec.get('name','-')}\n"
            f"Class Name：{rec.get('class_name','-')}\n"
            f"Create：{str(rec.get('created_at') or '-').replace('T',' ')}\n"
            f"模型：{rec.get('model','-')}\n"
            f"品質：{rec.get('quality','-')}\n"
            f"輸出張數：{rec.get('num_outputs','-')}\n"
            f"狀態：可開啟後繼續編輯或重新執行。"
        )

    def prompt_new_project_settings(self) -> tuple[str, str, str] | None:
        dialog = QDialog(self)
        dialog.setWindowTitle("Create Project")
        form = QFormLayout(dialog)
        project_edit = QLineEdit()
        object_edit = QLineEdit()
        mode_combo = QComboBox()
        for key, label in MODE_OPTIONS:
            mode_combo.addItem(label, key)
        mode_combo.setCurrentIndex(0 if APP_MODE == "defect" else 1)
        project_edit.setPlaceholderText("專案名稱")
        object_edit.setPlaceholderText("例如：fried_chicken、burr")
        form.addRow("專案名稱", project_edit)
        form.addRow("生成圖片物件的名稱", object_edit)
        form.addRow("模式", mode_combo)
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(dialog.accept)
        buttons.rejected.connect(dialog.reject)
        form.addWidget(buttons)
        project_edit.setFocus()
        if dialog.exec() != QDialog.DialogCode.Accepted:
            return None
        selected_mode = str(mode_combo.currentData() or APP_MODE)
        # 生成圖片物件的名稱 becomes the class name: it is the {class_name} prompt
        # template variable used in Step 5 (Prompt 編輯) for both modes and the
        # internal data/config folder name. Falls back to the project name if blank.
        return project_edit.text().strip(), object_edit.text().strip(), selected_mode

    def create_project_with_values(self, name: str, class_name: str, selected_mode: str | None = None) -> None:
        selected_mode = (selected_mode or APP_MODE).lower()
        if selected_mode not in {"defect", "food"}:
            selected_mode = APP_MODE
        if selected_mode != APP_MODE:
            callback = getattr(self, "_mode_switch_callback", None)
            if callable(callback):
                callback(selected_mode, name, class_name)
                return
            QMessageBox.information(self, "Switch mode", f"請從根目錄 launch_ui.py 啟動整合版，才能直接建立 {'Food' if selected_mode == 'food' else 'Defect'} 專案。")
            return
        if not name.strip():
            return
        normalized = name.strip().casefold()
        if normalized in self.all_project_names():
            self.show_project_error("專案名稱不可重複，請重新命名。")
            QMessageBox.warning(self, "專案名稱重複", "專案名稱不可重複，請重新命名。")
            return
        self.show_project_error("")
        pid = make_project_id(name)
        if pid in self.all_project_ids() or (self.projects_root / pid).exists():
            self.show_project_error("Project folder id already exists; please choose another project name.")
            QMessageBox.warning(self, "Project folder exists", "This project name maps to an existing project folder. Please choose another name.")
            return
        class_default = sanitize_name(class_name or name)
        self.state = UIState(project_id=pid, project_name=name.strip(), created_at=datetime.now().isoformat(timespec="seconds"), class_name=class_default, project_mode=APP_MODE, saved_project=True)
        self.dirty_steps = [False] + [True] * (STEP_COUNT - 1)
        self.state.completed_steps[0] = True
        self.state.completed_steps[1] = True
        self.dirty_steps[1] = False
        self.selected_project_card_id = self.state.project_id
        self.save_state(); self.log_event(f"project created: {self.state.project_id}"); self.load_state_to_widgets(); self.clear_all_visual_previews(); self.refresh_project_list(); self.status_label.setText("Status: project created"); self.goto_step(2)

    def create_project(self) -> None:
        settings = self.prompt_new_project_settings()
        if settings is None:
            return
        name, class_name, selected_mode = settings
        self.create_project_with_values(name, class_name, selected_mode)

    def selected_project_id(self) -> str:
        return getattr(self, "selected_project_card_id", "")

    def open_selected_project(self) -> None:
        self.open_project_by_id(self.selected_project_id())

    def open_project_by_id(self, pid: str) -> None:
        if not pid:
            QMessageBox.warning(self, "Missing", "請先選擇專案。")
            return
        target_mode = self.project_mode_for_id(pid)
        if target_mode != APP_MODE:
            callback = getattr(self, "_mode_open_callback", None)
            if callable(callback):
                callback(target_mode, pid)
                return
            QMessageBox.information(self, "Switch mode", f"請從根目錄 launch_ui.py 啟動整合版，才能開啟 {'Food' if target_mode == 'food' else 'Defect'} 專案。")
            return
        base = self.record_state_dict(pid)
        if base is None:
            QMessageBox.warning(self, "Missing", "找不到專案。")
            return
        if len(base.get("completed_steps", [])) != STEP_COUNT:
            done = [True] + [False] * (STEP_COUNT - 1)
            old = base.get("completed_steps", [])
            for i, v in enumerate(old[:STEP_COUNT]):
                done[i] = bool(v)
            base["completed_steps"] = done
        self.state = UIState(**base)
        self.state.saved_project = True
        # project_state.json is gone; restore the saved prompt from disk if present.
        try:
            pp = self.prompt_path()
            if pp.exists():
                self.state.prompt_input = pp.read_text(encoding="utf-8", errors="ignore").strip()
        except Exception:
            pass
        self.selected_project_card_id = pid
        self.reconcile_completed_steps_for_loaded_project(self.state)
        self.dirty_steps = [False if d else True for d in self.state.completed_steps]
        # Opened projects always land on Homepage, while preserving each project's
        # completed-step status and saved artifacts. Programmatic widget refresh is
        # dirty-suppressed so Step 5/6 cannot become red merely from setText()/selection signals.
        self.current_step = 0
        self.state.last_active_step = 0
        self._suspend_dirty_tracking = True
        try:
            self.save_state()
            self.load_state_to_widgets()
            self.clear_all_visual_previews()
            if hasattr(self, "stack"):
                self.stack.setCurrentIndex(0)
            self.refresh_all()
        finally:
            self._suspend_dirty_tracking = False
        self.reconcile_completed_steps_for_loaded_project(self.state)
        self.dirty_steps = [False if d else True for d in self.state.completed_steps]
        self.save_state()
        self.log_event(f"project opened: {self.state.project_id}")
        self.status_label.setText("Status: project opened - returned to homepage")
        self.update_step_buttons()

    def rename_selected_project(self) -> None:
        self.rename_project_by_id(self.selected_project_id())

    def rename_project_by_id(self, pid: str) -> None:
        """Rename a project card and its saved project folder safely.

        The class name and generated artifacts are intentionally left unchanged;
        this action only changes the project identity/name shown on Homepage and
        the saved project folder id.  The rename is blocked during generation and
        when it would duplicate another project's display name or folder id.
        """
        if not pid:
            QMessageBox.warning(self, "Missing", "請先選擇專案。")
            return
        if self.is_generation_running():
            QMessageBox.warning(self, "Generating", "生成圖片期間不可重新命名專案，請等待完成或停止目前程序。")
            return
        target_mode = self.project_mode_for_id(pid)
        if target_mode != APP_MODE:
            callback = getattr(self, "_mode_action_callback", None)
            if callable(callback):
                callback("rename", target_mode, pid)
                return
            QMessageBox.information(self, "Switch mode", "請先開啟該模式後再重新命名專案。")
            return
        rec = self.find_project_record(pid)
        project_dir = self.projects_root / sanitize_name(pid)
        if rec is None or not project_dir.exists():
            QMessageBox.warning(self, "Missing", "找不到要重新命名的專案。")
            return
        data = dict(rec)
        old_name = str(rec.get("name") or pid).strip()
        new_name, ok = QInputDialog.getText(self, "Rename Project", "New project name", text=old_name)
        if not ok:
            return
        new_name = new_name.strip()
        if not new_name:
            QMessageBox.warning(self, "Missing", "專案名稱不可空白。")
            return
        if new_name.casefold() == old_name.casefold() and make_project_id(new_name) == pid:
            return
        if new_name.casefold() in self.all_project_names(exclude_id=pid):
            QMessageBox.warning(self, "專案名稱重複", "專案名稱不可重複，請重新命名。")
            return
        new_id = make_project_id(new_name)
        if not new_id:
            QMessageBox.warning(self, "Invalid", "專案名稱無法轉成有效資料夾名稱，請重新命名。")
            return
        if new_id in self.all_project_ids(exclude_id=pid):
            QMessageBox.warning(self, "Project folder exists", "This project name maps to another existing project folder. Please choose another name.")
            return
        new_dir = self.projects_root / new_id
        if new_id != pid and new_dir.exists():
            QMessageBox.warning(self, "資料夾已存在", f"專案資料夾已存在：{new_id}\n請使用其他專案名稱。")
            return
        now_iso = datetime.now().isoformat(timespec="seconds")
        try:
            if new_id != pid:
                project_dir.rename(new_dir)
            data["id"] = new_id
            data["name"] = new_name
            data["updated_at"] = now_iso
            data.setdefault("project_mode", APP_MODE)
            idx = self.load_index()
            projects = []
            updated = False
            for item in idx.get("projects", []) or []:
                if str(item.get("id") or "") == pid:
                    item = dict(item)
                    item["id"] = new_id
                    item["name"] = new_name
                    item["updated_at"] = now_iso
                    item["mode"] = APP_MODE
                    updated = True
                projects.append(item)
            if not updated:
                projects.append({
                    "id": new_id,
                    "name": new_name,
                    "class_name": data.get("class_name", "-"),
                    "mode": APP_MODE,
                    "model": data.get("model", "-"),
                    "quality": data.get("quality", "-"),
                    "num_outputs": data.get("num_outputs", 1),
                    "created_at": data.get("created_at", now_iso),
                    "updated_at": now_iso,
                })
            idx["projects"] = sorted(projects, key=lambda p: p.get("updated_at", ""), reverse=True)
            if idx.get("active_project_id") == pid:
                idx["active_project_id"] = new_id
            self.save_index(idx)
        except Exception as exc:
            traceback.print_exc()
            QMessageBox.critical(self, "Rename failed", f"重新命名失敗：{type(exc).__name__}: {exc}")
            return
        if getattr(self, "selected_project_card_id", "") == pid:
            self.selected_project_card_id = new_id
        if self.state.project_id == pid:
            self.state.project_id = new_id
            self.state.project_name = new_name
            self.selected_project_card_id = new_id
            self.save_state()
        self.refresh_project_list()
        self.status_label.setText(f"Status: project renamed to {new_name}")
        QMessageBox.information(self, "Done", f"專案已重新命名為：{new_name}")

    def copy_selected_project(self) -> None:
        self.copy_project_by_id(self.selected_project_id())

    def make_duplicate_project_name(self, source_name: str) -> str:
        """Return a unique copy name without asking the user to type one."""
        base_name = (source_name.strip() or "project") + "_copy"
        existing = self.all_project_names()
        candidate = base_name
        n = 2
        while True:
            candidate_id = make_project_id(candidate)
            if candidate.strip().casefold() not in existing and not (self.projects_root / candidate_id).exists():
                return candidate
            candidate = f"{base_name}{n}"
            n += 1

    def copy_class_workspace(self, src_class: str, dst_class: str) -> None:
        # Project folders are copied wholesale under project/<project_name>/,
        # so no root-level data/config/runs duplication is needed.
        return

    def copy_project_by_id(self, pid: str) -> None:
        if not pid:
            QMessageBox.warning(self, "Missing", "請先選擇專案。")
            return
        target_mode = self.project_mode_for_id(pid)
        if target_mode != APP_MODE:
            callback = getattr(self, "_mode_action_callback", None)
            if callable(callback):
                callback("copy", target_mode, pid)
                return
            QMessageBox.information(self, "Switch mode", "請先開啟該模式後再複製專案。")
            return
        src = self.projects_root / pid
        rec = self.find_project_record(pid)
        if not src.exists() or rec is None:
            QMessageBox.warning(self, "Missing", "找不到要複製的專案。")
            return
        source_name = str(rec.get("name") or pid)
        new_name = self.make_duplicate_project_name(source_name)
        new_id = make_project_id(new_name)
        dst = self.projects_root / new_id
        # A duplicate copies only the uploaded inputs (data/), never the generated
        # outputs (runs/ and exports/), so it owns an independent run counter that
        # starts again at run1.
        shutil.copytree(src, dst, ignore=shutil.ignore_patterns("runs", "exports"))

        src_class_raw = rec.get("class_name")
        src_class = sanitize_name(str(src_class_raw) if src_class_raw not in (None, "", "-") else source_name)
        now_iso = datetime.now().isoformat(timespec="seconds")
        idx = self.load_index()
        projects = [p for p in idx.get("projects", []) if p.get("id") != new_id]
        projects.append({
            "id": new_id,
            "name": new_name,
            "class_name": src_class,
            "mode": APP_MODE,
            "model": rec.get("model", "-"),
            "quality": rec.get("quality", "-"),
            "num_outputs": rec.get("num_outputs", 1),
            "created_at": now_iso,
            "updated_at": now_iso,
        })
        idx["projects"] = sorted(projects, key=lambda p: p.get("updated_at", ""), reverse=True)
        self.save_index(idx)
        self.selected_project_card_id = new_id
        self.refresh_project_list()
        QMessageBox.information(self, "Done", f"專案已複製為：{new_name}")

    def delete_selected_project(self) -> None:
        self.delete_project_by_id(self.selected_project_id())

    def delete_project_by_id(self, pid: str) -> None:
        if not pid:
            QMessageBox.warning(self, "Missing", "請先選擇專案。")
            return
        target_mode = self.project_mode_for_id(pid)
        if target_mode != APP_MODE:
            callback = getattr(self, "_mode_action_callback", None)
            if callable(callback):
                callback("delete", target_mode, pid)
                return
            QMessageBox.information(self, "Switch mode", "請先開啟該模式後再刪除專案。")
            return
        box = QMessageBox(self); box.setIcon(QMessageBox.Icon.Warning); box.setWindowTitle("確認刪除專案"); box.setText("確定要刪除此專案嗎？"); box.setInformativeText("此操作會刪除該專案保存的設定與專案資料，無法復原。")
        yes = box.addButton("確認", QMessageBox.ButtonRole.AcceptRole); box.addButton("取消", QMessageBox.ButtonRole.RejectRole); box.exec()
        if box.clickedButton() != yes:
            return
        shutil.rmtree(self.projects_root / pid, ignore_errors=True)
        idx = self.load_index(); idx["projects"] = [p for p in idx.get("projects", []) if p.get("id") != pid]
        if idx.get("active_project_id") == pid:
            idx["active_project_id"] = ""
        self.save_index(idx)
        if getattr(self, "selected_project_card_id", "") == pid:
            self.selected_project_card_id = ""
        if self.state.project_id == pid:
            self.state = UIState(); self.dirty_steps = [True] * STEP_COUNT; self.load_state_to_widgets(); self.goto_step(0)
        self.refresh_project_list(); QMessageBox.information(self, "Done", "專案已刪除。")

    def save_current_project_to_home(self) -> None:
        if not self.state.project_id:
            QMessageBox.warning(self, "Missing", "目前沒有可儲存的專案。")
            return
        if self.state.project_name.strip().casefold() in self.all_project_names(exclude_id=self.state.project_id):
            self.show_project_error("專案名稱不可重複，請重新命名。")
            QMessageBox.warning(self, "專案名稱重複", "專案名稱不可重複，請重新命名。")
            return
        self.update_state_from_widgets()
        self.state.saved_project = True
        self.save_state(); self.refresh_project_list(); QMessageBox.information(self, "Done", "專案已儲存，之後可從 Homepage 開啟。")

    # ---------- widget/state ----------
    def _set_widget_value_blocked(self, widget, setter_name: str, value) -> None:
        """Set a widget value without firing connected dirty/refresh slots."""
        try:
            blocker = QSignalBlocker(widget)
            if value is None:
                getattr(widget, setter_name)()
            else:
                getattr(widget, setter_name)(value)
            del blocker
        except Exception:
            if value is None:
                getattr(widget, setter_name)()
            else:
                getattr(widget, setter_name)(value)

    def clear_all_visual_previews(self) -> None:
        """Clear UI-only previews when opening/creating a project to prevent stale images."""
        try:
            if hasattr(self, "upload_preview"):
                self.upload_preview.clear("尚無預覽，請選取圖片或直接拖曳圖片到此處。")
            if hasattr(self, "crop_canvas"):
                self.crop_canvas.set_image(None)
            if hasattr(self, "crop_done_preview"):
                self.crop_done_preview.clear("尚無 Step 4 輸入圖像")
            if hasattr(self, "region_canvas"):
                self.region_canvas.set_image(None)
            if hasattr(self, "region_status"):
                self.region_status.setText("尚未選擇圖像")
            if hasattr(self, "output_preview"):
                self.output_preview.clear("尚無預覽")
            if hasattr(self, "log_box"):
                self.log_box.clear()
            if hasattr(self, "command_preview"):
                self.command_preview.clear()
        except Exception:
            pass

    def load_state_to_widgets(self) -> None:
        if hasattr(self, "class_edit"):
            self._set_widget_value_blocked(self.class_edit, "setText", self.state.class_name)
        if hasattr(self, "quality_combo"):
            self._set_widget_value_blocked(self.quality_combo, "setCurrentText", self.state.quality)
        if hasattr(self, "model_combo"):
            self._set_widget_value_blocked(self.model_combo, "setCurrentText", self.state.model)
        if hasattr(self, "gpt2_size_mode_combo"):
            self._set_widget_value_blocked(self.gpt2_size_mode_combo, "setCurrentText", "與原圖尺寸相同" if self.state.size == SAME_AS_ORIGINAL_SIZE else "自訂尺寸")
            self.update_size_controls()
        if hasattr(self, "num_spin"):
            self._set_widget_value_blocked(self.num_spin, "setValue", self.state.num_outputs)
        if hasattr(self, "run_name_edit"):
            self._set_widget_value_blocked(self.run_name_edit, "setText", self.state.run_name or "run")
        if hasattr(self, "prompt_mode_combo"):
            self._set_widget_value_blocked(self.prompt_mode_combo, "setCurrentText", "使用模板" if self.state.prompt_mode == "template" else "自訂 prompt")
        if hasattr(self, "prompt_edit"):
            self._set_widget_value_blocked(self.prompt_edit, "setPlainText", self.state.prompt_input or "")
            self.update_actual_prompt_preview()
        # Re-sync the shared API key field from the repo-root .env so a key saved
        # earlier (in this or the other mode) shows as already saved instead of
        # prompting for re-entry when a project is created/opened.
        self.refresh_api_placeholder()
        if hasattr(self, "home_api_edit"):
            self._set_widget_value_blocked(self.home_api_edit, "clear", None)

    def update_state_from_widgets(self) -> None:
        if hasattr(self, "class_edit"):
            self.state.class_name = sanitize_name(self.class_edit.text())
        if hasattr(self, "crop_w_edit"):
            try: self.state.crop_width = int(self.crop_w_edit.text())
            except Exception: pass
            try: self.state.crop_height = int(self.crop_h_edit.text())
            except Exception: pass
        if hasattr(self, "prompt_edit"):
            self.state.prompt_input = self.prompt_edit.toPlainText()
        if hasattr(self, "model_combo"):
            self.state.model = self.model_combo.currentText(); self.state.quality = self.quality_combo.currentText(); self.state.size = self.current_size(); self.state.num_outputs = self.num_spin.value(); self.state.run_name = self.run_name_edit.text().strip(); self.state.dry_run = False

    def first_unfinished_step(self) -> int:
        for i, done in enumerate(self.state.completed_steps):
            if not done: return i
        return STEP_COUNT - 1

    # ---------- API key ----------
    def env_path(self) -> Path:
        # The shared OPENAI_API_KEY is stored once at the repository root (the
        # level that contains the modes/ folder), not duplicated under each mode.
        # self.root is modes/<mode>, so the repo root is its grandparent.
        return self.root.parent.parent / ".env"

    def read_env_key(self) -> str:
        # Only a key the user explicitly saved through the UI (written to .env) counts
        # as "saved". Do NOT fall back to the OS OPENAI_API_KEY environment variable —
        # otherwise a brand-new project would show a key as already saved before the
        # user has entered one. The field must start empty until the user saves a key.
        env = self.env_path()
        if env.exists():
            for line in env.read_text(encoding="utf-8", errors="ignore").splitlines():
                if line.startswith("OPENAI_API_KEY="):
                    return line.split("=", 1)[1].strip()
        return ""

    def masked_key(self, key: str) -> str:
        """Return a stable, non-sensitive API key preview.

        After the user saves an API key, Step 1 should show the real first
        three and last three characters so the user can confirm which key is
        currently stored, without exposing the full secret.  Keep this as
        placeholder text only; never write the masked preview back to .env.
        """
        if not key:
            return "第一次使用請輸入 OpenAI API Key"
        clean_key = key.strip()
        if len(clean_key) <= 6:
            # Very short values are unusual for OpenAI keys, but handle them
            # defensively without inventing random trailing characters.
            return f"{clean_key[:3]}***{clean_key[-3:]}（已保存，可沿用）"
        head = clean_key[:3]
        tail = clean_key[-3:]
        return f"{head}{'*' * 12}{tail}（已保存，可沿用）"

    def refresh_api_placeholder(self) -> None:
        key = self.read_env_key()
        for attr in ("home_api_edit", "api_edit"):
            editor = getattr(self, attr, None)
            if editor is not None:
                editor.setPlaceholderText(self.masked_key(key))
        self.state.api_key_set = bool(key)

    def save_api_key_from_editor(self, editor: QLineEdit) -> bool:
        key = editor.text().strip()
        if not key:
            # User pressed Save without typing a key. Always prompt them to
            # enter one — never silently report an existing key as "saved",
            # which made an empty Save misleadingly show 已保存／可沿用.
            QMessageBox.warning(self, "請輸入 API Key", "請輸入 API Key")
            return False
        old = self.read_env_key()
        if old and key != old:
            ret = QMessageBox.question(self, "Replace API Key", "Replace the shared OpenAI API Key for all projects?", QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)
            if ret != QMessageBox.StandardButton.Yes:
                editor.clear()
                return False
        env = self.env_path()
        lines = []
        if env.exists():
            lines = [ln for ln in env.read_text(encoding="utf-8").splitlines() if not ln.startswith("OPENAI_API_KEY=")]
        lines.append(f"OPENAI_API_KEY={key}")
        env.write_text("\n".join(lines) + "\n", encoding="utf-8")
        os.environ["OPENAI_API_KEY"] = key
        self.state.api_key_set = True
        self.save_state()
        editor.clear()
        self.refresh_api_placeholder()
        QMessageBox.information(self, "API Key", f"Shared API Key saved: {self.masked_key(key)}")
        return True

    def save_api_key(self) -> None:
        editor = getattr(self, "home_api_edit", None) or getattr(self, "api_edit", None)
        if editor is not None:
            self.save_api_key_from_editor(editor)

    # ---------- submit ----------
    def submit_home(self) -> bool:
        editor = getattr(self, "home_api_edit", None)
        if editor is not None and editor.text().strip():
            if not self.save_api_key_from_editor(editor):
                return False
        if not self.state.project_id:
            QMessageBox.warning(self, "Missing", "Please create or open a project first.")
            return False
        self.state.completed_steps[1] = True
        self.dirty_steps[1] = False
        self.save_state()
        return True

    def submit_project(self) -> bool:
        self.update_state_from_widgets()
        if not self.state.class_name:
            QMessageBox.warning(self, "Missing", "Class Name is required.")
            return False
        self.init_workspace_silent()
        self.save_state()
        self.status_label.setText("Status: project class ready")
        return True

    def submit_upload(self) -> bool:
        self.update_state_from_widgets(); self.save_state()
        if not list_images(self.raw_dir()): QMessageBox.warning(self, "Missing", "請先上傳圖片。") ; return False
        return True

    def submit_crop(self) -> bool:
        if not list_images(self.inputs_dir()):
            QMessageBox.warning(self, "Missing", "請至少完成一張 Step 4 輸入圖像。請先點選左側縮圖後按『使用原始圖片』，或用裁切框裁切原圖。")
            return False
        return True

    def submit_regions(self) -> bool:
        crops = list_images(self.inputs_dir())
        if not crops:
            QMessageBox.warning(self, "Missing", "請先完成裁切或使用原始圖片。")
            return False
        ready = self.all_region_ready_image_paths()
        if not ready:
            QMessageBox.warning(self, "Missing", "請至少完成 1 張圖像的 ROI 與 Target Area 框選後，再進入 Step 5。")
            return False
        # Step 4 no longer decides the final generation set.  The final 1～16
        # input groups are selected in Step 5 Prompt 編輯.
        self.refresh_prompt_groups()
        return True

    def submit_prompt(self) -> bool:
        return self.save_prompt(show_message=False)

    def submit_model(self) -> bool:
        return self.save_model_settings(show_message=False)

    def submit_aggregate(self) -> bool:
        summary = self.build_aggregate_text()
        if hasattr(self,"aggregate_box"): self.aggregate_box.setPlainText(summary)
        self.save_state(); return True

    def submit_run(self) -> bool:
        if self.current_process and self.current_process.state() != QProcess.NotRunning:
            QMessageBox.warning(self,"Running","請等待生成程序結束。"); return False
        if not self.current_run_has_successful_outputs():
            QMessageBox.warning(self,"Missing","目前 Run 尚未找到有效輸出圖，請先按「重新開始生成」或「從上次進度開始生成」並確認 log 顯示成功。")
            return False
        return True

    def submit_export(self) -> bool:
        return self.export_dataset(show_message=True)

    # ---------- Step 1 setup ----------
    def verify_environment(self) -> None:
        self.run_quick_command_dialog([sys.executable, str(self.root / "scripts" / "verify_env.py")], "檢查環境")

    def init_class_workspace(self) -> None:
        self.update_state_from_widgets(); self.save_state(); self.init_workspace_silent(); self.refresh_all(); QMessageBox.information(self, "Done", "工作區已建立 / 同步。")

    def init_workspace_silent(self) -> None:
        # UI-managed workspace: keep all user products inside project/<project_name>/.
        for d in [self.raw_dir(), self.reference_dir(), self.runs_dir()]:
            ensure_dir(d)
        self.save_state()

    # ---------- upload ----------
    def choose_source_folder(self) -> None:
        # Use a non-native, non-modal QFileDialog so the main UI remains active.
        # This avoids the confusing Windows behavior where dragging from a modal
        # folder picker to the disabled parent window always shows a forbidden cursor.
        dlg = QFileDialog(self, "選擇圖片資料夾")
        dlg.setFileMode(QFileDialog.Directory)
        dlg.setOption(QFileDialog.ShowDirsOnly, True)
        dlg.setOption(QFileDialog.DontUseNativeDialog, True)
        dlg.setWindowModality(Qt.NonModal)
        dlg.setAttribute(Qt.WA_DeleteOnClose, True)
        dlg.accepted.connect(self.safe_action("dialog_import_source_paths", lambda d=dlg: self.import_source_paths([Path(x) for x in d.selectedFiles() if x])))
        self._source_folder_dialog = dlg
        dlg.show()

    def import_source_paths(self, paths: list[Path]) -> None:
        self.update_state_from_widgets(); self.init_workspace_silent(); count=0
        last_imported: Optional[Path] = None
        for p in paths:
            if p.is_dir():
                for img in list_images(p): last_imported = copy_image_to(img, self.raw_dir()); count+=1
            elif p.is_file() and p.suffix.lower() in SUPPORTED_EXTS:
                last_imported = copy_image_to(p, self.raw_dir()); count+=1
        if count:
            self.clear_downstream_generation_artifacts(clear_inputs=True)
            self.clear_all_visual_previews()
        self.refresh_uploads(auto_select=True, select_path=last_imported); self.refresh_raw_thumbs(); self.refresh_crops(); self.refresh_region_thumbs(); self.refresh_prompt_groups(); self.mark_dirty(2); self.status_label.setText(f"Status: imported {count} image(s); downstream Step 3～9 data cleared")

    def delete_selected_uploads(self) -> None:
        for item in self.upload_list.selectedItems():
            p=Path(item.data(Qt.UserRole))
            try: p.unlink()
            except OSError: pass
        self.clear_downstream_generation_artifacts(clear_inputs=True)
        self.clear_all_visual_previews()
        self.refresh_uploads(); self.refresh_raw_thumbs(); self.refresh_crops(); self.refresh_region_thumbs(); self.refresh_prompt_groups(); self.mark_dirty(2)

    # ---------- crop ----------
    def crop_record_json_path(self) -> Path:
        return self.class_config_dir() / "crop_records.json"

    def load_crop_record(self, crop_path: Path) -> dict:
        # Defect mode no longer writes crop records; crops overwrite the source
        # image in place. Return an empty record so legacy callers stay safe.
        return {}

    def source_path_from_crop_record(self, rec: dict) -> Optional[Path]:
        candidates: list[Path] = []
        for key in ["source_path", "source_abs"]:
            val = rec.get(key)
            if val:
                candidates.append(Path(str(val)))
        source_name = rec.get("source")
        if source_name:
            candidates.append(self.raw_dir() / str(source_name))
            candidates.extend(self.raw_dir().glob(str(source_name)))
        for c in candidates:
            try:
                if c.exists():
                    return c
            except Exception:
                pass
        return None

    def on_crop_done_selected(self, path_str: str) -> None:
        crop_path = Path(path_str)
        self.crop_done_preview.set_path(crop_path)
        rec = self.load_crop_record(crop_path)
        rect_vals = rec.get("crop_xyxy") or []
        if not (isinstance(rect_vals, list) and len(rect_vals) == 4):
            # Older projects may not have crop metadata; still keep the cropped preview usable.
            return
        rect = tuple(int(x) for x in rect_vals)
        source_path = self.source_path_from_crop_record(rec)
        if not source_path:
            QMessageBox.warning(self, "Missing source", "找不到此裁切圖對應的原始圖像；只能顯示右側預覽，無法回到中間裁切框編輯。")
            return
        w, h = rect[2] - rect[0], rect[3] - rect[1]
        for edit, value in [(self.crop_w_edit, w), (self.crop_h_edit, h)]:
            old = edit.blockSignals(True)
            edit.setText(str(value))
            edit.blockSignals(old)
        self.state.crop_width = w
        self.state.crop_height = h
        self.crop_canvas.set_crop_size(w, h)
        self.crop_canvas.set_image(source_path, rect)
        self.validate_crop_inputs()
        self.save_state()
        self.status_label.setText(f"Status: loaded crop frame {w}*{h} from {crop_path.name}")

    def step3_mode_is_selected(self) -> bool:
        return bool(getattr(self.state, "crop_mode_selected", False))

    def reset_step3_visible_previews(self) -> None:
        """Clear Step 3 visible previews until the user actively selects a thumbnail."""
        if hasattr(self, "crop_canvas"):
            self.crop_canvas.set_image(None)
        if hasattr(self, "crop_done_grid"):
            self.crop_done_grid.clear()
        if hasattr(self, "crop_done_preview"):
            self.crop_done_preview.clear("尚無 Step 4 輸入圖像")
        if hasattr(self, "raw_thumb_grid"):
            self.raw_thumb_grid.clearSelection()

    def prepare_step3_entry(self) -> None:
        """Called whenever Step 3 is entered/reloaded; prevents stale preview carry-over."""
        if hasattr(self, "raw_thumb_grid"):
            self.refresh_raw_thumbs()
        self.reset_step3_visible_previews()

    def select_step3_mode(self, target_mode: str) -> None:
        """Initial big-card selection. This is not a cross-mode switch yet."""
        target_mode = "no_crop" if target_mode == "no_crop" else "crop"
        self.state.crop_mode = target_mode
        self.state.crop_mode_selected = True
        self.state.completed_steps[3] = False
        for j in range(4, STEP_COUNT):
            self.state.completed_steps[j] = False
            self.dirty_steps[j] = True
        if target_mode == "no_crop":
            self.use_original_images_without_crop(confirm=False, mark_completed=False, auto_select=False)
        else:
            self.clear_downstream_generation_artifacts(clear_inputs=True, clear_runs_exports=False)
            self.save_state()
            self.refresh_region_thumbs(); self.refresh_prompt_groups(); self.update_step_buttons()
        self.refresh_raw_thumbs()
        self.reset_step3_visible_previews()
        self.update_step3_mode_ui()
        self.status_label.setText("Status: 已選擇 Step 3 模式，請先點選左側縮圖再繼續。")

    def switch_step3_other_mode(self) -> None:
        current = getattr(self.state, "crop_mode", "crop") or "crop"
        self.switch_step3_mode("crop" if current == "no_crop" else "no_crop")

    def update_step3_mode_ui(self) -> None:
        """Compatibility no-op after reverting Step 3 to the original workflow.

        Step 3 no longer has a global mutually-exclusive mode page. The user
        simply selects a raw thumbnail, then either presses 『使用原始圖片』 or
        uses the crop frame. Older project states may still contain crop_mode
        fields, so keep this method harmless for old signal/loader calls.
        """
        if hasattr(self, "crop_controls_box"):
            self.crop_controls_box.setVisible(True)
        for name in ["crop_w_edit", "crop_h_edit", "crop_confirm_btn"]:
            if hasattr(self, name):
                getattr(self, name).setEnabled(True)

    def switch_step3_mode(self, target_mode: str) -> None:
        target_mode = "no_crop" if target_mode == "no_crop" else "crop"
        current = getattr(self.state, "crop_mode", "crop") or "crop"
        if self.step3_mode_is_selected() and current == target_mode:
            QMessageBox.information(self, "Info", "目前已經是此模式。")
            return
        target_label = "使用原圖尺寸" if target_mode == "no_crop" else "裁切尺寸"
        current_label = "使用原圖尺寸" if current == "no_crop" else "裁切尺寸"
        if self.step3_mode_is_selected():
            ret = QMessageBox.question(
                self,
                "切換 Step 3 模式",
                f"是否確定要切換到「{target_label}」模式？目前「{current_label}」模式下原圖尺寸/裁切好的所有圖片與後續 ROI、runs、exports 都會被清除。",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if ret != QMessageBox.StandardButton.Yes:
                return False
        self.state.crop_mode = target_mode
        self.state.crop_mode_selected = True
        self.state.completed_steps[3] = False
        for j in range(4, STEP_COUNT):
            self.state.completed_steps[j] = False
            self.dirty_steps[j] = True
        if target_mode == "no_crop":
            self.use_original_images_without_crop(confirm=False, mark_completed=False, auto_select=False)
        else:
            self.clear_downstream_generation_artifacts(clear_inputs=True, clear_runs_exports=False)
            self.save_state()
            self.refresh_region_thumbs(); self.refresh_prompt_groups(); self.update_step_buttons()
        self.refresh_raw_thumbs()
        self.reset_step3_visible_previews()
        self.update_step3_mode_ui()
        self.status_label.setText(f"Status: 已切換到{target_label}模式，請先點選左側縮圖。")

    def crop_paths_for_source(self, source: Path) -> list[Path]:
        all_inputs = list_images(self.inputs_dir())
        mode = getattr(self.state, "crop_mode", "crop") or "crop"
        if mode == "no_crop":
            source_stem = sanitize_name(source.stem)
            matched = [p for p in all_inputs if p.stem == source_stem or p.name == source.name]
            return matched or all_inputs
        records_dir = self.inputs_dir().parent / "crop_records"
        matched: list[Path] = []
        for p in all_inputs:
            rec = self.load_crop_record(p)
            src_name = str(rec.get("source", ""))
            src_path = str(rec.get("source_path", "") or rec.get("source_abs", ""))
            if src_name == source.name or (src_path and Path(src_path).name == source.name):
                matched.append(p)
        if not matched:
            base = sanitize_name(source.stem)
            matched = [p for p in all_inputs if p.stem.startswith(f"{base}_crop")]
        return sorted(matched, key=lambda p: p.name.lower())

    def refresh_crops_for_source(self, source: Path, auto_select: bool=False) -> None:
        paths = self.crop_paths_for_source(source)
        if hasattr(self, "crop_done_grid"):
            self.crop_done_grid.load_paths(paths)
            if auto_select and self.crop_done_grid.count():
                self.crop_done_grid.setCurrentRow(self.crop_done_grid.count() - 1)
        if not paths and hasattr(self, "crop_done_preview"):
            self.crop_done_preview.clear("尚無 Step 4 輸入圖像")

    def refresh_crops_for_current_raw(self, auto_select: bool=False) -> None:
        if hasattr(self, "raw_thumb_grid"):
            items = self.raw_thumb_grid.selectedItems()
            if items:
                self.refresh_crops_for_source(Path(str(items[0].data(Qt.UserRole))), auto_select=auto_select)
                return
        self.refresh_crops(auto_select=auto_select)

    def validate_crop_inputs(self) -> bool:
        """Validate Step 3 crop-frame size.

        Crop/input images are not constrained by the GPT-image-2 output-size
        rules. The crop frame only needs to be a positive pixel size and, once
        a source image is selected, fit inside that source image. Output-size
        validation remains centralized in Step 6.
        """
        edits = [getattr(self, "crop_w_edit", None), getattr(self, "crop_h_edit", None)]
        if any(e is None for e in edits):
            return False
        vals: list[int] = []
        ok = True
        for edit in edits:
            try:
                v = int(str(edit.text()).strip())
            except Exception:
                v = -1
            vals.append(v)
            if v <= 0:
                ok = False
        selected_size = getattr(getattr(self, "crop_canvas", None), "image_size", (0, 0))
        img_w, img_h = selected_size if selected_size else (0, 0)
        if img_w and img_h and len(vals) == 2:
            if vals[0] > img_w or vals[1] > img_h:
                ok = False
        if len(vals) == 2:
            bad_w = vals[0] <= 0 or (img_w and vals[0] > img_w)
            bad_h = vals[1] <= 0 or (img_h and vals[1] > img_h)
            self.crop_w_edit.setStyleSheet("border: 2px solid #dc2626;" if bad_w else "")
            self.crop_h_edit.setStyleSheet("border: 2px solid #dc2626;" if bad_h else "")
        if hasattr(self, "crop_error"):
            if not ok and img_w and img_h:
                self.crop_error.setText(f"裁切尺寸需大於 0，且不可超過目前原圖尺寸 {img_w}×{img_h} px。建議輸入值範圍：320(px)~1280(px)")
            elif not ok:
                self.crop_error.setText("裁切尺寸需大於 0。建議輸入值範圍：320(px)~1280(px)")
            else:
                self.crop_error.setText("建議輸入值範圍：320(px)~1280(px)")
            self.crop_error.setVisible(True)
        if hasattr(self, "crop_confirm_btn"):
            self.crop_confirm_btn.setVisible(ok)
        return ok

    def confirm_crop_size(self) -> None:
        if not self.validate_crop_inputs(): return
        w,h=int(self.crop_w_edit.text()), int(self.crop_h_edit.text())
        self.state.crop_width=w; self.state.crop_height=h; self.crop_canvas.set_crop_size(w,h); self.save_state(); self.status_label.setText(f"Status: crop size {w}*{h}")

    def clear_downstream_generation_artifacts(self, clear_inputs: bool = False, clear_runs_exports: bool = True) -> None:
        """Remove stale editable inputs, optionally preserving previous run/export products.

        Step 3 operations such as 「使用原始圖片」 or adding a crop should not erase
        historical runs/exports. Those folders are only cleared when explicitly
        starting the same Run name again in Step 8, or when an upstream data-upload
        change intentionally invalidates the whole workflow.
        """
        folders: list[Path] = []
        if clear_inputs:
            # raw_image holds the input images themselves and is preserved here;
            # only the downstream annotation references are invalidated. ROI/Target
            # geometry lives in project_state.json (state.regions).
            folders.append(self.reference_dir())
        if clear_runs_exports:
            folders.extend([self.runs_dir(), self.exports_dir()])
        for folder in folders:
            try:
                if folder.exists():
                    shutil.rmtree(folder)
            except Exception:
                pass
        if clear_inputs:
            self.state.selected_region_stems = []
            self.state.regions = {}
        self.init_workspace_silent()

    def selected_raw_image_path(self) -> Optional[Path]:
        if not hasattr(self, "raw_thumb_grid"):
            return None
        items = self.raw_thumb_grid.selectedItems()
        if not items:
            return None
        try:
            p = Path(str(items[0].data(Qt.UserRole)))
            return p if p.exists() else None
        except Exception:
            return None

    def use_selected_original_image(self) -> None:
        """Use the currently selected raw image as a Step 4 input without cropping.

        raw_image/ is already the generation input folder, so no copy is made and
        no crop record is written.
        """
        src = self.selected_raw_image_path()
        if not src:
            QMessageBox.warning(self, "Missing", "請先點選左側『已上傳圖像縮圖』中的一張圖片。")
            return
        self.state.crop_mode = "mixed"
        self.state.crop_mode_selected = True
        self.state.completed_steps[3] = False
        for j in range(4, STEP_COUNT):
            self.state.completed_steps[j] = False
            self.dirty_steps[j] = True
        # Preserve previous runs/exports when using original image as Step 4 input.
        self.save_state()
        self.refresh_crops(auto_select=True)
        self.refresh_region_thumbs(); self.refresh_prompt_groups(); self.update_step_buttons()
        self.status_label.setText(f"Status: 已將原始圖片加入 Step 4 輸入圖像：{src.name}")

    def step4_input_modes(self) -> set[str]:
        """Classify the current crop-step input images as original ('no_crop') vs cropped.

        Used to decide whether switching between 「使用原始圖片」 and cropping needs a
        confirmation prompt. A brand-new / clean project has no inputs (empty set), so no
        prompt is shown; only an actual mode switch (cropped -> original or original ->
        cropped) prompts the user and clears the previous products.
        """
        modes: set[str] = set()
        for p in list_images(self.inputs_dir()):
            rec = self.load_crop_record(p)
            if rec.get("mode") == "no_crop":
                modes.add("no_crop")
            elif re.search(r"_crop\d+$", p.stem):
                modes.add("crop")
            else:
                modes.add("no_crop")
        return modes

    def use_original_images_without_crop(self, confirm: bool = True, mark_completed: bool = False, auto_select: bool = False) -> bool:
        raws = list_images(self.raw_dir())
        if not raws:
            QMessageBox.warning(self, "Missing", "請先在 Step 2 上傳原始圖片。")
            return False
        # Only warn when there are previously cropped Step 4 inputs that this action would
        # discard. A clean / first-time project (no cropped inputs) switches silently.
        if confirm and ("crop" in self.step4_input_modes()):
            ret = QMessageBox.question(
                self,
                "直接使用原圖",
                "此動作會清除目前繪製完成的 ROI 與 Target Area ，並把所有原圖直接作為 Step 4 輸入使用。既有的 runs 會保留。是否繼續？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if ret != QMessageBox.StandardButton.Yes:
                return False
        # raw_image/ already holds the input images; just invalidate downstream
        # references/geometry without copying or writing crop records.
        self.clear_downstream_generation_artifacts(clear_inputs=True, clear_runs_exports=False)
        self.state.crop_mode = "no_crop"
        self.state.crop_mode_selected = True
        self.state.completed_steps[3] = bool(mark_completed)
        for j in range(4, STEP_COUNT):
            self.state.completed_steps[j] = False
            self.dirty_steps[j] = True
        self.save_state()
        if auto_select:
            self.refresh_crops(auto_select=True)
        elif hasattr(self, "crop_done_grid"):
            self.crop_done_grid.clear()
        if hasattr(self, "crop_done_preview"):
            self.crop_done_preview.clear("尚無 Step 4 輸入圖像")
        self.refresh_region_thumbs(); self.refresh_prompt_groups(); self.update_step_buttons()
        self.update_step3_mode_ui()
        self.status_label.setText(f"Status: 使用原圖作為 Step 4 輸入，共 {len(list_images(self.inputs_dir()))} 張。")
        return True

    def use_all_original_images_and_go_step4(self) -> None:
        """Footer action: use all Step 2 raw uploads as Step 4 inputs and continue."""
        ok = self.use_original_images_without_crop(confirm=True, mark_completed=True, auto_select=True)
        if not ok:
            return
        self.state.completed_steps[3] = True
        self.dirty_steps[3] = False
        self.save_state()
        self.update_step_buttons()
        self.goto_step(4)

    def crop_select_raw_image(self, path_str: str) -> None:
        src = Path(path_str)
        self.crop_canvas.set_image(src)
        self.refresh_crops(auto_select=False)
        if self.validate_crop_inputs():
            self.confirm_crop_size()
        self.status_label.setText("Status: 已載入原圖。可在中間圖像進行裁切；若要全部原圖直接送入 Step 4，請按底部『使用原始圖片』。")

    def make_crop_from_rect(self, src: Path, rect: tuple[int,int,int,int]) -> Optional[Path]:
        # Switching from 「使用原始圖片」 back to cropping must clear the previously copied
        # originals (and their ROI/Target Area) before the first crop is committed. A clean
        # project or one that is already in crop mode adds crops without any prompt.
        if "no_crop" in self.step4_input_modes():
            ret = QMessageBox.question(
                self,
                "改用裁切圖",
                "此動作會清除目前繪製完成的 ROI 與 Target Area ，並以裁切後的圖像就地取代原圖作為 Step 4 輸入。既有的 runs 會保留。是否繼續？",
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if ret != QMessageBox.StandardButton.Yes:
                return None
            self.clear_downstream_generation_artifacts(clear_inputs=True, clear_runs_exports=False)
        try:
            img=Image.open(src).convert("RGB"); crop=img.crop(rect)
            # Crop in place: overwrite the source image in raw_image/ so the input
            # keeps the same stem; no separate crop folder or crop records.
            out=src; crop.save(out)
            # Editing the raw image invalidates any previously rendered reference.
            try:
                ref = self.reference_dir()/f"{src.stem}.png"
                if ref.exists(): ref.unlink()
            except Exception:
                pass
            self.state.regions.pop(src.stem, None)
            self.state.crop_mode = "mixed"
            self.state.crop_mode_selected = True
            self.save_state()
            # Preserve previous runs/exports when cropping.
            self.mark_dirty(3); return out
        except Exception as exc:
            QMessageBox.critical(self,"Crop failed",str(exc)); return None


    def _rewrite_crop_record_text_output(self, txt_path: Path, output_name: str) -> None:
        if not txt_path.exists():
            return
        try:
            lines = txt_path.read_text(encoding="utf-8", errors="ignore").splitlines()
            found = False
            new_lines = []
            for line in lines:
                if line.strip().startswith("output:"):
                    new_lines.append(f"output: {output_name}")
                    found = True
                else:
                    new_lines.append(line)
            if not found:
                new_lines.insert(0, f"output: {output_name}")
            txt_path.write_text("\n".join(new_lines) + "\n", encoding="utf-8")
        except Exception:
            pass

    def _renumber_crop_outputs(self) -> None:
        """No-op for defect mode.

        Crops overwrite the source image in raw_image/ in place, so there are no
        sequential *_crop### outputs to renumber.
        """
        return


    def delete_selected_crop(self) -> None:
        items = self.crop_done_grid.selectedItems() if hasattr(self, "crop_done_grid") else []
        if not items:
            QMessageBox.information(self, "Info", "請先選擇要刪除的裁切圖。")
            return
        for item in items:
            p = Path(str(item.data(Qt.UserRole)))
            stem = p.stem
            for q in [p]:
                try:
                    if q.exists(): q.unlink()
                except Exception:
                    pass
            for ref in self.reference_dir().glob(f"{stem}.*"):
                try: ref.unlink()
                except Exception: pass
            self.state.regions.pop(stem, None)
        self.save_state()
        remaining_crops = list_images(self.inputs_dir())
        if hasattr(self, "crop_canvas"):
            self.crop_canvas.set_image(None)
        if hasattr(self, "crop_done_preview"):
            self.crop_done_preview.clear("尚無裁切完成圖" if not remaining_crops else "請選擇裁切完成圖")
        self.refresh_crops(auto_select=False); self.refresh_region_thumbs(); self.mark_dirty(3)

    def delete_all_crops(self) -> None:
        if not (hasattr(self, "crop_done_grid") and self.crop_done_grid.count()):
            QMessageBox.information(self, "Info", "目前沒有可刪除的裁切圖。")
            return
        # Remove every Step 4 input image together with its reference image and
        # ROI / Target Area geometry. Existing runs are preserved.
        for p in list_images(self.raw_dir()):
            try: p.unlink()
            except Exception: pass
        self.clear_downstream_generation_artifacts(clear_inputs=True, clear_runs_exports=False)
        if hasattr(self, "crop_canvas"):
            self.crop_canvas.set_image(None)
        if hasattr(self, "crop_done_preview"):
            self.crop_done_preview.clear("尚無 Step 4 輸入圖像")
        self.refresh_crops(auto_select=False); self.refresh_region_thumbs(); self.refresh_prompt_groups(); self.mark_dirty(3)
        self.status_label.setText("Status: 已刪除所有 Step 4 輸入裁切圖。")

    # ---------- regions ----------
    def region_select_crop_image(self, path_str: str) -> None:
        # Save the previous image before switching. This makes Step 4 truly
        # auto-save even when the user navigates to another image immediately.
        try:
            old = getattr(self.region_canvas, "image_path", None)
            newp = Path(path_str)
            if old and Path(old) != newp:
                self.save_current_regions(silent=True)
        except Exception:
            pass
        p=Path(path_str); self.region_canvas.set_image(p); self.load_regions_into_canvas(p); self.region_status.setText(self.region_status_text(p))

    def load_regions_into_canvas(self, p: Path) -> None:
        """Repopulate the ROI/Target canvas from project_state.json geometry."""
        if not hasattr(self, "region_canvas"):
            return
        geo = self.state.regions.get(p.stem, {}) if isinstance(self.state.regions, dict) else {}
        rois: list[tuple[int, int, int, int]] = []
        for r in geo.get("rois", []) or []:
            try:
                rois.append(tuple(int(v) for v in r[:4]))
            except Exception:
                pass
        targets: list[dict[str, Any]] = []
        for t in geo.get("targets", []) or []:
            if not isinstance(t, dict):
                continue
            if t.get("kind") == "polygon":
                pts = [(int(x), int(y)) for x, y in t.get("points", []) if len([x, y]) == 2]
                if len(pts) >= 3:
                    targets.append({"kind": "polygon", "points": pts})
            else:
                rect = t.get("rect")
                if rect and len(rect) >= 4:
                    targets.append({"kind": "rect", "rect": tuple(int(v) for v in rect[:4])})
        self.region_canvas.rois = rois
        self.region_canvas.selected_roi_indices = set()
        self.region_canvas.target_areas = targets
        self.region_canvas.selected_target_indices = set()
        self.region_canvas.update()

    def region_status_text(self, p: Path) -> str:
        geo = self.state.regions.get(p.stem, {}) if isinstance(self.state.regions, dict) else {}
        roi_count = len(geo.get("rois", []) or [])
        target_count = len(geo.get("targets", []) or [])
        return f"目前圖像：{p.name}｜ROI：{roi_count} 個｜Target Area：{target_count} 個"

    def delete_selected_roi(self) -> None:
        if not hasattr(self, "region_canvas") or not self.region_canvas.image_path:
            QMessageBox.information(self, "Info", "請先選擇一張裁切圖。")
            return
        if not self.region_canvas.delete_selected_roi():
            QMessageBox.information(self, "Info", "請先切換「選取 ROI」並點選一個或多個要刪除的 ROI。")
            return
        self.region_status.setText(self.region_status_text(self.region_canvas.image_path))

    def delete_target_area(self) -> None:
        if not hasattr(self, "region_canvas") or not self.region_canvas.image_path:
            QMessageBox.information(self, "Info", "請先選擇一張裁切圖。")
            return
        if not self.region_canvas.clear_target_areas():
            QMessageBox.information(self, "Info", "目前這張圖沒有已框選的 Target Area。")
            return
        self.region_status.setText(self.region_status_text(self.region_canvas.image_path))

    def delete_selected_target_area(self) -> None:
        if not hasattr(self, "region_canvas") or not self.region_canvas.image_path:
            QMessageBox.information(self, "Info", "請先選擇一張裁切圖。")
            return
        if not self.region_canvas.delete_selected_target_area():
            QMessageBox.information(self, "Info", "請先切換「選取 Target Area」並點選一個或多個要刪除的 Target Area。")
            return
        self.region_status.setText(self.region_status_text(self.region_canvas.image_path))

    def all_region_ready_image_paths(self) -> list[Path]:
        """All Step 4 input images that already have both ROI and Target Area.

        This is intentionally independent from selected_region_stems. Step 4
        selection controls which image groups are sent to generation, while
        Step 5 should still allow the user to inspect/prompt against every
        completed group.
        """
        ready: list[Path] = []
        regions = self.state.regions if isinstance(self.state.regions, dict) else {}
        for p in list_images(self.inputs_dir()):
            geo = regions.get(p.stem, {})
            has_geo = bool(geo.get("rois")) and bool(geo.get("targets"))
            # project_state.json is no longer persisted, so on reopen state.regions
            # is empty; fall back to the rendered annotation reference on disk
            # (data/reference_image/<stem>.png) — the artifact generation uses — so a
            # group completed in an earlier session still appears in 引用組別.
            has_reference = (self.reference_dir() / f"{p.stem}.png").exists()
            if has_geo or has_reference:
                ready.append(p)
        return ready

    def selected_region_image_paths(self) -> list[Path]:
        """Final input images sent to generation = the groups chosen in 引用組別.

        Maps the user's 引用組別 selection (state.selected_region_stems) to the
        prepared raw images, capped at 16. Empty when nothing is selected, so
        Step 5 enforces an explicit selection before generation.
        """
        all_paths = list_images(self.inputs_dir())
        by_stem = {p.stem: p for p in all_paths}
        stems = [s for s in getattr(self.state, "selected_region_stems", []) if s in by_stem]
        return [by_stem[s] for s in stems][:16]

    def selected_region_stems(self) -> list[str]:
        """Persist and return the 引用組別 selection as image stems (no file on disk).

        The selection is passed to the backend inline via --selected-stems, so no
        .selected_region_stems.txt is written anywhere."""
        stems = [p.stem for p in self.selected_region_image_paths()][:16]
        self.state.selected_region_stems = stems
        self.save_state()
        return stems

    def update_region_selection_status(self) -> None:
        if hasattr(self, "region_selection_status"):
            count = len(getattr(self, "step4_selected_view_stems", []) or [])
            self.region_selection_status.setText(f"目前選取：{count} 組（Step 5 決定輸入組合）")

    def region_selected_groups_changed(self, paths: object) -> None:
        """Step 4 selection is for viewing/navigation only — NOT an edit. It must not
        mark Step 4 dirty, otherwise merely clicking another framed image would
        wrongly un-complete the already-finished later steps."""
        try:
            selected_paths = [Path(str(p)) for p in (paths or [])]
        except Exception:
            selected_paths = []
        self.step4_selected_view_stems = [p.stem for p in selected_paths]
        self.update_region_selection_status()

    def auto_save_current_regions(self) -> None:
        self.save_current_regions(silent=True)
        self.refresh_prompt_groups()
        self.update_region_selection_status()

    def _format_regions_for_log(self, rois, targets) -> str:
        """Readable per-box coordinate list for the Step 4 change log."""
        roi_str = ", ".join(f"({r[0]},{r[1]},{r[2]},{r[3]})" for r in rois) or "-"
        tgt_parts: list[str] = []
        for t in targets:
            if t.get("kind") == "polygon":
                pts = t.get("points", [])
                tgt_parts.append("poly[" + " ".join(f"({int(x)},{int(y)})" for x, y in pts) + "]")
            else:
                rr = t.get("rect", [])
                tgt_parts.append(f"({rr[0]},{rr[1]},{rr[2]},{rr[3]})" if len(rr) >= 4 else str(rr))
        return f"ROI[{len(rois)}]={roi_str}; Target[{len(targets)}]={', '.join(tgt_parts) or '-'}"

    def save_current_regions(self, silent: bool=False) -> None:
        p=self.region_canvas.image_path
        if not p: return
        rois = [tuple(int(v) for v in roi[:4]) for roi in getattr(self.region_canvas, "rois", [])]
        targets: list[dict[str, Any]] = []
        for shape in getattr(self.region_canvas, "target_areas", []):
            if shape.get("kind") == "polygon":
                pts = [[int(x), int(y)] for x, y in shape.get("points", [])]
                targets.append({"kind": "polygon", "points": pts})
            else:
                targets.append({"kind": "rect", "rect": list(target_area_bbox(shape))})
        if not isinstance(self.state.regions, dict):
            self.state.regions = {}
        new_geo = {"rois": [list(r) for r in rois], "targets": targets}
        prev_geo = self.state.regions.get(p.stem)
        changed = prev_geo != new_geo
        self.state.regions[p.stem] = new_geo
        if not changed:
            # No actual edit (e.g. switching images, or re-entering Step 4 and only
            # selecting/viewing). Do NOT re-render, do NOT log, and crucially do NOT
            # mark Step 4 dirty — that would wrongly un-complete the later steps.
            if not silent:
                QMessageBox.information(self, "Done", "目前圖像的 ROI / Target Area 已儲存。")
            return
        # Render the Image 2 annotation reference to match the Step 4 editor exactly:
        # ROI box = red, Target Area box = blue, each a 3px outline ONLY (no fill).
        try:
            ensure_dir(self.reference_dir())
            ref_path = self.reference_dir()/f"{p.stem}.png"
            base = Image.open(p).convert("RGB")
            size = base.size
            roi_rgb, target_rgb = (255, 23, 68), (0, 194, 255)
            if rois or targets:
                overlay = Image.new("RGBA", size, (0, 0, 0, 0))
                odraw = ImageDraw.Draw(overlay)
                for shape in targets:
                    if shape.get("kind") == "polygon":
                        pts = [(int(x), int(y)) for x, y in shape.get("points", [])]
                        if len(pts) >= 3:
                            odraw.polygon(pts, fill=None, outline=(*target_rgb, 255), width=3)
                    else:
                        rect = shape.get("rect") or [0, 0, 0, 0]
                        odraw.rectangle([int(v) for v in rect[:4]], fill=None, outline=(*target_rgb, 255), width=3)
                for roi in rois:
                    odraw.rectangle(list(roi), fill=None, outline=(*roi_rgb, 255), width=3)
                composed = Image.alpha_composite(base.convert("RGBA"), overlay).convert("RGB")
                composed.save(ref_path)
            elif ref_path.exists():
                ref_path.unlink()
        except Exception:
            pass
        self.save_state()
        self.log_event(f"Step 4 ROI/Target changed for {p.stem}: {self._format_regions_for_log(rois, targets)}")
        self.region_status.setText(self.region_status_text(p)); self.mark_dirty(4)
        if hasattr(self, "prompt_group_grid"):
            self.refresh_prompt_groups()
        if not silent: QMessageBox.information(self,"Done","目前圖像的 ROI / Target Area 已儲存。")

    # ---------- prompt ----------
    def on_prompt_mode_changed(self) -> None:
        if not hasattr(self, "prompt_mode_combo"):
            return
        is_template = self.prompt_mode_combo.currentText() == "使用模板"
        for w in [self.prompt_template_label, self.prompt_template_combo, self.apply_template_btn]:
            if w is not None:
                w.setVisible(is_template)
        self.update_actual_prompt_preview()
        self.mark_dirty(5)

    def on_prompt_text_changed(self) -> None:
        self.update_actual_prompt_preview()
        self.mark_dirty(5)

    def on_prompt_group_selected(self, group_id: str) -> None:
        # Current item is only used for focus/preview.  The actual final input
        # groups are stored in state.selected_region_stems by
        # on_prompt_group_selection_changed().
        self.state.prompt_group = group_id or ""
        self.update_actual_prompt_preview()
        self.mark_dirty(5)

    def prompt_ready_stems(self) -> list[str]:
        ready = {p.stem for p in self.all_region_ready_image_paths()}
        return [s for s in getattr(self.state, "selected_region_stems", []) if s in ready][:16]

    def update_prompt_selection_status(self) -> None:
        if hasattr(self, "prompt_selection_status"):
            n = len(self.prompt_ready_stems())
            self.prompt_selection_status.setText(f"已選定：{n}/16 組")

    def _set_prompt_group_selection(self, stems: list[str]) -> None:
        if not hasattr(self, "prompt_group_grid"):
            return
        blocker = QSignalBlocker(self.prompt_group_grid)
        try:
            keep = set(stems)
            for i in range(self.prompt_group_grid.count()):
                item = self.prompt_group_grid.item(i)
                role = str(item.data(Qt.UserRole))
                item.setSelected(role in keep)
        finally:
            del blocker

    def on_prompt_group_selection_changed(self, group_ids: object) -> None:
        try:
            ids = [str(x) for x in (group_ids or []) if str(x)]
        except Exception:
            ids = []
        ready_paths = self.all_region_ready_image_paths()
        ready_stems = [p.stem for p in ready_paths]
        if "__ALL__" in ids:
            chosen = ready_stems
        else:
            chosen = [x for x in ids if x in ready_stems]
        # Preserve order of ready image list so generated outputs are stable.
        chosen_set = set(chosen)
        chosen = [s for s in ready_stems if s in chosen_set]
        if len(chosen) > 16:
            QMessageBox.warning(self, "超過上限", "Step 5 最終輸入組合一次最多只能選定 16 組；請取消部分組合後再繼續。")
            chosen = chosen[:16]
            self._set_prompt_group_selection(chosen)
        self.state.selected_region_stems = chosen
        self.state.prompt_group = "__SELECTED__" if len(chosen) != 1 else chosen[0]
        self.save_state()
        self.update_prompt_selection_status()
        self.update_actual_prompt_preview()
        self.mark_dirty(5)

    def current_prompt_group(self) -> str:
        stems = self.prompt_ready_stems()
        if stems:
            return "__SELECTED__" if len(stems) != 1 else stems[0]
        return self.state.prompt_group or ""

    def default_prompt_template(self) -> str:
        c = self.class_name()
        return (
            "你會收到兩張輸入圖像。\n"
            "Image 1 是未標註的原始待編輯影像，也是唯一要被實際編輯與保留風格的圖像。\n"
            "Image 2 是 ROI / Target Area 標註參考圖，只能用來判斷位置，不可被複製、描繪或當成輸出內容。\n\n"
            "請根據 Image 2 中 ROI 標示的位置，找到 Image 1 對應位置的原始瑕疵或指定物件，並將該位置自然還原成周圍一致的背景材質。\n"
            "請只把 ROI 中的瑕疵外觀當作參考特徵，刪除原位置的瑕疵，也不要保留 ROI 的任何標註痕跡。\n"
            f"接著請在 Image 2 中 Target Area 所標示的允許區域內，隨機生成 1~4 個與 ROI 來源特徵相似的 `{c}` 目標。\n"
            "新生成目標必須只出現在 Target Area 內；若目標屬於 burr，請讓它精準貼附在紙板邊緣或物件邊緣，不要漂浮在平面中央、背景或金屬導軌上。\n"
            "生成目標需小型、細微、不規則、低對比且自然，接近真實工業檢測影像中的微小邊緣毛邊、纖維破損、暗色小點或細微污染；可被檢測辨識，但不要誇張放大。\n"
            "請保持 Image 1 的背景、材質、紙纖維紋理、邊緣形狀、金屬導軌、陰影、模糊程度、光照、相機角度、透視、裁切、解析度與整體風格一致。\n"
            "禁止輸出或殘留任何 ROI 框線、Target Area 框線、藍色線、粉紅色線、半透明色塊、標籤文字、座標、箭頭、遮罩邊界、UI 標記、方形區塊、修補痕跡或全圖重繪痕跡。"
        )

    def apply_prompt_template(self) -> None:
        self.prompt_edit.setPlainText(self.default_prompt_template())
        self.update_actual_prompt_preview()
        self.mark_dirty(5)

    def refresh_prompt_groups(self) -> None:
        if not hasattr(self,"prompt_group_grid"):
            return
        current = self.state.prompt_group or ""
        selected_stems = self.prompt_ready_stems()
        self.prompt_group_grid.blockSignals(True)
        self.prompt_group_grid.load_groups(self.all_region_ready_image_paths(), self.regions_dir(), current, selected_stems)
        self.prompt_group_grid.blockSignals(False)
        self.update_prompt_selection_status()
        self.update_actual_prompt_preview()

    def _region_lines_for_stem(self, stem: str, include_name: bool = False) -> list[str]:
        txt = self.regions_dir() / f"{stem}.txt"
        if not txt.exists():
            return []
        raw = txt.read_text(encoding="utf-8", errors="ignore").strip()
        rois: list[str] = []
        for line in raw.splitlines():
            if re.match(r"^ROI(?:_\d+)?:", line):
                rois.append(line.split(":", 1)[1].strip())
        target_shapes = parse_target_areas(txt)
        lines: list[str] = []
        if include_name:
            lines.append(f"[{stem}]")
        def fmt_xyxy(raw_xyxy: str) -> str:
            vals = [v.strip() for v in raw_xyxy.split(",") if v.strip()]
            if len(vals) >= 4:
                return f"x1：{vals[0]}, y1：{vals[1]}, x2：{vals[2]}, y2：{vals[3]}"
            return raw_xyxy
        def fmt_rect(rect: tuple[int, int, int, int]) -> str:
            x1, y1, x2, y2 = rect
            return f"x1：{x1}, y1：{y1}, x2：{x2}, y2：{y2}"
        for i, roi in enumerate(rois, start=1):
            lines.append(f"ROI{i}: {fmt_xyxy(roi)}")
        for i, shape in enumerate(target_shapes, start=1):
            if shape.get("kind") == "polygon":
                pts = shape.get("points", [])
                point_text = "; ".join(f"x{j}：{int(x)}, y{j}：{int(y)}" for j, (x, y) in enumerate(pts, start=1))
                lines.append(f"Target Area{i}: polygon closed points = {point_text}")
            else:
                lines.append(f"Target Area{i}: {fmt_rect(target_area_bbox(shape))}")
        return lines

    def collect_region_context(self, selected: str = "") -> str:
        # Kept only for legacy preview/debug. Reference-guided mode does not append coordinates to the API prompt.
        if not selected:
            return ""
        if selected == "__SELECTED__":
            paths = self.selected_region_image_paths()
            blocks: list[str] = []
            include_name = len(paths) > 1
            for p in paths:
                lines = self._region_lines_for_stem(p.stem, include_name=include_name)
                if lines:
                    blocks.append("\n".join(lines))
            return "\n".join(blocks)
        if selected == "__ALL__":
            blocks: list[str] = []
            for p in self.all_region_ready_image_paths()[:16]:
                lines = self._region_lines_for_stem(p.stem, include_name=True)
                if lines:
                    blocks.append("\n".join(lines))
            return "\n".join(blocks)
        return "\n".join(self._region_lines_for_stem(selected, include_name=False))

    def build_actual_prompt(self) -> str:
        base = self.prompt_edit.toPlainText().strip() if hasattr(self,"prompt_edit") else (self.state.prompt_input or "").strip()
        # Reference-guided generation now sends Image 1 = clean original and
        # Image 2 = annotation reference.  ROI / Target Area coordinates are not
        # appended to the prompt, because location is conveyed visually by Image 2.
        return base

    def update_actual_prompt_preview(self) -> None:
        if hasattr(self, "actual_prompt_view"):
            text = self.build_actual_prompt().strip()
            self.actual_prompt_view.setPlainText(text)

    def validate_prompt_selection(self) -> bool:
        stems = self.prompt_ready_stems()
        if not stems:
            QMessageBox.warning(self, "Missing", "請在 Step 5『引用組別』中選定 1～16 組已完成 ROI / Target Area 的圖像組合。")
            return False
        if len(stems) > 16:
            QMessageBox.warning(self, "超過上限", "Step 5 最終輸入組合一次最多只能選定 16 組。")
            return False
        return True

    def save_prompt(self, show_message: bool=True) -> bool:
        if hasattr(self, "prompt_edit"):
            self.state.prompt_input = self.prompt_edit.toPlainText()
        self.update_actual_prompt_preview()
        self.update_state_from_widgets(); self.save_state()
        text=(self.state.prompt_input or "").strip()
        if not text:
            QMessageBox.warning(self,"Missing","Prompt 不能空白。") ; return False
        if show_message: QMessageBox.information(self,"Done","Prompt 已儲存。")
        return True

    # ---------- model ----------
    def parse_size_or_default(self, text: str) -> tuple[int,int]:
        m=re.match(r"^(\d+)x(\d+)$", str(text).lower())
        return (int(m.group(1)), int(m.group(2))) if m else (1280,1280)

    def update_size_controls(self) -> None:
        is_gpt2 = self.model_combo.currentText() == "gpt-image-2"
        same_as_original = is_gpt2 and hasattr(self, "gpt2_size_mode_combo") and self.gpt2_size_mode_combo.currentText() == "與原圖尺寸相同"
        self.size_label.setVisible(not is_gpt2); self.size_combo.setVisible(not is_gpt2)
        self.gpt2_size_mode_label.setVisible(is_gpt2); self.gpt2_size_mode_combo.setVisible(is_gpt2)
        for w in [self.width_label, self.width_edit, self.height_label, self.height_edit, self.size_warning]:
            w.setVisible(is_gpt2)
        for edit in [self.width_edit, self.height_edit]:
            edit.setReadOnly(same_as_original)
            edit.setFocusPolicy(Qt.NoFocus if same_as_original else Qt.StrongFocus)
            edit.setCursor(StepButton.forbidden_cursor() if same_as_original else QCursor(Qt.IBeamCursor))
            edit.setStyleSheet("background:#e5e7eb;color:#64748b;" if same_as_original else "")
        if same_as_original:
            w, h = self.first_input_or_raw_image_size()
            if w and h:
                self.width_edit.setText(str(w)); self.height_edit.setText(str(h))
            self.size_warning.setText(GPT2_SIZE_LIMIT_HINT)
        else:
            self.size_warning.setText(GPT2_SIZE_LIMIT_HINT)
        

    def first_input_or_raw_image_size(self) -> tuple[int, int]:
        candidates = list_images(self.inputs_dir()) or list_images(self.raw_dir())
        if not candidates:
            return self.parse_size_or_default(self.state.size)
        try:
            with Image.open(candidates[0]) as img:
                return int(img.width), int(img.height)
        except Exception:
            return self.parse_size_or_default(self.state.size)

    def current_size(self) -> str:
        if not hasattr(self,"model_combo"):
            return self.state.size
        if self.model_combo.currentText()=="gpt-image-2":
            if hasattr(self, "gpt2_size_mode_combo") and self.gpt2_size_mode_combo.currentText() == "與原圖尺寸相同":
                return SAME_AS_ORIGINAL_SIZE
            return f"{self.width_edit.text()}x{self.height_edit.text()}"
        return self.size_combo.currentText()

    def gpt2_size_error(self, w: int, h: int) -> str:
        try:
            w = int(w); h = int(h)
        except Exception:
            return "寬度與高度必須是整數。"
        if w <= 0 or h <= 0:
            return "寬度與高度必須大於 0。"
        if w % 16 != 0 or h % 16 != 0:
            return f"寬度與高度都必須是 16 的倍數；目前為 {w}×{h}。"
        long_edge = max(w, h)
        short_edge = max(1, min(w, h))
        if long_edge > 3840:
            return f"長邊不可超過 3840 px；目前長邊為 {long_edge} px。"
        ratio = long_edge / short_edge
        if ratio > 3.0:
            return f"長寬比不可超過 3:1；目前比例約為 {ratio:.2f}:1。"
        pixels = w * h
        if pixels < 655_360:
            return f"總像素 {pixels:,} px 低於最低限制 655,360 px。"
        if pixels > 8_294_400:
            return f"總像素 {pixels:,} px 高於最高限制 8,294,400 px。"
        return ""

    def show_gpt2_size_limit_dialog(self, w: int, h: int, reason: str, *, blocking: bool = True) -> None:
        title = "GPT-image-2 輸出尺寸不支援" if blocking else "GPT-image-2 尺寸提醒"
        action = (
            "此尺寸已確認不符合限制，因此系統會阻止繼續下一步。\n"
            "請回到 Step 6 調整寬度與高度後，再重新按『確認參數並估算本次成本』。"
            if blocking else
            "此模式會先以 API 支援尺寸生成，再將最終圖片回存為原圖尺寸。"
        )
        QMessageBox.warning(
            self,
            title,
            f"目前輸出尺寸：{w}×{h}\n"
            f"不支援原因：{reason}\n\n"
            "GPT-image-2 尺寸限制：\n"
            "1. 寬度與高度都必須是 16 的倍數。\n"
            "2. 長邊不可超過 3840 px。\n"
            "3. 長寬比不可超過 3:1。\n"
            "4. 總像素需介於 655,360 px～8,294,400 px。\n\n"
            f"{action}"
        )

    def validate_current_size(self) -> bool:
        """Validate GPT-image-2 custom output size.

        Tested invalid dimensions such as 640×640 should be blocked before
        Step 8, because the API rejects them. The only exception is the
        UI-level 'same as original' mode: that mode requests an API-safe
        intermediate size and then resizes the final saved file back to the
        source dimensions.
        """
        if str(self.model_combo.currentText()).strip() != "gpt-image-2":
            return True
        if hasattr(self, "gpt2_size_mode_combo") and str(self.gpt2_size_mode_combo.currentText()).strip() == "與原圖尺寸相同":
            w, h = self.first_input_or_raw_image_size()
            if not (w and h):
                QMessageBox.warning(self, "Missing size", "找不到可用的原圖尺寸，請先上傳或裁切圖片。")
                return False
            reason = self.gpt2_size_error(w, h)
            base_style = "background:#e5e7eb;color:#64748b;"
            bad_style = base_style + "border: 2px solid #ef4444;" if reason else base_style
            for edit in [self.width_edit, self.height_edit]:
                edit.setStyleSheet(bad_style)
            if reason:
                self.state.estimated_run_cost_usd = 0.0
                self.save_state()
                self.show_gpt2_size_limit_dialog(w, h, "原圖尺寸模式不支援：" + reason, blocking=True)
                return False
            return True
        try:
            w = int(self.width_edit.text())
            h = int(self.height_edit.text())
        except Exception:
            QMessageBox.warning(self, "Invalid size", "寬度與高度必須是整數。")
            return False
        reason = self.gpt2_size_error(w, h)
        bad = bool(reason)
        style = "border: 2px solid #ef4444;" if bad else ""
        self.width_edit.setStyleSheet(style)
        self.height_edit.setStyleSheet(style)
        if bad:
            self.state.estimated_run_cost_usd = 0.0
            self.save_state()
            self.show_gpt2_size_limit_dialog(w, h, reason, blocking=True)
            return False
        return True

    def fetch_openai_image2_pricing(self) -> dict[str, float]:
        """Best-effort live pricing scrape with local fallback.

        The UI can run offline, so this never blocks the workflow permanently.
        OpenAI's pricing page remains the source of truth; parsing is intentionally
        conservative and falls back to the embedded public snapshot when needed.
        """
        try:
            req = urllib.request.Request(OPENAI_PRICING_URL, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=8) as resp:
                html = resp.read().decode("utf-8", errors="ignore")
            section = html
            m = re.search(r"GPT-Image-2(.{0,2500})", html, flags=re.I | re.S)
            if m:
                section = m.group(1)
            prices = [float(x) for x in re.findall(r"\$\s*([0-9]+(?:\.[0-9]+)?)\s*/\s*1M\s*tokens", section, flags=re.I)]
            # Expected order on the pricing page: image input, image cached input, image output, text input, text cached input.
            if len(prices) >= 5:
                parsed = {
                    "image_input_tokens": prices[0],
                    "image_cached_input_tokens": prices[1],
                    "image_output_tokens": prices[2],
                    "text_input_tokens": prices[3],
                    "text_cached_input_tokens": prices[4],
                }
                return parsed
        except Exception:
            pass
        return dict(DEFAULT_GPT_IMAGE2_PRICING_PER_1M)

    def estimate_model_cost(self) -> None:
        if not self.validate_current_size():
            return
        self.update_state_from_widgets()
        pricing = self.fetch_openai_image2_pricing() if self.state.model == "gpt-image-2" else dict(DEFAULT_GPT_IMAGE2_PRICING_PER_1M)
        quality_mult = {"low": 1.0, "medium": 1.8, "high": 3.0, "auto": 2.0}.get(self.state.quality, 1.0)
        w, h = self.first_input_or_raw_image_size() if self.state.size == SAME_AS_ORIGINAL_SIZE else self.parse_size_or_default(self.state.size)
        area_mult = max(0.25, (max(1, w) * max(1, h)) / (1280 * 1280)) if w and h else 1.0
        # Approximate tokens from recent gpt-image-2 edit runs in this project; rates are fetched live when possible.
        text_tokens = 450
        image_input_tokens = int(1521 * area_mult)
        image_output_tokens = int(233 * area_mult * quality_mult)
        est = (text_tokens/1_000_000.0)*pricing.get("text_input_tokens",5.0) + (image_input_tokens/1_000_000.0)*pricing.get("image_input_tokens",8.0) + (image_output_tokens/1_000_000.0)*pricing.get("image_output_tokens",30.0)
        est = round(est * max(1, int(self.state.num_outputs)), 6)
        self.state.estimated_run_cost_usd = est
        self.save_state()
        if hasattr(self, "estimated_cost_label"):
            self.estimated_cost_label.setText(f"預估成本：約 USD ${est:.6f}（已嘗試讀取 OpenAI 最新 pricing；實際以 API usage / 帳單為準）")
        QMessageBox.information(self, "成本預估", f"本次預估成本：約 USD ${est:.6f}\n費率來源：{OPENAI_PRICING_URL}（無法連線時使用本地快取/備援費率）。")

    def save_model_settings(self, show_message: bool=True) -> bool:
        if not self.validate_current_size(): return False
        self.update_state_from_widgets()
        if not self.state.estimated_run_cost_usd:
            self.estimate_model_cost()
        self.save_state()
        if show_message: QMessageBox.information(self,"Done","模型參數已儲存。")
        return True

    # ---------- aggregate ----------
    def aggregate_output_size_text(self) -> str:
        """Return the real numeric output size shown in Aggregate.

        The internal sentinel 'same_as_original' is a UI mode, not a useful
        confirmation value. Show the actual width*height that the user should
        expect as the final saved image size.
        """
        if str(self.state.size).strip().lower() == SAME_AS_ORIGINAL_SIZE:
            w, h = self.first_input_or_raw_image_size()
            if w and h:
                return f"{w}*{h}"
            return "尚未取得原圖尺寸"
        w, h = self.parse_size_or_default(self.state.size)
        return f"{w}*{h}"

    def build_aggregate_text(self) -> str:
        self.update_state_from_widgets()
        raw_count=len(list_images(self.raw_dir())); crop_count=len(list_images(self.inputs_dir()))
        selected_paths = self.selected_region_image_paths()
        selected_count = len(selected_paths)
        regions = self.state.regions if isinstance(self.state.regions, dict) else {}
        annotated=sum(1 for p in selected_paths if ((regions.get(p.stem, {}).get("rois") or regions.get(p.stem, {}).get("targets")) or (self.reference_dir()/f"{p.stem}.png").exists()))
        prompt=self.build_actual_prompt() if hasattr(self,"prompt_edit") else ""
        lines=[
            "GPT GenImage UI - Aggregate Settings",
            f"確認時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "[專案]",
            f"Project：{self.state.project_name or '-'}",
            f"Class Name：{self.state.class_name}",
            f"OpenAI API Key：{'已設定' if self.read_env_key() else '未設定'}",
            "",
            "[資料與裁切]",
            f"原始圖片數量：{raw_count}",
            f"Step 4 輸入圖像數量：{crop_count}",
            f"已選定生成組數：{selected_count}",
            f"已選定且完成 ROI / Target Area 的圖像數量：{annotated}",
            f"裁切尺寸：{self.state.crop_width}*{self.state.crop_height}",
            "",
            "[Prompt]",
            f"Prompt 模式：{self.state.prompt_mode}",
            f"模板樣式：{self.state.prompt_template if self.state.prompt_mode == 'template' else '-'}",
            "實際 Prompt：",
            prompt or "<empty>",
            "",
            "[模型參數]",
            f"模型：{self.state.model}",
            f"品質：{self.state.quality}",
            f"輸出尺寸：{self.aggregate_output_size_text()}",
            f"輸出張數：{self.state.num_outputs}",
            f"輸出資料夾名稱：{self.state.run_name or '<auto>'}",
            f"預估成本 USD：{self.state.estimated_run_cost_usd:.6f}" if self.state.estimated_run_cost_usd else "預估成本 USD：尚未估算",
        ]
        return "\n".join(lines)

    def refresh_aggregate_summary(self) -> None:
        if hasattr(self,"aggregate_box"): self.aggregate_box.setPlainText(self.build_aggregate_text())

    # ---------- run ----------
    def check_generation_runtime(self) -> bool:
        """Validate the exact Python interpreter used by the UI before Step 8."""
        env = os.environ.copy()
        key = self.read_env_key()
        if key:
            env["OPENAI_API_KEY"] = key
        cmd = [sys.executable, "-c", "import sys; from openai import OpenAI; import openai; print(sys.executable); print(getattr(openai, '__version__', 'unknown'))"]
        proc = subprocess.run(cmd, cwd=self.root, capture_output=True, text=True, encoding="utf-8", errors="replace", env=env)
        if proc.returncode == 0:
            return True
        detail = (proc.stdout or "") + (proc.stderr or "")
        install_cmd = f'"{sys.executable}" -m pip install -r "{self.root.parent.parent / "requirements.txt"}"'
        QMessageBox.critical(
            self,
            "Generation runtime not ready",
            "目前啟動 UI 的 Python 環境缺少 Step 8 所需套件，因此不會開始生成，也不會把 Step 8 標成完成。\n\n"
            f"Python：{sys.executable}\n"
            "缺少或無法匯入：openai\n\n"
            "請在同一個環境執行：\n"
            f"{install_cmd}\n\n"
            "原始錯誤：\n" + self.scrub_text(detail[-1600:])
        )
        self.state.generation_status = "dependency_error"
        self.state.last_generation_return_code = int(proc.returncode)
        self.state.last_generation_error = self.scrub_text(detail[-2000:])
        self.save_state()
        return False

    def _resolve_generated_path(self, raw_path: str, meta_path: Path) -> Optional[Path]:
        if not raw_path:
            return None
        s = str(raw_path).strip().strip('"')
        normalized = s.replace("\\", os.sep).replace("/", os.sep)
        candidates: list[Path] = []
        p0 = Path(normalized)
        candidates.append(p0)
        if not p0.is_absolute():
            candidates.extend([meta_path.parent / p0, self.root / p0, self.project_dir() / p0])
            # Logs/metadata from older Windows runs may start with \project\...
            stripped = normalized.lstrip("\\/")
            candidates.extend([self.root / stripped, self.root.parent / stripped])
        for c in candidates:
            try:
                if c.exists() and c.is_file() and c.suffix.lower() in SUPPORTED_EXTS:
                    return c.resolve()
            except Exception:
                pass
        return None

    def current_log_tail(self, limit: int = 10000) -> str:
        """Return recent Step 8 log text for error analysis without exposing secrets."""
        try:
            if hasattr(self, "log_box"):
                return self.scrub_text(self.log_box.toPlainText()[-max(1000, int(limit)):])
        except Exception:
            pass
        return ""

    def analyze_generation_failure(self, log_text: str, return_code: int | None = None) -> tuple[str, str, str]:
        """Convert backend errors into actionable UI guidance.

        Returns (dialog_title, dialog_message, short_status_message). The parser
        intentionally keys off stable error markers and OpenAI API fields rather
        than a single exact traceback format.
        """
        text = self.scrub_text(str(log_text or ""))
        lower = text.lower()
        size_patterns = [
            "[user_error][size]",
            "invalid size",
            "param': 'size'",
            'param": "size"',
            "longest edge",
            "multiples of 16",
            "aspect ratio",
            "total pixels",
            "invalid_value",
        ]
        if any(p in lower for p in size_patterns):
            m = re.search(r"Invalid size ['\"]?([0-9]+x[0-9]+|[0-9]+\*[0-9]+)['\"]?", text, flags=re.I)
            bad_size = m.group(1).replace("*", "x") if m else (self.aggregate_output_size_text() if hasattr(self, "aggregate_output_size_text") else str(self.state.size))
            title = "輸出尺寸不支援"
            message = (
                f"生成失敗原因：目前輸出尺寸 {bad_size} 不符合 GPT-image-2 API 的尺寸限制。\n\n"
                "常見限制包含：\n"
                "1. 長邊不可超過 3840 px。\n"
                "2. 寬與高需要是 16 的倍數。\n"
                "3. 長寬比不可超過 3:1。\n"
                "4. 總像素需落在模型支援範圍內。\n\n"
                "處理方式：請回到 Step 6，改成較小且合法的輸出尺寸後重新執行生成。\n"
                "若你選的是『與原圖尺寸相同』，新版程式會先用 API 支援尺寸生成，再把最終檔案回存為原圖尺寸；若仍看到此錯誤，請確認你已使用最新版本重新啟動 UI。\n\n"
                "錯誤摘要：\n" + text[-1600:]
            )
            return title, message, "生成失敗：輸出尺寸不支援，請回到 Step 6 修正。"
        if "no module named 'openai'" in lower or "missing or incompatible dependency: openai" in lower:
            title = "缺少 OpenAI 套件"
            message = (
                "生成失敗原因：目前啟動 UI 的 Python 環境缺少 openai 套件。\n\n"
                "請在同一個環境執行：\n"
                f'"{sys.executable}" -m pip install -r "{self.root.parent.parent / "requirements.txt"}"\n\n'
                "錯誤摘要：\n" + text[-1600:]
            )
            return title, message, "生成失敗：缺少 openai 套件。"
        if "openai_api_key" in lower or "api key" in lower or "authentication" in lower:
            title = "API Key / 驗證錯誤"
            message = (
                "生成失敗原因：OpenAI API Key 未設定、無效，或目前環境沒有正確讀取。\n\n"
                "處理方式：請回到 Step 0 重新儲存共用 API Key，再回到 Step 8 執行生成。\n\n"
                "錯誤摘要：\n" + text[-1600:]
            )
            return title, message, "生成失敗：API Key / 驗證錯誤。"
        title = "生成失敗"
        message = (
            f"生成程序回傳錯誤 code={return_code if return_code is not None else '-'}。\n"
            "請查看 Step 8 log，依錯誤內容回到對應步驟修正後重新執行。\n\n"
            "錯誤摘要：\n" + (text[-2000:] if text else "無可用 log。")
        )
        return title, message, f"生成失敗：return code={return_code if return_code is not None else '-'}，請查看 log。"

    def show_generation_failure_dialog(self, return_code: int | None = None) -> None:
        log_text = self.current_log_tail()
        title, message, status_msg = self.analyze_generation_failure(log_text, return_code)
        self.state.last_generation_error = self.scrub_text(message[-3000:])
        self.set_generation_progress(self.generation_completed, status_msg)
        try:
            QMessageBox.critical(self, title, message)
        except Exception:
            pass

    def export_run_names(self) -> list[str]:
        """Run folder names under runs/ that contain at least one generated image."""
        root = self.runs_dir()
        if not root.exists():
            return []
        names: list[str] = []
        for child in sorted(root.iterdir(), key=lambda p: p.name.lower()):
            if child.is_dir() and self.run_image_paths(child.name):
                names.append(child.name)
        return names

    def run_image_paths(self, run_name: str) -> list[Path]:
        """All generated images for one run, taken from its Gen_Images folder.

        Falls back to a filtered recursive search for legacy runs created before the
        Gen_Images layout existed."""
        run_dir = self.runs_dir() / sanitize_name(str(run_name))
        gen = run_dir / "Gen_Images"
        if gen.exists():
            return list_images(gen)
        if not run_dir.exists():
            return []
        found: list[Path] = []
        for p in run_dir.rglob("*"):
            if not (p.is_file() and p.suffix.lower() in SUPPORTED_EXTS):
                continue
            stem = p.stem.lower()
            if stem == "input" or "mask" in stem or "preview" in stem or "annotation" in stem:
                continue
            if re.match(r"^(edited|generated|repaired)_seed", p.name) or p.stem == p.parent.name or p.parent.name == "Gen_Images":
                found.append(p.resolve())
        return sorted(found, key=lambda x: x.name)

    def successful_output_files(self, metas: Optional[list[Path]] = None) -> list[Path]:
        """Return final generated image paths for the current run.

        With the Gen_Images run layout the current run name maps directly to
        runs/<run_name>/Gen_Images/.  When no current run is set, fall back to the
        most recently modified run folder so Step 9 still shows something."""
        paths: list[Path] = []
        run_name = sanitize_name(str(self.state.run_name or "").strip())
        if run_name:
            paths = self.run_image_paths(run_name)
        if not paths:
            names = self.export_run_names()
            if names:
                latest = max((self.runs_dir() / n for n in names), key=lambda d: d.stat().st_mtime if d.exists() else 0)
                paths = self.run_image_paths(latest.name)
        seen: set[str] = set()
        unique: list[Path] = []
        for path in sorted(paths, key=lambda x: x.stat().st_mtime if x.exists() else 0):
            key = str(path.resolve())
            if key not in seen:
                unique.append(path); seen.add(key)
        return unique

    def current_run_has_successful_outputs(self) -> bool:
        return bool(self.successful_output_files())


    def next_restart_run_name(self, base_name: str | None = None) -> str:
        """Return a non-overwriting run folder name for Restart Generation.

        The visible Step 6 value is the user's base run name. Restart creates a
        new batch folder every time with no separator: <base>1, <base>2,
        <base>3, ... across both runs/ and exports/ so completed outputs are
        never overwritten.
        """
        raw_base = sanitize_name(base_name or self.state.run_name or "run") or "run"
        used: set[str] = set()
        for root in (self.runs_dir(), self.exports_dir()):
            if root.exists():
                for child in root.iterdir():
                    if child.is_dir():
                        used.add(child.name)
        # If Step 6 currently shows the previously generated folder name
        # (for example run1 or run2), recover the original user base name so
        # the next restart becomes run2 / run3 instead of run11 / run21.
        base = raw_base
        m = re.match(r"^(?P<base>.+?)(?P<num>\d+)$", raw_base)
        if m and raw_base in used and f"{m.group('base')}1" in used:
            base = m.group("base")
        candidate = f"{base}1"
        if candidate not in used:
            return candidate
        n = 2
        while True:
            candidate = f"{base}{n}"
            if candidate not in used:
                return candidate
            n += 1

    def assign_restart_run_name(self) -> str:
        base = self.run_name_edit.text().strip() if hasattr(self, "run_name_edit") else self.state.run_name
        new_name = self.next_restart_run_name(base or "run")
        self.state.run_name = new_name
        if hasattr(self, "run_name_edit"):
            self._set_widget_value_blocked(self.run_name_edit, "setText", new_name)
        self.save_state()
        return new_name

    def build_generation_cmd(self, resume_existing: bool = False) -> list[str]:
        self.update_state_from_widgets()
        self.current_run_name(create_if_empty=True)
        self.save_state()
        cmd=[
            sys.executable,str(self.root/"scripts"/"batch_from_folders.py"),
            "--class-name",self.state.class_name,
            "--images-dir",str(self.inputs_dir()),
            "--reference-dir",str(self.reference_dir()),
            "--output-dir",str(self.project_dir()/"runs"),
            "--workflow","reference-guided-edit",
            "--model",self.state.model,"--size",self.state.size,"--quality",self.state.quality,
            "--num-outputs","1","--total-outputs",str(self.state.num_outputs),
            "--min-defects","1","--max-defects","4",
            "--prompt",(self.state.prompt_input or ""),
        ]
        # Honor the 引用組別 selection: only the chosen groups are generated. The
        # selection is passed inline (no .selected_region_stems.txt file on disk).
        stems = self.selected_region_stems()
        if stems:
            cmd += ["--selected-stems", *stems]
        if self.state.run_name:
            cmd += ["--run-name", self.state.run_name]
        if resume_existing:
            cmd.append("--resume-existing")
        return cmd

    def expected_generation_steps(self)->int:
        return max(1, int(self.state.num_outputs))

    def reset_generation_progress(self,total:int)->None:
        self.generation_total=max(1,total); self.generation_completed=0; self._generation_progress_buffer=""
        self.generation_progress_bar.setRange(0,self.generation_total); self.generation_progress_bar.setValue(0); self.generation_progress_bar.setFormat(f"0/{self.generation_total} 張 (%p%)"); self.generation_progress_label.setText(f"準備生成：共 {self.generation_total} 張。")

    def set_generation_progress(self,completed:int,message:str|None=None)->None:
        self.generation_completed=max(0,min(completed,max(1,self.generation_total))); self.generation_progress_bar.setValue(self.generation_completed); self.generation_progress_bar.setFormat(f"{self.generation_completed}/{max(1,self.generation_total)} 張 (%p%)")
        if message: self.generation_progress_label.setText(message)

    def update_generation_progress_from_text(self,text:str)->None:
        self._generation_progress_buffer += text
        lines=self._generation_progress_buffer.splitlines(keepends=True)
        if lines and not(lines[-1].endswith("\n") or lines[-1].endswith("\r")): self._generation_progress_buffer=lines.pop()
        else: self._generation_progress_buffer=""
        for line in lines:
            clean=line.strip()
            if re.search(r"^\[OK\]\s+(edited|generated|repaired):", clean) or re.search(r"^\[DONE\]\s+output\s+\d+/\d+\s+saved:", clean):
                self.set_generation_progress(self.generation_completed+1, f"已完成 {min(self.generation_completed+1,self.generation_total)} / {self.generation_total} 張。")

    def clear_current_run_artifacts_only(self) -> None:
        """Clear only the current run/export folder when reusing the same Run name."""
        run_name = str(self.state.run_name or "").strip()
        if not run_name:
            return
        for folder in [self.runs_dir() / sanitize_name(run_name), self.exports_dir() / sanitize_name(run_name)]:
            try:
                if folder.exists():
                    shutil.rmtree(folder)
            except Exception:
                pass

    def start_generation(self, resume: bool=False)->None:
        if self.is_generation_running():
            QMessageBox.warning(self, "Generating", "目前仍在生成中，請先等待完成或按『停止目前程序』後再開始新的生成。")
            return
        if not self.save_prompt(show_message=False) or not self.save_model_settings(show_message=False): return
        if not list_images(self.inputs_dir()): QMessageBox.warning(self,"Missing","請先完成裁切與 ROI / Target Area。") ; return
        selected_paths = self.selected_region_image_paths()
        if not selected_paths:
            QMessageBox.warning(self,"Missing","請先在 Step 5『引用組別』選定 1～16 組已完成 ROI / Target Area 的圖像。") ; return
        if len(selected_paths) > 16:
            QMessageBox.warning(self,"超過上限","Step 5 最終輸入組合一次最多只能選定 16 組。") ; return
        regions = self.state.regions if isinstance(self.state.regions, dict) else {}
        missing_ref = []
        for p in selected_paths:
            geo = regions.get(p.stem, {})
            has_geo = bool(geo.get("rois")) and bool(geo.get("targets"))
            if not (self.reference_dir()/f"{p.stem}.png").exists() and not has_geo:
                missing_ref.append(p.name)
        if missing_ref:
            QMessageBox.warning(self,"Missing", "新版參考圖模式需要每張圖像都有 ROI 與 Target Area 標註，才能產生 Image 2 參考圖。請為下列圖像繪製 ROI / Target Area：\n" + "\n".join(missing_ref[:10])) ; return
        if not self.read_env_key(): QMessageBox.warning(self,"Missing","OPENAI_API_KEY 尚未設定，請先到 Step 1 儲存共用 API Key。") ; return
        if not self.check_generation_runtime(): return
        if not resume:
            new_run_name = self.assign_restart_run_name()
            self.status_label.setText(f"Status: new run folder prepared: {new_run_name}")
        else:
            self.current_run_name(create_if_empty=True)
        self.state.generation_status = "running"
        self.state.last_generation_return_code = 0
        self.state.last_generation_error = ""
        self.generation_started_step = int(self.current_step)
        self.save_state()
        self._stop_requested = False
        self.log_event(f"generation start (resume={resume}, total={self.state.num_outputs}, run={self.state.run_name})")
        self.set_generation_ui_locked(True)
        cmd=self.build_generation_cmd(resume_existing=resume)
        existing_completed = min(len(self.successful_output_files()), self.expected_generation_steps()) if resume else 0
        self.reset_generation_progress(self.expected_generation_steps())
        if resume and existing_completed:
            self.set_generation_progress(existing_completed, f"已找到 {existing_completed} 張既有輸出，將從上次進度繼續。")
        elif resume:
            self.set_generation_progress(0, "未找到既有輸出，將從第 1 張開始生成。")
        else:
            self.set_generation_progress(0, f"已建立新的 run：{self.state.run_name}，準備重新開始生成。")
        self.command_preview.setPlainText(self.scrub_text(" ".join(f'"{c}"' if " " in c else c for c in cmd)))
        def _after_generation():
            self.update_actual_generation_cost()
            if getattr(self, "current_step", None) == 9:
                self.refresh_outputs()
            outputs = self.successful_output_files()
            if outputs:
                self.state.generation_status = "success"
                self.state.completed_steps[8] = True
                self.dirty_steps[8] = False
                self.save_state()
                self.update_step_buttons()
                self.set_generation_progress(min(len(outputs), self.generation_total), f"生成完成：找到 {len(outputs)} 張輸出圖。")
            else:
                self.state.generation_status = "no_outputs"
                self.state.completed_steps[8] = False
                self.dirty_steps[8] = True
                self.save_state()
                self.update_step_buttons()
                self.set_generation_progress(0, "生成程序結束，但沒有找到任何輸出圖；Step 8 不會被標記完成。")
                QMessageBox.warning(self, "No output", "生成程序結束，但沒有找到任何輸出圖。請查看下方 log。")
        self.run_command(cmd,"generation",on_finished=_after_generation)

    def _current_run_summary_file(self) -> Optional[Path]:
        run_name = sanitize_name(str(self.state.run_name or "").strip())
        if not run_name:
            return None
        run_dir = self.runs_dir() / run_name
        for fname in ("generation_summary.xlsx", "generation_summary.csv"):
            p = run_dir / fname
            if p.exists():
                return p
        return None

    def update_actual_generation_cost(self) -> None:
        """Sum the Cost($USD) column of the current run's summary table."""
        total = 0.0
        found = False
        summary = self._current_run_summary_file()
        if summary is not None:
            try:
                if summary.suffix.lower() == ".xlsx":
                    from openpyxl import load_workbook
                    wb = load_workbook(summary, read_only=True, data_only=True)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if row and len(row) >= 4 and row[3] not in (None, ""):
                            try:
                                total += float(row[3]); found = True
                            except Exception:
                                pass
                    wb.close()
                else:
                    import csv
                    with open(summary, encoding="utf-8-sig", newline="") as fh:
                        reader = csv.reader(fh); next(reader, None)
                        for row in reader:
                            if len(row) >= 4 and row[3] not in (None, ""):
                                try:
                                    total += float(row[3]); found = True
                                except Exception:
                                    pass
            except Exception:
                found = False
        self.state.actual_run_cost_usd = round(total, 8) if found else 0.0
        self.save_state()
        if hasattr(self, "actual_cost_label"):
            if found:
                self.actual_cost_label.setText(f"本次實際成本：約 USD ${self.state.actual_run_cost_usd:.6f}（依 API 回傳 usage 估算）")
            else:
                self.actual_cost_label.setText("本次實際成本：API 未回傳 usage，請以平台帳單為準")

    def stop_process(self) -> None:
        if self.current_process and self.current_process.state() != QProcess.NotRunning and self.current_process_label == "generation":
            # Graceful pause: tell the batch process to stop BEFORE sending the next image
            # by writing a single "STOP" line to its stdin (no sentinel file on disk).
            # Do NOT kill the process — the in-flight image finishes so its API call (and cost) is not wasted.
            self._stop_requested = True
            try:
                self.current_process.write(b"STOP\n")
                self.current_process.waitForBytesWritten(1000)
            except Exception:
                pass
            self.status_label.setText("Status: stop requested — finishing current image")
            self.log_event("generation stop requested by user")
            self._show_receiving_dialog()
        elif self.current_process and self.current_process.state() != QProcess.NotRunning:
            self.current_process.kill(); self.status_label.setText("Status: process stopped"); self.set_generation_ui_locked(False)

    def _show_receiving_dialog(self) -> None:
        try:
            self._close_receiving_dialog()
            dlg = QMessageBox(self); dlg.setIcon(QMessageBox.Information); dlg.setWindowTitle("暫停中")
            dlg.setText("目前生成圖像尚在接收中...")
            dlg.setStandardButtons(QMessageBox.StandardButton.NoButton)
            self._receiving_dialog = dlg; dlg.show()
            app = QApplication.instance()
            if app is not None: app.processEvents()
        except Exception:
            pass

    def _close_receiving_dialog(self) -> None:
        dlg = getattr(self, "_receiving_dialog", None)
        if dlg is not None:
            try: dlg.close(); dlg.deleteLater()
            except Exception: pass
            self._receiving_dialog = None

    # ---------- export ----------
    def find_run_metadata(self)->list[Path]:
        root=self.runs_dir(); return sorted(root.rglob("metadata.json"),key=lambda p:p.stat().st_mtime,reverse=True) if root.exists() else []

    def _read_meta_file(self, meta_path: Path) -> dict:
        try:
            return json.loads(meta_path.read_text(encoding="utf-8"))
        except Exception:
            return {}

    def _meta_group_key(self, meta_path: Path) -> str:
        meta = self._read_meta_file(meta_path)
        batch = str(meta.get("batch_run_name") or "").strip()
        if batch:
            return batch
        parent = meta_path.parent.name
        if "_seed" in parent:
            return parent.split("_seed", 1)[0]
        return parent

    def _meta_matches_run(self, meta_path: Path, run_name: str) -> bool:
        if not run_name:
            return False
        meta = self._read_meta_file(meta_path)
        batch = str(meta.get("batch_run_name") or "").strip()
        parent = meta_path.parent.name
        grandparent = meta_path.parent.parent.name if meta_path.parent.parent else ""
        return batch == run_name or parent == run_name or grandparent == run_name or parent.startswith(run_name + "_")

    def filtered_run_metadata(self) -> list[Path]:
        metas = self.find_run_metadata()
        if not metas:
            return []
        self.update_state_from_widgets()
        if getattr(self.state, "export_scope", "current") == "all":
            return metas
        run_name = str(self.state.run_name or "").strip()
        if run_name:
            selected = [m for m in metas if self._meta_matches_run(m, run_name)]
            if selected:
                return selected
        # Fallback: latest batch group.
        latest_group = self._meta_group_key(metas[0])
        return [m for m in metas if self._meta_group_key(m) == latest_group]

    def selected_export_runs(self) -> list[str]:
        combo = getattr(self, "export_runs_combo", None)
        if combo is None:
            return []
        return [str(n) for n in combo.checked_runs() if n]

    def refresh_export_runs_combo(self) -> None:
        if not hasattr(self, "export_runs_combo"):
            return
        runs = self.export_run_names()
        previously = set(self.selected_export_runs())
        self.export_runs_combo.set_runs(runs)
        keep = [r for r in runs if r in previously]
        if not keep:
            cur = sanitize_name(str(self.state.run_name or "").strip())
            if cur in runs:
                keep = [cur]
        if keep:
            self.export_runs_combo.set_checked(keep)

    def _collect_selected_run_images(self) -> tuple[list[str], list[Path]]:
        runs = self.selected_export_runs()
        images: list[Path] = []
        seen: set[str] = set()
        for rn in runs:
            for p in self.run_image_paths(rn):
                key = str(p.resolve())
                if key not in seen:
                    seen.add(key); images.append(p)
        return runs, images

    def load_selected_runs_preview(self) -> None:
        """Confirm button: load every selected run's images into the left sidebar."""
        self.update_state_from_widgets()
        runs, images = self._collect_selected_run_images()
        if not runs:
            QMessageBox.warning(self, "Missing", "請先在『匯出範圍』勾選至少一個 run。")
            return
        self.output_list.clear()
        if not images:
            self.output_preview.clear("所選 runs 沒有任何生成圖像。")
            return
        progress = QProgressDialog("生成圖像正在 loading...", "", 0, len(images), self)
        progress.setWindowTitle("Loading")
        progress.setCancelButton(None)
        progress.setWindowModality(Qt.WindowModal)
        progress.setMinimumDuration(0)
        progress.setValue(0)
        app = QApplication.instance()
        icon_size = self.output_list.iconSize()
        for i, p in enumerate(images, start=1):
            item = QListWidgetItem(elide_middle_stars(f"{p.parent.parent.name} / {p.name}", 34))
            item.setData(Qt.UserRole, str(p))
            item.setTextAlignment(Qt.AlignCenter)
            try:
                item.setIcon(cached_thumbnail_icon(p, icon_size))
            except Exception:
                pass
            item.setToolTip(str(p))
            self.output_list.addItem(item)
            progress.setValue(i)
            if app is not None:
                app.processEvents()
        progress.close()
        if self.output_list.count():
            self.output_list.setCurrentRow(0)
        self.status_label.setText(f"Status: 已載入 {len(images)} 張預覽圖（{len(runs)} 個 run）。")

    def export_dataset(self, show_message: bool=True)->bool:
        """Copy the selected runs' generated images into a user-chosen folder.

        The destination path is always chosen by the user. A folder named
        <project_name>-<timestamp> is created there holding every image; when the
        zip option is ticked the folder is packed into <project_name>-<timestamp>.zip
        instead of being left loose."""
        self.update_state_from_widgets()
        runs, images = self._collect_selected_run_images()
        if not runs:
            QMessageBox.warning(self, "Missing", "請先在『匯出範圍』勾選至少一個 run。")
            return False
        if not images:
            QMessageBox.warning(self, "Missing", "所選 runs 沒有任何生成圖像可匯出。")
            return False
        dest = QFileDialog.getExistingDirectory(self, "選擇匯出路徑")
        if not dest:
            return False
        project_label = sanitize_name(self.state.project_name or self.state.project_id or "export") or "export"
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        folder_name = f"{project_label}-{stamp}"
        want_zip = bool(self.zip_check.isChecked()) if hasattr(self, "zip_check") else True
        try:
            dest_root = Path(dest)
            out_dir = dest_root / folder_name
            if out_dir.exists():
                shutil.rmtree(out_dir)
            ensure_dir(out_dir)
            for p in images:
                target = out_dir / p.name
                if target.exists():
                    target = out_dir / f"{sanitize_name(p.parent.parent.name)}_{p.name}"
                shutil.copy2(p, target)
            result_path: Path = out_dir
            if want_zip:
                archive = shutil.make_archive(str(dest_root / folder_name), "zip", root_dir=str(dest_root), base_dir=folder_name)
                shutil.rmtree(out_dir, ignore_errors=True)
                result_path = Path(archive)
            self.state.completed_steps[9] = True
            self.dirty_steps[9] = False
            self.save_state(); self.update_step_buttons()
            if show_message:
                QMessageBox.information(self, "Done", f"已匯出 {len(images)} 張圖像至：\n{result_path}")
            return True
        except Exception as exc:
            QMessageBox.critical(self, "Export failed", self.scrub_text(str(exc)))
            return False

    def go_home_from_export(self)->None:
        """Return to Step 1 without deleting the current project or generated files."""
        self.update_state_from_widgets(); self.save_state(); self.refresh_project_list(); self.goto_step(0); self.status_label.setText("Status: returned to homepage")

    # ---------- subprocess and logs ----------
    def scrub_text(self,text:str)->str:
        if not text: return text
        root=str(self.root).replace("\\","/")
        out=text.replace(str(self.root),"").replace(root,"")
        out=re.sub(r"[A-Za-z]:[\\/][^\n\r\t ]+", "", out)
        out=re.sub(r"/[^\n\r\t ]*/GPT_GenImage_UI", "", out)
        out=re.sub(r"data/[0-9A-Za-z_./<>-]+", "", out)
        out=re.sub(r"configs/[0-9A-Za-z_./<>-]+", "", out)
        out=re.sub(r"runs/[0-9A-Za-z_./<>-]+", "", out)
        return out

    def run_quick_command_dialog(self, cmd:list[str], label:str, on_success:Optional[Callable]=None)->bool:
        try:
            env=os.environ.copy()
            key=self.read_env_key()
            if key:
                env["OPENAI_API_KEY"] = key
            proc=subprocess.run([str(x) for x in cmd],cwd=self.root,capture_output=True,text=True,encoding="utf-8",errors="replace",check=False, env=env)
            output=self.scrub_text((proc.stdout or "")+(proc.stderr or ""))
            if proc.returncode==0:
                if on_success: on_success()
                QMessageBox.information(self,label,output.strip() or "完成。"); return True
            QMessageBox.critical(self,label,output[-4000:] or "Command failed."); return False
        except Exception as exc:
            QMessageBox.critical(self,label,str(exc)); return False

    def run_command(self, cmd:list[str], label:str, on_finished:Optional[Callable]=None)->None:
        if self.current_process and self.current_process.state()!=QProcess.NotRunning: QMessageBox.warning(self,"Running","已有程序執行中。"); return
        if label == "generation":
            self.set_generation_ui_locked(True)
        self.log_box_append("\n===== " + label + " =====\n[RUN] " + " ".join(str(c) for c in cmd) + "\n")
        self.status_label.setText(f"Status: running - {label}")
        p=QProcess(self); p.setWorkingDirectory(str(self.root)); qenv=QProcessEnvironment.systemEnvironment(); qenv.insert("PYTHONIOENCODING","utf-8"); qenv.insert("PYTHONUNBUFFERED","1"); qenv.insert("PYTHONDONTWRITEBYTECODE","1")
        key=self.read_env_key()
        if key:
            qenv.insert("OPENAI_API_KEY", key)
        p.setProcessEnvironment(qenv)
        self.current_process_label=label; p.readyReadStandardOutput.connect(self.safe_action("process_stdout", lambda: self._read_process(p,False))); p.readyReadStandardError.connect(self.safe_action("process_stderr", lambda: self._read_process(p,True)))
        def finished(code,status):
            try:
                self.log_box_append(f"\n===== {label} finished: code={code} =====\n")
                self.status_label.setText("Status: completed" if code==0 else f"Status: error code={code}")
                generation_finished_away = bool(label == "generation" and self.current_step != 8)
                if label == "generation":
                    self._close_receiving_dialog()
                    self.state.last_generation_return_code = int(code)
                    if getattr(self, "_stop_requested", False):
                        self._stop_requested = False
                        self._close_receiving_dialog()
                        self.state.generation_status = "paused"
                        done = self.generation_completed
                        remaining = max(0, self.generation_total - done)
                        self.state.completed_steps[8] = False
                        self.dirty_steps[8] = True
                        self.set_generation_progress(done, f"已暫停：完成 {done} / {self.generation_total} 張。")
                        try: self.update_actual_generation_cost()
                        except Exception: pass
                        self.save_state(); self.update_step_buttons()
                        self.log_event(f"generation paused at {done}/{self.generation_total}")
                        self.current_process = None; self.current_process_label = ""
                        self.set_generation_ui_locked(False)
                        QMessageBox.information(self, "已暫停", f"現在生成至第 {done} 張，還剩 {remaining} 張尚未生成，已暫停")
                        return
                    if code != 0:
                        self.state.generation_status = "failed"
                        self.state.completed_steps[8] = False
                        self.dirty_steps[8] = True
                        self.show_generation_failure_dialog(int(code))
                        self.save_state(); self.update_step_buttons()
                        self.log_event(f"generation failed (code={code})")
                    else:
                        self.set_generation_progress(max(1,self.generation_total),f"生成程序完成，正在檢查輸出圖。")
                        self.log_event("generation finished: success")
                self.current_process=None; self.current_process_label=""
                if label == "generation":
                    self.set_generation_ui_locked(False)
                else:
                    self.update_step_buttons()
                if code == 0 and on_finished:
                    on_finished()
                if label == "generation" and code == 0 and generation_finished_away:
                    QMessageBox.information(self, "生成完成", "背景生成程序已完成。請回到 Step 8 查看 log，或進入 Step 9 檢視輸出結果。")
            except Exception as exc:
                self._log_ui_exception(f"process_finished_{label}", exc)
                try:
                    QMessageBox.critical(self,"Process callback failed",f"程序完成後更新 UI 時發生錯誤：\n{type(exc).__name__}: {exc}")
                except Exception:
                    pass
        def proc_error(error):
            try:
                msg = f"{error}: {p.errorString()}"
                self.log_box_append(f"\n[QProcess ERROR] {msg}\n")
                self.status_label.setText("Status: process error")
                if label == "generation":
                    self.state.generation_status = "process_error"
                    self.state.last_generation_error = msg
                    self.state.completed_steps[8] = False
                    self.dirty_steps[8] = True
                    self.save_state(); self.set_generation_ui_locked(False)
            except Exception as exc:
                self._log_ui_exception(f"process_error_{label}", exc)
                if label == "generation":
                    self.set_generation_ui_locked(False)
        p.errorOccurred.connect(proc_error)
        if label == "generation":
            p.started.connect(self.safe_action("process_started_generation", lambda: self.set_generation_ui_locked(True)))
        p.finished.connect(finished); self.current_process=p; self.update_step_buttons(); p.start(str(cmd[0]), [str(x) for x in cmd[1:]])
        if label == "generation":
            self.set_generation_ui_locked(True)

    def _read_process(self,p:QProcess,err:bool)->None:
        data=p.readAllStandardError() if err else p.readAllStandardOutput(); text=bytes(data).decode("utf-8",errors="replace"); self.log_box_append(text); self.update_generation_progress_from_text(text) if self.current_process_label=="generation" else None

    def log_box_append(self,text:str)->None:
        text=self.scrub_text(text)
        if hasattr(self,"log_box"):
            try: end=QTextCursor.MoveOperation.End
            except AttributeError: end=QTextCursor.End
            self.log_box.moveCursor(end); self.log_box.insertPlainText(text); self.log_box.moveCursor(end)
        else: print(text,end="")

    # ---------- refresh / preview ----------
    def refresh_all(self, light: bool=False)->None:
        """Refresh only the active page during normal navigation.

        Large runs can contain hundreds of generated images.  Refreshing every
        list and rebuilding output thumbnails on every button click makes Qt's
        GUI thread decode many images unnecessarily, so light refreshes are
        scoped to the current step.
        """
        step = int(getattr(self, "current_step", 0))
        if step == 0:
            if hasattr(self,"project_cards_layout"): self.refresh_project_list()
        elif step == 2:
            if hasattr(self,"upload_list"): self.refresh_uploads()
        elif step == 3:
            if hasattr(self,"raw_thumb_grid"): self.refresh_raw_thumbs()
            if hasattr(self,"crop_done_grid"): self.refresh_crops()
        elif step == 4:
            if hasattr(self,"region_thumb_grid"): self.refresh_region_thumbs()
        elif step == 5:
            if hasattr(self,"prompt_group_grid"): self.refresh_prompt_groups()
        elif step == 7:
            if hasattr(self,"aggregate_box"): self.refresh_aggregate_summary()
        elif step == 9:
            if hasattr(self,"output_list"): self.refresh_outputs()
        self.update_step_buttons()

    def refresh_uploads(self, auto_select: bool=False, select_path: Optional[Path]=None)->None:
        self.upload_list.clear()
        for p in list_images(self.raw_dir()):
            item=QListWidgetItem(p.name); item.setData(Qt.UserRole,str(p)); self.upload_list.addItem(item)
        if select_path is not None:
            target=select_path.name
            for i in range(self.upload_list.count()):
                if Path(self.upload_list.item(i).data(Qt.UserRole)).name == target:
                    self.upload_list.setCurrentRow(i); return
        if auto_select and self.upload_list.count(): self.upload_list.setCurrentRow(0)
        elif not self.upload_list.selectedItems(): self.upload_preview.clear("尚無預覽，請選取圖片或直接拖曳圖片到此處。")

    def refresh_raw_thumbs(self)->None:
        self.raw_thumb_grid.load_paths(list_images(self.raw_dir()))

    def refresh_crops(self, auto_select: bool=False, select_path: Optional[Path]=None)->None:
        # The grid always shows every crop in inputs_dir together (not filtered by the
        # selected raw source). select_path selects a specific crop (e.g. the one just
        # made) so its preview/canvas update without changing which crops are listed.
        # Block signals during (re)load so the transient selection churn while clearing
        # the grid does not fire on_crop_done_selected and hijack the middle canvas.
        paths=list_images(self.inputs_dir())
        blocker=QSignalBlocker(self.crop_done_grid)
        try:
            self.crop_done_grid.load_paths(paths)
        finally:
            del blocker
        if select_path is not None:
            target=Path(select_path).name
            for i in range(self.crop_done_grid.count()):
                if Path(self.crop_done_grid.item(i).data(Qt.UserRole)).name == target:
                    self.crop_done_grid.setCurrentRow(i); return
        if auto_select and self.crop_done_grid.count(): self.crop_done_grid.setCurrentRow(self.crop_done_grid.count()-1)
        elif not paths and hasattr(self,"crop_done_preview"): self.crop_done_preview.clear("尚無裁切完成圖")

    def refresh_region_thumbs(self)->None:
        if not hasattr(self, "region_thumb_grid"):
            return
        paths = list_images(self.inputs_dir())
        valid = {str(p) for p in paths}
        blocker = QSignalBlocker(self.region_thumb_grid)
        try:
            self.region_thumb_grid.load_paths(paths)
        finally:
            del blocker
        cur = getattr(getattr(self, "region_canvas", None), "image_path", None)
        if not paths or (cur and str(Path(cur)) not in valid):
            if hasattr(self, "region_canvas"):
                self.region_canvas.set_image(None)
            if hasattr(self, "region_status"):
                self.region_status.setText("尚未選擇圖像")
        elif cur and hasattr(self, "region_canvas") and not (
            getattr(self.region_canvas, "rois", None) or getattr(self.region_canvas, "target_areas", None)
        ):
            # The canvas is empty but saved geometry exists for this image (e.g.
            # after reopening the project or revisiting Step 4 from a later step):
            # restore the ROI/Target boxes so they do not appear to have vanished.
            self.load_regions_into_canvas(Path(cur))
        self.step4_selected_view_stems = []
        self.update_region_selection_status()

    def refresh_outputs(self, auto_select: bool=False)->None:
        self.refresh_export_runs_combo()
        self.output_list.clear()
        clean = sorted(self.successful_output_files(), key=lambda x: x.stat().st_mtime if x.exists() else 0, reverse=True)
        if not clean and hasattr(self, "output_preview"):
            self.output_preview.clear("尚無預覽")
        if hasattr(self.output_list, "load_paths"):
            self.output_list.load_paths(clean[:MAX_INITIAL_THUMBNAILS])
        else:
            self.output_list.clear()
            for p in clean[:MAX_INITIAL_THUMBNAILS]:
                item=QListWidgetItem(elide_middle_stars(f"{p.parent.name} / {p.name}", 34)); item.setData(Qt.UserRole,str(p)); item.setTextAlignment(Qt.AlignCenter); self.output_list.addItem(item)
        if auto_select and self.output_list.count(): self.output_list.setCurrentRow(0)
        elif not self.output_list.selectedItems(): self.output_preview.clear("尚無預覽")

    def preview_upload_selection(self)->None:
        items=self.upload_list.selectedItems(); self.upload_preview.set_path(Path(items[0].data(Qt.UserRole)) if items else None, "尚無預覽，請選取圖片或直接拖曳圖片到此處。")

    def preview_output_selection(self)->None:
        items=self.output_list.selectedItems(); self.output_preview.set_path(Path(items[0].data(Qt.UserRole)) if items else None, "尚無預覽")


def main() -> None:
    root = project_root()
    app = QApplication(sys.argv)
    # Single shared log file at the repo root (replaces the old logs/ folder).
    app_log = root.parent.parent / "log.txt"
    try:
        crash_log = app_log.open("a", encoding="utf-8")
        faulthandler.enable(file=crash_log, all_threads=True)
        app._ui_crash_log = crash_log  # keep file handle alive for faulthandler
    except Exception:
        pass

    def _excepthook(exc_type, exc, tb):
        try:
            with app_log.open("a", encoding="utf-8") as f:
                f.write(f"\n===== {datetime.now().isoformat(timespec='seconds')} | uncaught =====\n")
                traceback.print_exception(exc_type, exc, tb, file=f)
        except Exception:
            pass
        try:
            QMessageBox.critical(None, "UI error", f"{exc_type.__name__}: {exc}")
        except Exception:
            pass

    sys.excepthook = _excepthook
    win = MainWindow()
    win.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()

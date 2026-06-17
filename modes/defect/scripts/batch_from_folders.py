#!/usr/bin/env python3
"""Batch-run OpenAI GPT Image 2 from paired image/mask/target_area folders.

Default input structure:
  data/01_inputs/<defect_type>/images/<name>.png
  data/01_inputs/<defect_type>/masks/<name>.png
  data/01_inputs/<defect_type>/target_area_masks/<name>_target_area.png

Use --workflow reference-guided-edit for clean image + annotation reference image + prompt only.
Use --workflow reference-guided-edit for the default mode: unannotated image + annotation reference image + prompt.
Use --workflow repair-and-random-generate for source mask + target_area guided relocation.
"""
from __future__ import annotations

import argparse
import importlib.util
import json
import re
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))
from run_gpt_image2 import (
    SUPPORTED_GPT_IMAGE_MODELS,
    load_user_prompt,
    project_root,
    sanitize_defect_type,
    sanitize_name,
    sanitize_path_component,
)

SUPPORTED = [".png", ".jpg", ".jpeg", ".webp", ".bmp", ".tif", ".tiff"]


def list_images(folder: Path) -> list[Path]:
    return sorted([p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in SUPPORTED]) if folder.exists() else []


def find_mask(image: Path, masks_dir: Path) -> Path | None:
    candidates = [masks_dir / f"{image.stem}{ext}" for ext in SUPPORTED]
    candidates += [masks_dir / f"{image.stem}_mask{ext}" for ext in SUPPORTED]
    candidates += [masks_dir / f"{image.stem}_source{ext}" for ext in SUPPORTED]
    for p in candidates:
        if p.exists():
            return p
    return None


def find_target_area(image: Path, target_dir: Path) -> Path | None:
    candidates = [target_dir / f"{image.stem}{ext}" for ext in SUPPORTED]
    candidates += [target_dir / f"{image.stem}_target_area{ext}" for ext in SUPPORTED]
    candidates += [target_dir / f"{image.stem}_area{ext}" for ext in SUPPORTED]
    for p in candidates:
        if p.exists():
            return p
    return None


def make_batch_run_name(base_run_name: str | None) -> str:
    """Return the parent batch folder name under runs/<class_name>/."""
    if base_run_name and base_run_name.strip():
        return sanitize_name(base_run_name)
    return "run_" + datetime.now().strftime("%Y%m%d_%H%M%S")


def make_child_run_name(image_stem: str, time_label: str, counter: int) -> str:
    """Create one child folder per generated image inside the batch run folder.

    Naming scheme: ``<image_stem>_<time_label>_<counter>`` where ``time_label`` is
    the fixed moment the generation step started (YYYY年MM月DD日HH時MM分) and
    ``counter`` starts at 1 and increases by one per generated image, so child
    names never repeat or overwrite earlier outputs."""
    return sanitize_path_component(f"{image_stem}_{time_label}_{counter}")


def child_run_has_successful_output(child_dir: Path) -> bool:
    """Return True when a child run already contains a final generated image.

    Detection is metadata-driven first (the recorded final_outputs/outputs paths),
    then falls back to the legacy ``edited_seed*`` prefixes and to the current UI
    scheme where the final image file base equals its child-run folder name."""
    if not child_dir.exists():
        return False
    meta_path = child_dir / "metadata.json"
    if meta_path.exists():
        try:
            data = json.loads(meta_path.read_text(encoding="utf-8"))
            for out in (data.get("final_outputs") or data.get("outputs") or []):
                if Path(str(out)).exists():
                    return True
        except Exception:
            pass
    final_prefixes = ("edited_seed", "generated_seed", "repaired_seed")
    for path in child_dir.rglob("*"):
        if path.is_file() and path.suffix.lower() in SUPPORTED:
            if path.name.startswith(final_prefixes) or path.stem == child_dir.name:
                return True
    return False


def resolve_size_for_image(size: str, image: Path) -> str:
    # Keep same_as_original as a semantic mode. run_gpt_image2.py will turn it
    # into an API-valid size and resize final outputs back to the source size.
    s = str(size).strip().lower()
    if s in {"same_as_original", "original", "source"}:
        return "same_as_original"
    return size


def original_size_args_if_needed(size: str, image: Path) -> list[str]:
    if str(size).strip().lower() not in {"same_as_original", "original", "source"}:
        return []
    from PIL import Image
    with Image.open(image) as img:
        return ["--final-width", str(img.width), "--final-height", str(img.height)]


def run_and_tee(cmd: list[str], log_path: Path) -> None:
    """Run a subprocess while writing exactly what we print to log.txt."""
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with open(log_path, "a", encoding="utf-8", newline="") as log:
        header = f"\n===== RUN START {datetime.now().isoformat(timespec='seconds')} =====\n"
        log.write(header)
        print(header, end="", flush=True)
        run_line = "[RUN] " + " ".join(cmd) + "\n"
        log.write(run_line)
        print(run_line, end="", flush=True)
        log.flush()

        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            bufsize=1,
        )
        assert proc.stdout is not None
        for line in proc.stdout:
            print(line, end="", flush=True)
            log.write(line)
            log.flush()
        return_code = proc.wait()
        footer = f"===== RUN END {datetime.now().isoformat(timespec='seconds')} return_code={return_code} =====\n"
        log.write(footer)
        print(footer, end="", flush=True)
        if return_code != 0:
            raise subprocess.CalledProcessError(return_code, cmd)


def read_usage_from_metadata(meta_path: Path) -> tuple:
    """Return (total_tokens, cost_usd) summed across api_calls in a run metadata file.

    Returns empty strings when usage was not reported, so the spreadsheet cell stays
    blank instead of showing a misleading zero."""
    if not meta_path.exists():
        return "", ""
    try:
        data = json.loads(meta_path.read_text(encoding="utf-8"))
    except Exception:
        return "", ""
    tokens_total = 0
    cost_total = 0.0
    have_tokens = False
    have_cost = False
    for call in data.get("api_calls", []) or []:
        if not isinstance(call, dict):
            continue
        usage = call.get("usage") if isinstance(call.get("usage"), dict) else {}
        tok = usage.get("total_tokens")
        if tok is not None:
            try:
                tokens_total += int(tok)
                have_tokens = True
            except Exception:
                pass
        cost = call.get("estimated_cost_usd")
        if cost is None:
            cost = usage.get("estimated_cost_usd")
        if cost is not None:
            try:
                cost_total += float(cost)
                have_cost = True
            except Exception:
                pass
    return (tokens_total if have_tokens else ""), (round(cost_total, 8) if have_cost else "")


def find_final_image(work_dir: Path, base_name: str, ext: str) -> Path | None:
    """Locate the final generated image inside a per-image work folder."""
    direct = work_dir / f"{base_name}.{ext}"
    if direct.exists():
        return direct
    meta_path = work_dir / "metadata.json"
    if meta_path.exists():
        try:
            data = json.loads(meta_path.read_text(encoding="utf-8"))
            for item in (data.get("final_outputs") or data.get("outputs") or []):
                p = Path(str(item))
                if not p.is_absolute():
                    p = work_dir / p
                if p.exists():
                    return p
        except Exception:
            pass
    for p in sorted(work_dir.glob(f"*.{ext}")):
        stem = p.stem.lower()
        if stem == "input" or "mask" in stem or "preview" in stem or "annotation" in stem:
            continue
        return p
    return None


SUMMARY_HEADER = ["原圖像名稱", "生成圖像名稱", "Total Token", "Cost($USD)"]


def read_existing_summary(batch_output_root: Path) -> list:
    """Read previously written summary rows (xlsx or csv) so resumed runs keep them."""
    rows: list = []
    xlsx = batch_output_root / "generation_summary.xlsx"
    csv_path = batch_output_root / "generation_summary.csv"
    if xlsx.exists():
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx, read_only=True, data_only=True)
            ws = wb.active
            for r in ws.iter_rows(min_row=2, values_only=True):
                if r is None:
                    continue
                vals = list(r) + ["", "", "", ""]
                rows.append(tuple(vals[:4]))
            wb.close()
            return rows
        except Exception:
            pass
    if csv_path.exists():
        try:
            import csv
            with open(csv_path, encoding="utf-8-sig", newline="") as fh:
                reader = csv.reader(fh)
                next(reader, None)
                for r in reader:
                    vals = list(r) + ["", "", "", ""]
                    rows.append(tuple(vals[:4]))
        except Exception:
            pass
    return rows


def write_summary_table(batch_output_root: Path, rows: list) -> None:
    """Write the per-image summary as Excel (.xlsx); fall back to .csv when openpyxl is missing."""
    try:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "summary"
        ws.append(SUMMARY_HEADER)
        for row in rows:
            ws.append([row[0], row[1], row[2], row[3]])
        wb.save(str(batch_output_root / "generation_summary.xlsx"))
        stale_csv = batch_output_root / "generation_summary.csv"
        if stale_csv.exists():
            try:
                stale_csv.unlink()
            except Exception:
                pass
        return
    except Exception as exc:
        print(f"[WARN] openpyxl unavailable, writing CSV summary instead: {exc}", flush=True)
    try:
        import csv
        with open(batch_output_root / "generation_summary.csv", "w", encoding="utf-8-sig", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(SUMMARY_HEADER)
            for row in rows:
                writer.writerow([row[0], row[1], row[2], row[3]])
    except Exception as exc:
        print(f"[WARN] could not write summary table: {exc}", flush=True)


def main() -> None:
    root = project_root()
    p = argparse.ArgumentParser(description="Batch OpenAI GPT Image 2 edits from data/01_inputs/<defect_type>/ folders.")
    group = p.add_mutually_exclusive_group(required=True)
    group.add_argument("--class-name", help="Class name for folder/config/output naming")
    group.add_argument("--defect-type", help="Backward-compatible alias for --class-name")
    p.add_argument("--images-dir", type=Path, default=None)
    p.add_argument("--masks-dir", type=Path, default=None)
    p.add_argument("--target-area-dir", type=Path, default=None)
    p.add_argument("--selected-stems-file", type=Path, default=None, help="Optional newline-delimited image stems to include in this batch. Used by UI Step 4 multi-selection.")
    p.add_argument("--workflow", choices=["reference-guided-edit", "prompt-only-edit", "repair-and-random-generate", "repair-only", "generate-only"], default="reference-guided-edit")
    p.add_argument("--placement-mode", choices=["fixed-mask", "random-in-target"], default="random-in-target")
    p.add_argument("--seed", type=int, default=5000)
    p.add_argument("--num-outputs", type=int, default=1, help="Per-image output count when --total-outputs is not used.")
    p.add_argument("--total-outputs", type=int, default=None, help="Total number of generated images across all input images. When set, jobs are distributed over inputs and stop at this total.")
    p.add_argument("--run-name", default=None, help="Custom output folder name under runs/<defect_type>/. Example: --run-name test_001")
    p.add_argument("--output-dir", type=Path, default=None, help="Root folder for run outputs. Default: project root/runs")
    p.add_argument(
        "--model",
        default="gpt-image-2",
        choices=SUPPORTED_GPT_IMAGE_MODELS,
        help="OpenAI GPT Image model. Choices: " + ", ".join(SUPPORTED_GPT_IMAGE_MODELS),
    )
    p.add_argument("--size", default="1280x1280")
    p.add_argument("--quality", default="high", choices=["low", "medium", "high", "auto"])
    p.add_argument("--output-format", default="png", choices=["png", "jpeg", "webp"])
    p.add_argument("--output-compression", type=int, default=None)
    p.add_argument("--background", default="auto", choices=["auto", "opaque"])
    p.add_argument("--prompt", default="", help="Direct Chinese or English prompt. Overrides prompt file.")
    p.add_argument("--prompt-file", type=Path, default=None, help="Optional .txt prompt file. Default: configs/defect_types/<defect_type>/prompt.txt")
    p.add_argument("--prompt-config", type=Path, default=None, help="Deprecated compatibility alias for --prompt-file")
    p.add_argument("--prompt-extra", default="")
    p.add_argument("--mask-dilate", type=int, default=0)
    p.add_argument("--mask-feather", type=int, default=3)
    p.add_argument("--auto-resize-multiple", type=int, default=16)
    p.add_argument("--min-defects", type=int, default=1)
    p.add_argument("--max-defects", type=int, default=3)
    p.add_argument("--random-scale-min", type=float, default=0.85)
    p.add_argument("--random-scale-max", type=float, default=1.15)
    p.add_argument("--placement-attempts", type=int, default=500)
    p.add_argument("--exclude-repair-padding", type=int, default=8)
    p.add_argument("--no-clip-mask-to-target", action="store_true")
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--resume-existing", action="store_true", help="Skip child runs in this batch run folder that already have final generated images.")
    args = p.parse_args()

    if not args.dry_run:
        try:
            from openai import OpenAI as _OpenAI  # noqa: F401
        except Exception as exc:
            raise SystemExit(
                "Missing or incompatible dependency: openai. Install project requirements in the same Python environment that launches the UI:\n"
                f"  {sys.executable} -m pip install -r {root / 'requirements.txt'}\n"
                f"Original import error: {type(exc).__name__}: {exc}"
            )

    defect_type = sanitize_defect_type(args.class_name or args.defect_type)
    images_dir = args.images_dir or root / "data" / "01_inputs" / defect_type / "images"
    masks_dir = args.masks_dir or root / "data" / "01_inputs" / defect_type / "masks"
    target_area_dir = args.target_area_dir or root / "data" / "01_inputs" / defect_type / "target_area_masks"

    if not images_dir.exists():
        raise SystemExit(f"Images folder not found: {images_dir}\nRun: python scripts/init_defect_type.py --class-name {defect_type}")

    images = list_images(images_dir)
    if args.selected_stems_file and args.selected_stems_file.exists():
        selected_stems = {line.strip() for line in args.selected_stems_file.read_text(encoding="utf-8", errors="ignore").splitlines() if line.strip()}
        if selected_stems:
            images = [img for img in images if img.stem in selected_stems]
    if not images:
        raise SystemExit(f"No images found in: {images_dir}")

    needs_mask = args.workflow in {"reference-guided-edit", "repair-and-random-generate", "repair-only"} or (args.workflow == "generate-only" and args.placement_mode == "fixed-mask")
    needs_target = args.workflow in {"reference-guided-edit", "repair-and-random-generate"} or (args.workflow == "generate-only" and args.placement_mode == "random-in-target")

    jobs = []
    missing = []
    missing_target = []
    for img in images:
        mask = find_mask(img, masks_dir) if masks_dir.exists() else None
        target_area = find_target_area(img, target_area_dir) if target_area_dir.exists() else None
        if needs_mask and mask is None:
            missing.append(img.name)
        elif needs_target and target_area is None:
            missing_target.append(img.name)
        else:
            jobs.append((img, mask, target_area))

    if missing:
        hint = "ROI masks are required to render Image 2 annotation references." if args.workflow == "reference-guided-edit" else "Missing matching masks."
        raise SystemExit(f"{hint} Missing for {len(missing)} image(s) in {masks_dir}:\n" + "\n".join(missing[:20]))
    if missing_target:
        hint = "Target Area masks are required to render Image 2 annotation references." if args.workflow == "reference-guided-edit" else "This workflow requires target-area masks."
        raise SystemExit(f"{hint} Missing for {len(missing_target)} image(s) in {target_area_dir}:\n" + "\n".join(missing_target[:20]))

    # Build the execution plan.
    # Legacy behavior: --num-outputs means "per input image".
    # UI behavior: --total-outputs means "total images for the whole batch".
    plan = []
    if args.total_outputs is not None:
        total = max(1, int(args.total_outputs))
        full_rounds, remainder = divmod(total, max(1, len(jobs)))
        for idx, (img, mask, target_area) in enumerate(jobs):
            count = full_rounds + (1 if idx < remainder else 0)
            if count > 0:
                plan.append((img, mask, target_area, count))
    else:
        plan = [(img, mask, target_area, max(1, int(args.num_outputs))) for img, mask, target_area in jobs]

    emitted = 0
    batch_run_name = make_batch_run_name(args.run_name)
    output_root = args.output_dir or root / "runs"
    # New run layout (no <class_name> folder layer). runs/<run_name>/ holds:
    #   Gen_Images/                        all generated images for this run
    #   generation_summary.xlsx (or .csv)  per-image original/generated name, tokens, cost
    #   prompt.txt                         the prompt actually sent to the API
    batch_output_root = output_root / batch_run_name
    gen_images_dir = batch_output_root / "Gen_Images"
    gen_images_dir.mkdir(parents=True, exist_ok=True)

    # Clean any leftover working folders from an interrupted earlier run.
    for stray in batch_output_root.glob("work_*"):
        if stray.is_dir():
            shutil.rmtree(stray, ignore_errors=True)

    # Expand the execution plan into a deterministic flat queue, so a resumed run
    # can continue purely by counting images already present in Gen_Images.
    queue: list = []
    for img, mask, target_area, output_count in plan:
        for _ in range(max(1, int(output_count))):
            queue.append((img, mask, target_area))
    total_jobs = len(queue)

    # Reuse the fixed YYYY-MM-DD-HH-mm label and continue the counter from any
    # outputs that already exist (resume); otherwise start fresh at counter 1.
    name_re = re.compile(r"^" + re.escape(batch_run_name) + r"-(\d{4})-(\d{2})-(\d{2})-(\d{2})-(\d{2})-(\d+)$")
    existing_imgs = [p for p in list_images(gen_images_dir) if name_re.match(p.stem)]
    if existing_imgs:
        existing_imgs.sort(key=lambda p: int(name_re.match(p.stem).group(6)))
        last = name_re.match(existing_imgs[-1].stem)
        time_label = f"{last.group(1)}-{last.group(2)}-{last.group(3)}-{last.group(4)}-{last.group(5)}"
        already_done = len(existing_imgs)
    else:
        now = datetime.now()
        time_label = f"{now.year:04d}-{now.month:02d}-{now.day:02d}-{now.hour:02d}-{now.minute:02d}"
        already_done = 0

    # Persist the prompt that is actually sent to the API.
    try:
        resolved_prompt = load_user_prompt(defect_type, args.prompt_file, args.prompt, args.size).get("user_prompt", "")
    except Exception:
        resolved_prompt = args.prompt or ""
    try:
        (batch_output_root / "prompt.txt").write_text(resolved_prompt, encoding="utf-8")
    except Exception as exc:
        print(f"[WARN] could not write prompt.txt: {exc}", flush=True)

    start_index = already_done if args.resume_existing else 0
    summary_rows: list = []
    for index in range(start_index, total_jobs):
        img, mask, target_area = queue[index]
        counter = index + 1
        seed = args.seed + index
        new_name = sanitize_path_component(f"{batch_run_name}-{time_label}-{counter}")
        final_image = gen_images_dir / f"{new_name}.{args.output_format}"
        if args.resume_existing and final_image.exists():
            print(f"[SKIP] existing output: {final_image.name}", flush=True)
            continue
        work_dir = batch_output_root / f"work_{counter}"
        if work_dir.exists():
            shutil.rmtree(work_dir, ignore_errors=True)
        log_path = work_dir / "log.txt"
        size_is_original = str(args.size).strip().lower() in {"same_as_original", "original", "source"}

        cmd = [
            sys.executable, "-u", str(root / "scripts" / "run_gpt_image2.py"),
            "--class-name", defect_type,
            "--image", str(img),
            "--workflow", args.workflow,
            "--placement-mode", args.placement_mode,
            "--seed", str(seed),
            "--num-outputs", "1",
            "--run-name", work_dir.name,
            "--output-name", new_name,
            "--batch-run-name", batch_run_name,
            "--output-dir", str(batch_output_root),
            "--model", args.model,
            "--size", resolve_size_for_image(args.size, img),
            "--quality", args.quality,
            "--output-format", args.output_format,
            "--background", args.background,
            "--mask-dilate", str(args.mask_dilate),
            "--mask-feather", str(args.mask_feather),
            "--auto-resize-multiple", "1" if size_is_original else str(args.auto_resize_multiple),
            *original_size_args_if_needed(args.size, img),
            "--min-defects", str(args.min_defects),
            "--max-defects", str(args.max_defects),
            "--random-scale-min", str(args.random_scale_min),
            "--random-scale-max", str(args.random_scale_max),
            "--placement-attempts", str(args.placement_attempts),
            "--exclude-repair-padding", str(args.exclude_repair_padding),
        ]
        if mask is not None:
            cmd += ["--mask", str(mask)]
        if target_area is not None:
            cmd += ["--target-area", str(target_area)]
        if args.output_compression is not None:
            cmd += ["--output-compression", str(args.output_compression)]
        if args.no_clip_mask_to_target:
            cmd.append("--no-clip-mask-to-target")
        if args.prompt.strip():
            cmd += ["--prompt", args.prompt]
        if args.prompt_file:
            cmd += ["--prompt-file", str(args.prompt_file)]
        if args.prompt_config:
            cmd += ["--prompt-config", str(args.prompt_config)]
        if args.prompt_extra.strip():
            cmd += ["--prompt-extra", args.prompt_extra]
        if args.dry_run:
            cmd.append("--dry-run")

        run_and_tee(cmd, log_path)

        # Move the final image into Gen_Images, capture token/cost, drop the work folder.
        tokens, cost = read_usage_from_metadata(work_dir / "metadata.json")
        produced = find_final_image(work_dir, new_name, args.output_format)
        if produced is not None and produced.exists():
            if final_image.exists():
                final_image.unlink()
            shutil.move(str(produced), str(final_image))
            summary_rows.append((img.name, final_image.name, tokens, cost))
            print(f"[OK] saved: {final_image}", flush=True)
        else:
            print(f"[WARN] no final image produced for counter {counter}", flush=True)
        shutil.rmtree(work_dir, ignore_errors=True)
        emitted += 1

    # Drop any leftover work folders and (re)write the per-image summary table.
    for stray in batch_output_root.glob("work_*"):
        if stray.is_dir():
            shutil.rmtree(stray, ignore_errors=True)
    by_name: dict = {}
    for row in read_existing_summary(batch_output_root):
        if len(row) >= 2 and row[1]:
            by_name[row[1]] = row
    for row in summary_rows:
        by_name[row[1]] = row

    def _row_counter(row: tuple) -> int:
        m = name_re.match(Path(str(row[1])).stem)
        return int(m.group(6)) if m else 0

    write_summary_table(batch_output_root, sorted(by_name.values(), key=_row_counter))


if __name__ == "__main__":
    main()
